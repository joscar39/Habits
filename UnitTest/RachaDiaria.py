import time
import unittest
import warnings
from calendar import monthrange

import allure
import numpy
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
from appium import webdriver
import openpyxl
import openpyxl as xl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import requests
from datetime import datetime
from datetime import timedelta

from PageObjectModel.FlowsCostant.LoginAndRegisterFlows import LoginAndRegisterFlowsUtils
from PageObjectModel.FlowsCostant.ModalsNpsQualificate import NpsAndRatingStore
from PageObjectModel.Global_variables.Const_General import ConstGeneral
from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.action_app.RequestUtils import RequestMethod
from PageObjectModel.action_app.Scroll_util import ScrollUtil
from PageObjectModel.action_app.Tap_Drag_Util import TapAndDragUtil

ngrok_ruta = "https://ecf7-45-236-30-219.ngrok.io"


class Racha(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            # "platformName": "Android",
            # "appium:platformVersion": "10",
            # "appium:deviceName": "dandelion_global ",
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_HAB01(self):  # Marcar Habito diaro por 28 dias y confirmar puntuacion
        suma = None
        driver = self.driver
        driver.implicitly_wait(20)
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RachaDiaria.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        # Verificar mes y fecha actual para obtener total de dias del mes en curso
        now = datetime.now()
        num_days = monthrange(now.year, now.month)[1]
        print(f"La fecha actual es: {now.month}/{now.year}, por lo tanto este mes posee {num_days} días")

        # Ejecutar endpoint para dejar al usaurio en el dia 1 de la racha diaria
        # Obtener token
        headers = None
        headers = RequestMethod.header_auth()

        # Obtener id del usuario
        id_user = RequestMethod.user_id(email.value, headers)[0]
        # Procesar racha diaria PAra que el usuario inicie desde cero
        RequestMethod.Streak_Daily(id_user, 0, headers)
        # Iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)
        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {name.value}")
        #
        # Verificar si el boton Encuesta care esta visible
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Ver más')]").click()

        time.sleep(1)

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\MarcarHabito"
                                      "\\HAB01_Puntos antes de Marcar habito.png")


        # Iterar las veces que se marcara el reto diario

        for i in range(1, num_days + 1):
            if i <= num_days:
                # ScrollUtil.swipe_up_short(1, driver)
                valuepoint = ConstGeneral.ScoreDashboard(driver)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\MarcarHabito"
                    f"\\HAB01_Dia {i} de racha .png")
                time.sleep(1)
                # Hacer click en si lo hice
                driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Si, lo cumplí']").click()
                WebDriverWait(driver, 90). \
                    until(
                    EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='Continuar']")))
                time.sleep(1)
                pto_racha = None
                pto_racha = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.widget.TextView[2]").text
                pto_dia_racha = None
                pto_dia_racha = pto_racha.split()[0]
                suma = int(valuepoint) + int(pto_dia_racha)
                point_succes = False
                # Verificar si se otorgan los puntos acordes al esquema de puntos por dia

                if i == 5:
                    if int(pto_dia_racha) == 20:
                        point_succes = True
                elif i == 10:
                    if int(pto_dia_racha) == 30:
                        point_succes = True
                elif i == 15:
                    if int(pto_dia_racha) == 40:
                        point_succes = True
                elif i == 20:
                    if int(pto_dia_racha) == 50:
                        point_succes = True
                elif i == 25:
                    if int(pto_dia_racha) == 60:
                        point_succes = True
                elif i == 30:
                    if int(pto_dia_racha) == 70:
                        point_succes = True
                elif i >= 31:
                    if int(pto_dia_racha) == 40:
                        point_succes = True
                elif [i <= 4] or [(i >= 6) and (i <= 9)] or [(i >= 11) and (i <= 14)] or [
                    (i >= 16) and (i <= 19)] \
                        or [(i >= 21) and (i <= 24)] or [(i >= 26) and (i <= 29)]:
                    if int(pto_dia_racha) == (9 + i):
                        point_succes = True
                assert point_succes is True, f"Error Para el dia {i} No se otrogaron " \
                                             f"los puntos correctos en la animacion," \
                                             f" se obtuvieron solo {pto_dia_racha}"
                print(f"Se otorgaron los {pto_dia_racha}pts Correctamente en la animacion para el dia {i}")
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\MarcarHabito"
                    f"\\HAB01_Animacion De Racha Cumplida del Dia {i}.png")
                allure.attach(driver.get_screenshot_as_png(),
                              name=f"HAB01 HAB01_Animacion De Racha Cumplida del Dia {i}",
                              attachment_type=AttachmentType.PNG)
                # Ejecutar cambio de dia desde el endpoint
                # Cambiar datos del logpoints para que no se eliminen los puntos por concepto de duplicidad
                # Obtener id logpoint del último ingreso de puntos al score del usuario
                id_logpoint = None
                id_logpoint = RequestMethod.Get_id_UserLogPoint(id_user, headers)
                # Cambiar datos del description
                RequestMethod.Put_description_of_point(i, id_logpoint, headers)
                # Procesar racha diaria
                RequestMethod.Streak_Daily(id_user, i, headers)
                # Pulsar boton continuar en la animacion de racha diaria
                driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
                time.sleep(5)
                # Comparar si se sumo el valor de la racha al score}
                # ScrollUtil.swipe_down_large(1, driver)

                newpoint = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                value_new_pt = newpoint.text

                if int(value_new_pt) == int(suma):
                    print(f"Se sumo correctamente los puntos de racha diaria para el dia {i}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\MarcarHabito"
                        f"\\HAB01_Suma De Racha En Score Para El Dia {i}.png")
                else:
                    print("Error no se sumo correctamente la racha diaria")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\MarcarHabito"
                        f"\\HAB01_Error Suma De Racha En Score Para El Dia {i}.png")

        # # ENVIAR DATOS DEL USUARIO A LA BD DE USUARIOS LISTOS
        # wb = xl.load_workbook(filesheet)
        # wd = wb.active
        # ws = wb.worksheets[0]
        # receives = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        #
        # wb3 = xl.load_workbook(receives)
        # ws3 = wb3.active
        #
        # ws3.insert_rows(1)
        # for o in range(1, 2):
        #     for j in range(1, 11):
        #         c = ws.cell(row=o, column=j)
        #
        #         ws3.cell(row=o, column=j).value = c.value
        # wb3.save(str(receives))
        # # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        # wd.delete_rows(1)  # para la fila 1
        # wb.save(filesheet)
        # time.sleep(1)
        # driver.swipe(263, 1152, 263, 837)  # Scroll uo
        # time.sleep(1.5)
        print("Actualization de BD")

    def test_HAB02(self):
        # Realizar avance de 5 dias en racha y luego perder un dia de la racha para reiniciar la racha
        driver = self.driver
        driver.implicitly_wait(20)
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RachaDiaria.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        # Obtener token
        headers = RequestMethod.header_auth()
        # Obtener id del usuario
        id_user = RequestMethod.user_id(email.value, headers)[0]

        # Procesar racha diaria PAra que el usuario inicie desde cero
        RequestMethod.Streak_Daily(id_user, 0, headers)

        # Iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {name.value}")

        # Desplegar ver mas en racha
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Ver más')]").click()

        status_of_point = False
        for i in range(1, 7):
            # Guardar Score inicial del usuario
            valuepoint = ConstGeneral.ScoreDashboard(driver)
            time.sleep(1)
            driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Si, lo cumplí']").click()
            # Esperar animacion de racha diaria
            WebDriverWait(driver, 90).until(EC.presence_of_element_located
                                            ((By.XPATH, "//android.widget.TextView[@text='Continuar']")))
            pto_racha = None
            pto_racha = driver.find_element(By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.widget.TextView[2]").text
            pto_dia_racha = None
            pto_dia_racha = pto_racha.split()[0]
            suma = int(valuepoint) + int(pto_dia_racha)
            # Obtener id logpoint del último ingreso de puntos al score del usuario
            id_logpoint = None
            id_logpoint = RequestMethod.Get_id_UserLogPoint(id_user, headers)
            # Cambiar datos del description
            RequestMethod.Put_description_of_point(i, id_logpoint, headers)
            if (i == 5) and (pto_dia_racha == 20):
                status_of_point = True
                RequestMethod.Break_streak(id_user, 2, 5, headers)
            elif (i < 5) and [pto_dia_racha == (9 + i)]:
                status_of_point = True
                RequestMethod.Streak_Daily(id_user, i, headers)
            elif (i > 5) and (int(pto_dia_racha) == (9 + i)):
                status_of_point = True
                RequestMethod.Streak_Daily(id_user, 1, headers)
            else:
                status_of_point = False
            assert status_of_point is True, f"Error no se asignaron los puntos correctamente para el dia {i}"
            print(f"Puntuacion otorgada exitosamente para el dia {i}")
            # Pulsar boton continuar en la animacion de racha diaria
            driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
            time.sleep(5)
            final_score = None
            final_score = ConstGeneral.ScoreDashboard(driver)
            assert int(suma) == int(final_score), f"Error no se esta agregando la puntuacion correctamente {suma} al score en el dia {i}"
            print(f"Se sumo correctamente la puntuacion al score final del usuario para el dia {i}")
            if i == 5:
                elements = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                for ele in elements:
                    print(elements[elements.index(ele)].text, "--position--", elements.index(ele))

    def test_HAB03(self):
        # Cambiar de habito al transcurrir 28 dias
        # global habits_selected
        pass

    def test_HAB04(self):
        # Volver a realizar encuenta care despuesd e 28 dia
        # global habits_sele
        pass

    def test_HAB05(self):
        # Pasar 112 dias sin responder el Care
        # Abrir Aplicacion y configurar modo Desarrollo
        # global habits_sele_two
        pass

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(Racha)
    unittest.TextTestRunner(verbosity=2).run(suite)
