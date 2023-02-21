import time
import unittest
import warnings

import pytest
from selenium.webdriver.common.by import By
from appium import webdriver
import openpyxl
import openpyxl as xl
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.action_app.Scroll_util import ScrollUtil
from PageObjectModel.action_app.Tap_Drag_Util import TapAndDragUtil


class Registro(unittest.TestCase):

    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_RE01(self):  # HACER UN REGISTRO CON LICENCIA YA UTILIZADA POR OTRO USUARIO
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        print("Activado modo desarrollo")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Registro\\RE01_ActivadoModoOculto.png")
        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup")

        # Confirmar si la modal para modo desarrollo esta visible
        modalisvisible = modal.is_displayed()

        if modalisvisible is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
            time.sleep(3)
        else:
            print("No se desplego la modal")

        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView")

        modal2isvisible = modal2.is_displayed()
        # Seleccionar opcion Desarrollo
        if modal2isvisible is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(3)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE01_ActivadoMODODESARROLLO.png")

        # Registrar usuario

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget."
                                      "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                      ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup[3]").click()
        time.sleep(2)

        codigo = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                               "widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view."
                                               "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                               "android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android."
                                               "view.ViewGroup/android.widget.ScrollView/android.view."
                                               "ViewGroup/android.view.ViewGroup[1]/android.widget.EditText")

        codigo.send_keys(PathAndVariablesUtil.cod_licencia())
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Registro\\RE01_IngresarCodigo.png")
        time.sleep(2)

        # Ingresar codigo de licencia
        camp_licencia = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                      "android.widget.LinearLayout/android.widget"
                                                      ".FrameLayout/android.widget.LinearLayout/android"
                                                      ".widget.FrameLayout/android.widget.FrameLayout/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.widget.FrameLayout/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.FrameLayout"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.ScrollView"
                                                      "/android.view.ViewGroup/android.view.ViewGroup[3]"
                                                      "/android.widget.EditText")

        # Cargar licencia Ya utilizada en registro anterior para esta compañia

        camp_licencia.send_keys("LQnSJ")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      "screenshots\\Registro\\RE01_MensajeLicenciaUtilizada.png")
        driver.execute_script('mobile: longClickGesture', {'x': 621, 'y': 790, 'duration': 50})

        time.sleep(1)
        # Enviar datos de codigo y licencia
        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget.Frame"
                                      "Layout/android.view.ViewGroup/android.view.ViewGroup/android.view."
                                      "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                      "widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.widget.ScrollView/android.view.ViewGroup/android"
                                      ".view.ViewGroup[4]/android.widget.TextView").click()

        time.sleep(1)

        # Mensaje de error indicando que la licencia ya ha sido utilizada pro otro usaurio

        error = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Esta licencia ya esta en uso')]").is_displayed()

        assert error is True, "No se mostro la alerta de licencia ya se encuentra en uso"
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      "screenshots\\Registro\\RE01_MensajeErrorLicenciaEnUso.png")
        print("Se mostro la alerta correctamente")

        cerraralerta = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
        cerraralerta.click()

    def test_RE02(self):  # HACER UN REGISTRO CON LICENCIA NO EXISTENTE
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        print("Activado modo desarrollo")
        time.sleep(2)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE02_ActivadoModoOculto.png")
        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup")

        # Confirmar si la modal para modo desarrollo esta visible
        modalisvisible = modal.is_displayed()

        if modalisvisible is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
            time.sleep(3)
        else:
            print("No se desplego la modal")

        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView")

        modal2isvisible = modal2.is_displayed()
        # Seleccionar opcion Desarrollo
        if modal2isvisible is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(3)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE02_EntrarModoDesarrollo.png")

        # Registrar usuario

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget."
                                      "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                      ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup[3]").click()

        codigo = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                               "widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view."
                                               "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                               "android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android."
                                               "view.ViewGroup/android.widget.ScrollView/android.view."
                                               "ViewGroup/android.view.ViewGroup[1]/android.widget.EditText")

        codigo.send_keys(PathAndVariablesUtil.cod_licencia())
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE02_IngresarCodigoCompañia.png")
        time.sleep(2)

        # Ingresar codigo de licencia
        camp_licencia = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                      "android.widget.LinearLayout/android.widget"
                                                      ".FrameLayout/android.widget.LinearLayout/android"
                                                      ".widget.FrameLayout/android.widget.FrameLayout/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.widget.FrameLayout/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.FrameLayout"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.ScrollView"
                                                      "/android.view.ViewGroup/android.view.ViewGroup[3]"
                                                      "/android.widget.EditText")

        # Cargar licencia inventada posteriormente mostrara error
        camp_licencia.send_keys("5Dwfs")
        time.sleep(2)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE02_Licencia_Inventada.png")
        TapAndDragUtil.tap_screen(621, 790, 50, driver)

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout"
                                      "/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                      "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.widget.ScrollView/android.view.ViewGroup").click()
        time.sleep(2)
        # Pulsar boton continuar y esperar a que muestre mensaje de error
        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget.Frame"
                                      "Layout/android.view.ViewGroup/android.view.ViewGroup/android.view."
                                      "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                      "widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.widget.ScrollView/android.view.ViewGroup/android"
                                      ".view.ViewGroup[4]/android.widget.TextView").click()

        time.sleep(2)

        error = driver.find_element(By.XPATH, "	//android.widget.TextView[contains(@text, 'Esta licencia no existe')]").is_displayed()

        assert error is True, "No se mostro la alerta de Licencia No existente"
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE02_MensajeErrorLicenciaNoExiste.png")
        print("Alerta mostrada correctamente")
        cerraralerta = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
        cerraralerta.click()

    def test_RE03(self):  # HACER UN REGISTRO EXITOSO

        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for a in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        print("Activado modo desarrollo")
        time.sleep(1)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE03_ActivadoModoOculto.png")
        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
            time.sleep(3)
        else:
            print("No se desplego la modal")

        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(3)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE03_ActivadoMODODESARROLLO.png")

        time.sleep(0.5)

        # Registrar usuario

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget."
                                      "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                      ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup[3]").click()

        codigo = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                               "widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view."
                                               "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                               "android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android."
                                               "view.ViewGroup/android.widget.ScrollView/android.view."
                                               "ViewGroup/android.view.ViewGroup[1]/android.widget.EditText")

        codigo.send_keys(PathAndVariablesUtil.cod_licencia())

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE03_IngresarCodigo.png")
        time.sleep(2)

        # Ingresar codigo de licencia
        camp_licencia = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                      "android.widget.LinearLayout/android.widget"
                                                      ".FrameLayout/android.widget.LinearLayout/android"
                                                      ".widget.FrameLayout/android.widget.FrameLayout/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.widget.FrameLayout/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.FrameLayout"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.ScrollView"
                                                      "/android.view.ViewGroup/android.view.ViewGroup[3]"
                                                      "/android.widget.EditText")

        # Cargar licencia de la empresa seleccionada a traves del archivo excel extraido del módulo licencias
        filesheet1 = f"{PathAndVariablesUtil.db_path()}/recursos/Licencias.xlsx"
        wb1 = openpyxl.load_workbook(filesheet1)
        datos1 = wb1["Sheet1"]
        licencia = datos1['C2']
        camp_licencia.send_keys(licencia.value)
        time.sleep(1)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE03_LicenciaAgregada.png")

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout"
                                      "/android.widget.FrameLayout/android.widget.LinearLayout"
                                      "/android.widget.FrameLayout/android.widget.FrameLayout"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android."
                                      "view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android"
                                      ".view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.widget.ScrollView/android.view.ViewGroup/android.widget."
                                      "TextView[1]").click()
        time.sleep(2)
        # Concretar proceso de ingresar codigo y licencia
        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget.Frame"
                                      "Layout/android.view.ViewGroup/android.view.ViewGroup/android.view."
                                      "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                      "widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.widget.ScrollView/android.view.ViewGroup/android"
                                      ".view.ViewGroup[4]/android.widget.TextView").click()

        # Llenado de formulario de registro
        filesheet2 = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb2 = openpyxl.load_workbook(filesheet2)
        datos2 = wb2["Hoja1"]
        name = datos2['A1']
        email = datos2['B1']
        passw = datos2['C1']
        gender = datos2['D1']
        area = datos2['H1']
        range_years = datos2['I1']

        # Ingresar Full name
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name.value)
        driver.hide_keyboard()
        time.sleep(2)
        # driver.keyevent(61) # hacer tab siguiente campo
        print("Nombre cargado")
        # Ingresar Correo

        campos[1].send_keys(email.value)
        driver.hide_keyboard()
        print("correo cargado")
        time.sleep(2)

        # Ingresar contraseña

        campos[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        # ingresar verificacion de contraseña

        campos[3].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")
        time.sleep(1)
        ScrollUtil.swipe_up_large(1, driver)

        # seleccionar género
        genero = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]")
        genero.click()
        genero.click()
        time.sleep(2)

        camp_gender = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        #  Seleccionar de la lista de genero
        for gene in camp_gender:
            if gender.value == camp_gender[camp_gender.index(gene)].text:
                camp_gender[camp_gender.index(gene)].click()
                time.sleep(1)
                break
        # Cerrar listado de generos
        for gen in camp_gender:
            if camp_gender[camp_gender.index(gen)].text == "Done":
                camp_gender[camp_gender.index(gen)].click()
                time.sleep(1)
                break

        print("genero cargado")

        # Seleccionar fecha de nacimiento
        time.sleep(2)
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth)+1)].click()
                time.sleep(1)
                break
        cerrar = driver.find_element(By.ID, "android:id/button1")
        cerrar.click()
        time.sleep(2)

        # Aceptar terminos y condiciones
        terms = None
        terms = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for ter in terms:
            if terms[terms.index(ter)].text == "Acepto los términos y condiciones":
                terms[(terms.index(ter) - 1)].click()
                time.sleep(1)
                break
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE03_FormularioLleno.png")

        # Concretar registro
        finalizar = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finalizar.click()

        exitoso = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, '¡Bienvenido de nuevo!')]")))

        # Verificacion de registro exitoso
        review_register = exitoso.is_displayed()
        assert review_register is True, "No se concreto el registro por lo tanto no redirecciono a la pantalla de login"
        print("Se concreto el registro exitosamente")

        iniciosesion = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")
        iniciosesion[1].send_keys(passw.value)
        acceder = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout"
                                                "/android.widget.FrameLayout/android.widget.LinearLayout/android"
                                                ".widget.FrameLayout/android.widget.FrameLayout/android.view."
                                                "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                "/android.view.ViewGroup/android.widget.FrameLayout/android"
                                                ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                                ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]"
                                                "/android.widget.ScrollView/android.view.ViewGroup/android."
                                                "view.ViewGroup[4]")

        acceder.click()

        time.sleep(6)
        element = driver.find_element(By.XPATH, "//android.view.ViewGroup[2]/android.view."
                                                "ViewGroup/android.widget.TextView[1]").is_displayed()

        assert element is True, "No se logro hacer login de confirmacion con el nuevo registro de usuario"
        # Verificacion de usuario registrado iniciando sesion

        wb1.close()
        wb2.close()

        # Eliminar datos del excel sobre la licencia utilizada en
        # el registro para descartar las que ya estan usadas
        wb1 = xl.load_workbook(filesheet1)
        ws = wb1.active
        ws.delete_rows(2)  # para la fila 2
        wb1.save(filesheet1)

        # Enviar Datos de registro al archivo par hacer proceso de login

        wb2 = xl.load_workbook(filesheet2)
        ws2 = wb2.worksheets[0]

        receives = f"{PathAndVariablesUtil.db_path()}/recursos/LoginECV.xlsx"
        wb3 = xl.load_workbook(receives)
        ws3 = wb3.active

        ws3.insert_rows(1)
        for o in range(1, 2):
            for j in range(1, 11):
                c = ws2.cell(row=o, column=j)

                ws3.cell(row=o, column=j).value = c.value

        wb3.save(str(receives))

        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd = wb2.active
        wd.delete_rows(1)  # para la fila 1
        wb2.save(filesheet2)
        print("Se ajusto la BD al confirmar que el registro fue exitoso")

    def test_RE04(self):  # REGISTRAR USU1ARIO CON UNA CUENTA DE CORREO EXISTENTE
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        print("Activado modo desarrollo")
        time.sleep(1)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Registro\\RE04_ActivadoModoOculto.png")
        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
            time.sleep(3)
        else:
            print("No se desplego la modal")

        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(3)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE04_ActivadoMODODESARROLLO.png")

        time.sleep(0.5)

        # Registrar usuario

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget."
                                      "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                      ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup[3]").click()

        codigo = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                               "widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view."
                                               "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                               "android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android."
                                               "view.ViewGroup/android.widget.ScrollView/android.view."
                                               "ViewGroup/android.view.ViewGroup[1]/android.widget.EditText")

        codigo.send_keys(PathAndVariablesUtil.cod_licencia())

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Registro\\RE04_IngresarCodigo.png")
        time.sleep(2)

        # Ingresar codigo de licencia
        camp_licencia = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                      "android.widget.LinearLayout/android.widget"
                                                      ".FrameLayout/android.widget.LinearLayout/android"
                                                      ".widget.FrameLayout/android.widget.FrameLayout/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.view.ViewGroup/android.view.ViewGroup/"
                                                      "android.widget.FrameLayout/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.FrameLayout"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android.widget.ScrollView"
                                                      "/android.view.ViewGroup/android.view.ViewGroup[3]"
                                                      "/android.widget.EditText")

        # Cargar licencia de la empresa seleccionada a traves del archivo excel extraido del módulo licencias
        filesheet1 = f"{PathAndVariablesUtil.db_path()}/recursos/Licencias.xlsx"
        wb1 = openpyxl.load_workbook(filesheet1)
        datos1 = wb1["Sheet1"]
        licencia = datos1['C2']
        camp_licencia.send_keys(licencia.value)
        time.sleep(1)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE04_LicenciaAgregada.png")

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout"
                                      "/android.widget.FrameLayout/android.widget.LinearLayout"
                                      "/android.widget.FrameLayout/android.widget.FrameLayout"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android."
                                      "view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android"
                                      ".view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.widget.ScrollView/android.view.ViewGroup/android.widget."
                                      "TextView[1]").click()
        time.sleep(1)
        # Concretar proceso de ingresar codigo y licencia
        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget."
                                      "LinearLayout/android.widget.FrameLayout/android.widget.Frame"
                                      "Layout/android.view.ViewGroup/android.view.ViewGroup/android.view."
                                      "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android"
                                      ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                      "widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.widget.ScrollView/android.view.ViewGroup/android"
                                      ".view.ViewGroup[4]/android.widget.TextView").click()

        # Llenado de formulario de registro
        filesheet2 = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb2 = openpyxl.load_workbook(filesheet2)
        datos2 = wb2["Hoja1"]
        name = datos2['A1']
        passw = datos2['C1']
        gender = datos2['D1']

        time.sleep(3)
        # Ingresar Full name
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name.value)
        driver.hide_keyboard()
        time.sleep(2)
        # driver.keyevent(61) # hacer tab siguiente campo
        print("Nombre cargado")
        # Ingresar Correo

        campos[1].send_keys("joscar@habits.ai")
        driver.hide_keyboard()
        print("correo cargado")
        time.sleep(2)

        # Ingresar contraseña

        campos[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)
        driver.swipe(303, 1290, 321, 371)

        # ingresar verificacion de contraseña

        campos[3].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")
        time.sleep(1)
        ScrollUtil.swipe_up_large(1, driver)

        # seleccionar género
        genero = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]")
        genero.click()
        genero.click()
        time.sleep(2)

        camp_gender = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        #  Seleccionar de la lista de genero
        for gene in camp_gender:
            if gender.value == camp_gender[camp_gender.index(gene)].text:
                camp_gender[camp_gender.index(gene)].click()
                time.sleep(1)
                break
        # Cerrar listado de generos
        for gen in camp_gender:
            if camp_gender[camp_gender.index(gen)].text == "Done":
                camp_gender[camp_gender.index(gen)].click()
                time.sleep(1)
                break

        print("genero cargado")

        # Seleccionar fecha de nacimiento
        time.sleep(2)
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        cerrar = driver.find_element(By.ID, "android:id/button1")
        cerrar.click()
        time.sleep(2)

        # Aceptar terminos y condiciones
        terminos = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for term in terminos:
            if terminos[terminos.index(term)].text == "Acepto los términos y condiciones":
                terminos[(terminos.index(term) - 1)].click()
                time.sleep(1)
                break
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE04_FormularioLleno.png")

        # Concretar registro
        finalizar = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finalizar.click()
        time.sleep(3)

        # Verificacion alerta de correo de usaurio ya registrado

        alerta = driver.find_element(By.XPATH, " //android.widget.TextView[contains(@text,'Correo de usuario  ya registrado')]").is_displayed()
        assert alerta is True, "No se mostro la alerta indicando que el correo ingresado ya esta en uso"
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE04_AlertaError.png")
        print("Se mostro la alerta correctamente")
        cerraralerta = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
        cerraralerta.click()

    def test_REG05(self):
        # LUEGO DE REALIZAR REGISTRO PROCEDER A RETROCEDER E INGRESAR A
        # FORMULARIO DE REGISTRO PARA REGISTRAR UN SEGUNDO USUARIO
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for a in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        print("Activado modo desarrollo")
        time.sleep(1)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      f"\\Registro\\RE05_ActivadoModoOculto.png")
        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
            time.sleep(3)
        else:
            print("No se desplego la modal")

        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(3)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                      "\\Registro\\RE05_ActivadoMODODESARROLLO.png")

        time.sleep(0.5)

        # Registrar usuario

        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget"
                                      ".LinearLayout/android.widget.FrameLayout/android.widget."
                                      "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                      ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                      "android.view.ViewGroup[3]").click()

        codigo = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                               "widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view."
                                               "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                               "android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android."
                                               "view.ViewGroup/android.widget.ScrollView/android.view."
                                               "ViewGroup/android.view.ViewGroup[1]/android.widget.EditText")

        codigo.send_keys(PathAndVariablesUtil.cod_habits())

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE05_IngresarCodigo.png")
        time.sleep(2)
        # Llenado de formulario de registro
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        gender = datos['D1']
        area = datos['H1']
        range_years = datos['I1']

        # Ingresar datos de Area y rango de edad
        are = "/hierarchy/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.widget" \
              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView" \
              "/android.view.ViewGroup/android.view.ViewGroup[4]"

        camp_area = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, str(are))))
        camp_area.click()
        time.sleep(2)
        select_area = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        # Escoger una opcion de la lista
        for sele in select_area:
            if area.value == select_area[select_area.index(sele)].text:
                select_area[select_area.index(sele)].click()
                time.sleep(1)
                break
        # Pulsar boton Done (Hecho)
        for sel in select_area:
            if select_area[select_area.index(sel)].text == "Done":
                select_area[select_area.index(sel)].click()
                time.sleep(1)
                break
        # Seleccionar un rango de edad
        camp_range_year = driver.find_element(By.XPATH,
                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[5]")
        camp_range_year.click()
        time.sleep(2)
        select_range = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        # Escoger un rango de la lista
        for rang in select_range:
            if range_years.value == select_range[select_range.index(rang)].text:
                print(select_range[select_range.index(rang)].text)
                select_range[select_range.index(rang)].click()
                time.sleep(1)
                break
        # Pulsar boton de listo
        done = driver.find_element(By.XPATH,
                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
        done.click()
        # Avanzar al siguiente formulario
        time.sleep(2)
        next_step = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[6]/android.widget.TextView")
        next_step.click()
        time.sleep(2)

        # Ingresar Full name
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name.value)
        driver.hide_keyboard()
        time.sleep(2)
        # driver.keyevent(61) # hacer tab siguiente campo
        print("Nombre cargado")
        # Ingresar Correo

        campos[1].send_keys(email.value)
        driver.hide_keyboard()
        print("correo cargado")
        time.sleep(2)

        # Ingresar contraseña

        campos[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        # ingresar verificacion de contraseña

        campos[3].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")
        time.sleep(1)
        ScrollUtil.swipe_up_large(1, driver)

        # seleccionar género
        genero = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]")
        genero.click()
        genero.click()
        time.sleep(2)

        camp_gender = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        #  Seleccionar de la lista de genero
        for gene in camp_gender:
            if gender.value == camp_gender[camp_gender.index(gene)].text:
                camp_gender[camp_gender.index(gene)].click()
                time.sleep(1)
                break
        # Cerrar listado de generos
        for gen in camp_gender:
            if camp_gender[camp_gender.index(gen)].text == "Done":
                camp_gender[camp_gender.index(gen)].click()
                time.sleep(1)
                break

        print("genero cargado")

        # Seleccionar fecha de nacimiento
        time.sleep(2)
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        cerrar = driver.find_element(By.ID, "android:id/button1")
        cerrar.click()
        time.sleep(2)

        # Aceptar terminos y condiciones
        terminos = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for term in terminos:
            if terminos[terminos.index(term)].text == "Acepto los términos y condiciones":
                terminos[(terminos.index(term) - 1)].click()
                time.sleep(1)
                break
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE05_FormularioLleno.png")

        # Concretar registro
        finalizar = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finalizar.click()

        exitoso = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.XPATH, "//android.widget.TextView[contains(@text, '¡Bienvenido de nuevo!')]")))

        # Verificacion de registro exitoso
        review_register = exitoso.is_displayed()
        assert review_register is True, "No se concreto el registro por lo tanto no redirecciono a la pantalla de login"

        # Verificacion de usuario registrado iniciando sesion
        wb.close()

        # Enviar Datos de registro al archivo par hacer proceso de login

        wb = xl.load_workbook(filesheet)
        ws = wb.worksheets[0]

        receives = f"{PathAndVariablesUtil.db_path()}/recursos/Login.xlsx"
        wb3 = xl.load_workbook(receives)
        ws3 = wb3.active

        ws3.insert_rows(1)
        for o in range(1, 2):
            for j in range(1, 11):
                c = ws.cell(row=o, column=j)

                ws3.cell(row=o, column=j).value = c.value

        wb3.save(str(receives))

        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd = wb.active
        wd.delete_rows(1)  # para la fila 1
        wb.save(filesheet)
        print("Se logro iniciar sesion exitosamente")

        driver.hide_keyboard()
        backtosignup = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                     "LinearLayout/android.widget.FrameLayout/android.widget"
                                                     ".LinearLayout/android.widget.FrameLayout/android.widget"
                                                     ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                     "/android.view.ViewGroup/android.view.ViewGroup/android"
                                                     ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                                     ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                     "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                     "view.ViewGroup/android.view.ViewGroup[1]/android.widget"
                                                     ".TextView")
        backtosignup.click()
        time.sleep(3)

        # HACER PROCESO DE REGISTRO POR SEGUNDA VEZ
        # ACCEDER A MODULO REGISTRO
        driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/"
                                      "android.widget.FrameLayout/android.widget.LinearLayout/android.widget"
                                      ".FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/"
                                      "android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                      "/android.widget.FrameLayout/android.view.ViewGroup/android.view."
                                      "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view."
                                      "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                      ".ViewGroup[3]").click()
        # INGRESAR CODIGO
        codigo = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                               "widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.LinearLayout/android.widget.FrameLayout/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view."
                                               "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                               "android.widget.FrameLayout/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android."
                                               "view.ViewGroup/android.widget.ScrollView/android.view."
                                               "ViewGroup/android.view.ViewGroup[1]/android.widget.EditText")

        codigo.send_keys(PathAndVariablesUtil.cod_habits())

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE05_IngresarCodigo.png")

        time.sleep(2)

        # Llenado de formulario de registro
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        gender = datos['D1']
        area = datos['H1']
        range_years = datos['I1']

        # Ingresar datos de Area y rango de edad
        are = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[4]"
        camp_area = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, str(are))))
        camp_area.click()
        time.sleep(2)
        select_area = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        # Escoger una opcion de la lista
        for sele in select_area:
            if area.value == select_area[select_area.index(sele)].text:
                select_area[select_area.index(sele)].click()
                time.sleep(1)
                break
        # Pulsar boton Done (Hecho)
        for sel in select_area:
            if select_area[select_area.index(sel)].text == "Done":
                select_area[select_area.index(sel)].click()
                time.sleep(1)
                break
        # Seleccionar un rango de edad
        camp_range_year = driver.find_element(By.XPATH,
                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[5]")
        camp_range_year.click()
        time.sleep(1)
        select_range = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        # Escorger un rango de la lista
        for rang in select_range:
            if range_years.value == select_range[select_range.index(rang)].text:
                select_range[select_range.index(rang)].click()
                time.sleep(1)
                break

        # Pulsar boton de listo
        done = driver.find_element(By.XPATH,
                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
        done.click()
        # Avanzar al siguiente formulario
        time.sleep(2)
        next_step = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[6]/android.widget.TextView")
        next_step.click()
        time.sleep(2)

        # Ingresar Full name
        campos_new = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos_new[0].send_keys(name.value)
        driver.hide_keyboard()
        time.sleep(2)
        # driver.keyevent(61) # hacer tab siguiente campo
        print("Nombre cargado")
        # Ingresar Correo

        campos_new[1].send_keys(email.value)
        driver.hide_keyboard()
        print("correo cargado")
        time.sleep(2)

        # Ingresar contraseña

        campos_new[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        # ingresar verificacion de contraseña
        campos_new[3].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")
        ScrollUtil.swipe_up_large(1, driver)

        time.sleep(2)
        # seleccionar género
        genero = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]")
        genero.click()
        genero.click()
        time.sleep(2)

        camp_gender_new = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        #  Seleccionar de la lista de genero
        for gene_new in camp_gender_new:
            if gender.value == camp_gender_new[camp_gender_new.index(gene_new)].text:
                camp_gender_new[camp_gender_new.index(gene_new)].click()
                time.sleep(1)
                break
        # Cerrar listado de generos
        search_button_done = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for search_done in search_button_done:
            if search_button_done[search_button_done.index(search_done)].text == "Done":
                search_button_done[search_button_done.index(search_done)].click()
                time.sleep(1)
                break

        print("genero cargado")

        # Seleccionar fecha de nacimiento
        time.sleep(2)
        birthday_new = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth_new in birthday_new:
            if birthday_new[birthday_new.index(birth_new)].text == "Fecha de nacimiento":
                birthday_new[(birthday_new.index(birth_new) + 1)].click()
                time.sleep(1)
                break
        cerrar_new = driver.find_element(By.ID, "android:id/button1")
        cerrar_new.click()
        time.sleep(2)

        # Aceptar terminos y condiciones
        terminos_new = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for term_new in terminos_new:
            if terminos_new[terminos_new.index(term_new)].text == "Acepto los términos y condiciones":
                terminos_new[(terminos_new.index(term_new) - 1)].click()
                time.sleep(1)
                break
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE05_FormularioLleno_user 2.png")

        # Concretar registro
        finalizar_new = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finalizar_new.click()

        exitoso_new = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.XPATH, "//android.widget.TextView[contains(@text, '¡Bienvenido de nuevo!')]")))

        # Verificacion de registro exitoso
        review_register_new = exitoso_new.is_displayed()
        assert review_register_new is True, "No se concreto el registro por lo tanto no redirecciono a la pantalla de login"

        # Verificacion de usuario registrado iniciando sesion

        wb.close()
        # Enviar Datos de registro al archivo par hacer proceso de login

        wb = xl.load_workbook(filesheet)
        ws2 = wb.worksheets[0]

        receives = f"{PathAndVariablesUtil.db_path()}/recursos/Login.xlsx"
        wb3 = xl.load_workbook(receives)
        ws3 = wb3.active

        ws3.insert_rows(1)
        for o in range(1, 2):
            for j in range(1, 11):
                c = ws2.cell(row=o, column=j)

                ws3.cell(row=o, column=j).value = c.value

        wb3.save(str(receives))

        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd = wb.active
        wd.delete_rows(1)  # para la fila 1
        wb.save(filesheet)
        print("Registro concretado")

    def teardown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(Registro)
    unittest.TextTestRunner(verbosity=2).run(suite)
