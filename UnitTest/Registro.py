import time
import unittest
import warnings

import allure
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
from appium import webdriver
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.FlowsCostant.Set_Dev_Mode import SetUtilDevMode
from PageObjectModel.action_app.Scroll_util import ScrollUtil
from PageObjectModel.action_app.Tap_Drag_Util import TapAndDragUtil


class RegistroNew(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_REG01(self, email="lenyn@habits.ai"):
        # En la pantalla de registro ingresar un correo ya en uso
        # Al ingresar el correo existente mostrará alerta indicando ue el correo está en uso por lo que debera
        # redireccionar al login con el correo precargado para solo ingresar contraseña
        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Regístrate aquí')]").click()
        time.sleep(1)
        input_mail = None
        input_mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        input_mail.send_keys(email)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        alert_mail_registered = None
        alert_mail_registered = driver.find_elements(By.XPATH, "//android.widget.TextView")
        status = False
        for alert in alert_mail_registered:
            if alert_mail_registered[alert_mail_registered.index(alert)].text == "Ya te encuentras registrado":
                status = True
                break
            else:
                status = False
        assert status is True, "Error no se mostro alerta de correo ya existe en BD"
        print("Alerta de correo ya registrado en BD se mostró exitosamente")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE01_Alerta de licencia no existe.png")
        allure.attach(driver.get_screenshot_as_png(), name="REG01 Alerta de licencia no existe",
                      attachment_type=AttachmentType.PNG)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        email_input = None
        email_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.EditText[contains(@text, 'lenyn@habits.ai')]")))
        assert email_input.is_displayed(), "Error el usuario registrado no tiene precargado el correo en el input correspondiente"
        print("Registro Exitoso")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE01 El correo se mostro correctamente en el login.png")
        allure.attach(driver.get_screenshot_as_png(), name="RE01 El correo se mostro correctamente en el login", attachment_type=AttachmentType.PNG)

    def test_REG02(self, email="pedropablo@pabi.ai"):  # Ingresar a registrarte aqui introduccir el codigo empresa e introducir licencia errada
        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Regístrate aquí')]").click()

        time.sleep(1)
        input_mail = None
        input_mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        input_mail.send_keys(email)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys(
            PathAndVariablesUtil.cod_licencia())
        input_licence = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                        ((By.XPATH,
                                                          "//android.widget.EditText[contains(@text, 'ABC123')]")))
        input_licence.send_keys("As520.")
        TapAndDragUtil.tap_screen(918, 535, 50, driver)
        time.sleep(1)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        alert_licence_fail = None
        alert_licence_fail = driver.find_elements(By.XPATH, "//android.widget.TextView")
        status = False
        for alert in alert_licence_fail:
            if alert_licence_fail[alert_licence_fail.index(alert)].text == "Esta licencia no existe":
                status = True
                break
            else:
                status = False
        assert status is True, "La alerta de licencia no existe no se mostró"
        print("Alerta mostrada exitosamente, para cuando la licencia no existe")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE02_Alerta de licencia no existe.png")
        allure.attach(driver.get_screenshot_as_png(), name="REG02 Alerta de licencia no existe",
                      attachment_type=AttachmentType.PNG)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

    def test_REG03(self, email="fran@haba.ai"):  # Ingresar a registrate aqui e introducir un codigo de empresa incorrecto
        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Regístrate aquí')]").click()

        time.sleep(1)
        input_mail = None
        input_mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        input_mail.send_keys(email)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys("PAM-SUC")
        alert_cod_fail = None
        alert_cod_fail = driver.find_elements(By.XPATH, "//android.widget.TextView")
        status = False
        for alert in alert_cod_fail:
            if alert_cod_fail[alert_cod_fail.index(alert)].text == "Codigo no encontrado":
                status = True
                break
            else:
                status = False

        assert status is True, "La alerta de codigo no encontrado no se mostró"
        print("Alerta mostrada exitosamente, para cuando el codigo empresa no existe")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE03_Alerta de codigo de empresa no encontrado.png")
        allure.attach(driver.get_screenshot_as_png(), name="RE03 Alerta de codigo de empresa no encontrado",
                      attachment_type=AttachmentType.PNG)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

    def test_REG04(self):  # Registro correo inexistente para empresa con licencia
        # Ingresar a Registrate aqui, ingresar un correo que NO se encuentre registrado en la BD, debera
        # mostrar una alerta indicando que el correo NO esta registrado y si desea registrar dicho correo
        # redireccionara a la pantalla de registrarte aqui, con el correo precargado, se pulsa
        # continuar y debera cargar el codigo empresa y la licencia para luego avanzar al formulario de registro
        # sin que se muestre el campo de correo, al terminar debera llevar al
        # inicio con el correo precargado en el campo de correo

        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        # Ingresar a la seccion de registrate
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Regístrate aquí')]").click()
        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        gender = datos['D1']
        # Cargar correo a registrar para confirmar si existe en BD
        time.sleep(1)
        input_mail = None
        input_mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        input_mail.send_keys(email.value)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        # Ingresar codigo empresa y licencia para registrar usaurio nuevo
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys(
            PathAndVariablesUtil.cod_licencia())
        input_licence = None
        input_licence = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                        ((By.XPATH,
                                                          "//android.widget.EditText[contains(@text, 'ABC123')]")))
        filesheet2 = f"{PathAndVariablesUtil.db_path()}/recursos/Licencias.xlsx"
        wb2 = openpyxl.load_workbook(filesheet2)
        datos1 = wb2["Sheet1"]
        licencia = datos1['C2']
        input_licence.send_keys(licencia.value)
        TapAndDragUtil.tap_screen(918, 535, 50, driver)
        time.sleep(1)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                        "//android.widget.TextView[contains(@text, 'Completa tu información personal')]")))
        # llenado de formulario de registro
        campos = None
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name.value)
        print("Nombre cargado")

        campos[1].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        campos[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")

        ScrollUtil.swipe_up_large(1, driver)

        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Masculino')]").click()
        gender_select = None
        gender_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text, '{gender.value}')]")))
        gender_select.click()
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        print("genero seleccionado")
        time.sleep(2)
        birthday = None
        birth = None
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.ID, "android:id/button1").click()
        print("Fecha de nacimiento cargada")
        time.sleep(1)
        terms = None
        terms = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for ter in terms:
            if terms[terms.index(ter)].text == "Acepto los términos y condiciones":
                terms[(terms.index(ter) - 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        # registro satisfactorio y comprobación de correo cargado en input correspondiente
        registre_succes = None
        registre_succes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.EditText[contains(@text, '{email.value}')]")))
        assert registre_succes.is_displayed(), "Error el usuario registrado no tiene precargado el correo en el input correspondiente"
        print("Registro Exitoso")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE04 Registro exitoso.png")
        allure.attach(driver.get_screenshot_as_png(), name="RE04 Registro exitoso", attachment_type=AttachmentType.PNG)
        # Eliminar datos del excel sobre la licencia utilizada en
        # el registro para descartar las que ya estan usadas

        ws2 = wb2.active
        ws2.delete_rows(2)  # para la fila 2
        wb2.save(filesheet2)

        # Enviar Datos de registro al archivo par hacer proceso de login

        wb1 = openpyxl.load_workbook(filesheet)
        ws1 = wb1.worksheets[0]
        receives = f"{PathAndVariablesUtil.db_path()}/recursos/LoginECV.xlsx"
        wb3 = openpyxl.load_workbook(receives)
        ws3 = wb3.active

        for o in range(1, 2):
            ws3.insert_rows(1)
            for j in range(1, 11):
                c = ws1.cell(row=o, column=j)

                ws3.cell(row=o, column=j).value = c.value

        wb3.save(str(receives))

        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd = wb1.active
        wd.delete_rows(1)  # para la fila 1
        wb1.save(filesheet)
        print("Se ajusto la BD al confirmar que el registro fue exitoso")

    def test_REG05(self):  # Registro tradicional con correo inexistente en empresa sin licencia
        # Registrar de forma tradicional, ingresando al enlace Registrate aqui, posterioremente seleccionando un correo
        # Inexistente, luego debera seguir a la pantalla de Ingresar codigo empresa, una vez ingresado el codigo
        # Desplegara el formulario de registro, donde el correo debe estar precargado sin mostrarse en el formulario de registro
        # Al finalizar mostrara el correo en el campo correo de iniciar sesion y solo se debera anexar contraseña

        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        # Ingresar a la seccion de registrate
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Regístrate aquí')]").click()
        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        gender = datos['D1']
        area = datos['H1']
        range_years = datos['I1']

        # Cargar correo a registrar para confirmar si existe en BD
        time.sleep(1)
        input_mail = None
        input_mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        input_mail.send_keys(email.value)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        # Ingresar codigo empresa y licencia para registrar usaurio nuevo
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys(
            PathAndVariablesUtil.cod_habits())
        # Ingresar datos de Area y rango de edad
        are = None
        are = "/hierarchy/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.widget" \
              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView" \
              "/android.view.ViewGroup/android.view.ViewGroup[4]"
        camp_area = None
        camp_area = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, str(are))))
        camp_area.click()
        time.sleep(2)
        # Escoger una opcion de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{area.value}')]").click()
        time.sleep(1)
        # Pulsar boton Done (Hecho)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Seleccionar un rango de edad
        camp_range_year = None
        camp_range_year = driver.find_element(By.XPATH,
                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[5]")
        camp_range_year.click()
        time.sleep(2)

        # Escoger un rango de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{range_years.value}')]").click()
        time.sleep(1)
        # Pulsar boton de listo
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Ir al formulario de registro
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text, 'Completa tu información personal')]")))
        # Llenar formulario de registro
        campos = None
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name.value)
        print("Nombre cargado")

        campos[1].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        campos[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")

        ScrollUtil.swipe_up_large(1, driver)

        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Masculino')]").click()
        gender_select = None
        gender_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text, '{gender.value}')]")))
        gender_select.click()
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        print("genero seleccionado")
        time.sleep(2)
        birthday = None
        birth = None
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.ID, "android:id/button1").click()
        print("Fecha de nacimiento cargada")
        time.sleep(1)
        terms = None
        terms = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for ter in terms:
            if terms[terms.index(ter)].text == "Acepto los términos y condiciones":
                terms[(terms.index(ter) - 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        # registro satisfactorio y comprobación de correo cargado en input correspondiente
        registre_succes = None
        registre_succes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.EditText[contains(@text, '{email.value}')]")))
        assert registre_succes.is_displayed(), "Error el usuario registrado no tiene precargado el correo en el input correspondiente"
        print("Registro Exitoso")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE05 Registro exitoso.png")
        allure.attach(driver.get_screenshot_as_png(), name="RE05 Registro exitoso", attachment_type=AttachmentType.PNG)
        wb = openpyxl.load_workbook(filesheet)
        ws2 = wb.worksheets[0]
        receives = f"{PathAndVariablesUtil.db_path()}/recursos/Login.xlsx"
        wb3 = openpyxl.load_workbook(receives)
        ws3 = wb3.active

        for o in range(1, 2):
            ws3.insert_rows(1)
            for j in range(1, 11):
                c = ws2.cell(row=o, column=j)

                ws3.cell(row=o, column=j).value = c.value

        wb3.save(str(receives))

        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd = wb.active
        wd.delete_rows(1)  # para la fila 1
        wb.save(filesheet)
        wb.close()
        print("Registro concretado")

    def test_REG06(self):
        # Registrar dos usaurio en empresa sin licencia, uno a traves de acceso registrate aqui y otro a traves de correo inexistente
        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        # Ingresar a la seccion de registrate
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Regístrate aquí')]").click()
        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name_one = datos['A1']
        email_one = datos['B1']
        passw_one = datos['C1']
        gender_one = datos['D1']
        area_one = datos['H1']
        range_years_one = datos['I1']
        name_two = datos['A2']
        email_two = datos['B2']
        passw_two = datos['C2']
        gender_two = datos['D2']
        area_two = datos['H2']
        range_years_two = datos['I2']


        # Cargar correo a registrar para confirmar si existe en BD
        time.sleep(1)
        input_mail = None
        input_mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        input_mail.send_keys(email_one.value)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        # Ingresar codigo empresa y licencia para registrar usaurio nuevo
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys(
            PathAndVariablesUtil.cod_habits())
        # Ingresar datos de Area y rango de edad
        are = None
        are = "/hierarchy/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.widget" \
              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView" \
              "/android.view.ViewGroup/android.view.ViewGroup[4]"
        camp_area = None
        camp_area = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, str(are))))
        camp_area.click()
        time.sleep(2)
        # Escoger una opcion de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{area_one.value}')]").click()
        time.sleep(1)
        # Pulsar boton Done (Hecho)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Seleccionar un rango de edad
        camp_range_year = None
        camp_range_year = driver.find_element(By.XPATH,
                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[5]")
        camp_range_year.click()
        time.sleep(2)

        # Escoger un rango de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{range_years_two.value}')]").click()
        time.sleep(1)
        # Pulsar boton de listo
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Ir al formulario de registro
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text, 'Completa tu información personal')]")))
        # Llenar formulario de registro
        campos = None
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name_two.value)
        print("Nombre cargado")

        campos[1].send_keys(passw_two.value)
        driver.hide_keyboard()
        print("contraseña cargada")

        campos[2].send_keys(passw_two.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")

        ScrollUtil.swipe_up_large(1, driver)

        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Masculino')]").click()
        gender_select = None
        gender_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text, '{gender_two.value}')]")))
        gender_select.click()
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        print("genero seleccionado")
        time.sleep(1)
        birthday = None
        birth = 0
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.ID, "android:id/button1").click()
        print("Fecha de nacimiento cargada")
        time.sleep(1)
        terms = None
        ter = 0
        terms = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for ter in terms:
            if terms[terms.index(ter)].text == "Acepto los términos y condiciones":
                terms[(terms.index(ter) - 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        # registro satisfactorio y comprobación de correo cargado en input correspondiente
        registre_succes = None
        registre_succes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.EditText[contains(@text, '{email_one.value}')]")))
        assert registre_succes.is_displayed(), "Error el usuario registrado no tiene precargado el correo en el input correspondiente"
        print("Registro Exitoso")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE06 Primer Registro exitoso desde la secciona registrate aqui.png")
        allure.attach(driver.get_screenshot_as_png(), name="RE06 Primer Registro exitoso desde la secciona registrate aqui", attachment_type=AttachmentType.PNG)
        registre_succes.clear()
        time.sleep(1)
        registre_succes.send_keys(email_two.value)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        alert_email_nonexistent = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text, 'No te encuentras registrado')]"))).is_displayed()
        assert alert_email_nonexistent is True, "Error no se esta mostrando la alerta de correo inexistente"
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text,'Regístrate')]").click()
        secund_email = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, f"//android.widget.EditText[contains(@text, '{email_two.value}')]")))
        assert secund_email.is_displayed(), "Error el correo no se esta cargando en el campo email para registrar usaurio"
        print("Se mostro el correo cargado correctamente luego de mostrar alerta de usaurio inexistente")
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        # Ingresar codigo empresa y licencia para registrar usaurio nuevo
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys(
            PathAndVariablesUtil.cod_habits())
        # Ingresar datos de Area y rango de edad
        are = None
        are = "/hierarchy/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.widget" \
              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView" \
              "/android.view.ViewGroup/android.view.ViewGroup[4]"
        camp_area = None
        camp_area = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, str(are))))
        camp_area.click()
        time.sleep(2)
        # Escoger una opcion de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{area_two.value}')]").click()
        time.sleep(1)
        # Pulsar boton Done (Hecho)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Seleccionar un rango de edad
        camp_range_year = None
        camp_range_year = driver.find_element(By.XPATH,
                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[5]")
        camp_range_year.click()
        time.sleep(2)

        # Escoger un rango de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{range_years_one.value}')]").click()
        time.sleep(1)
        # Pulsar boton de listo
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Ir al formulario de registro
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text, 'Completa tu información personal')]")))
        # Llenar formulario de registro
        campos = None
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name_one.value)
        print("Nombre cargado")

        campos[1].send_keys(passw_one.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        campos[2].send_keys(passw_one.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")

        ScrollUtil.swipe_up_large(1, driver)

        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Masculino')]").click()
        gender_select = None
        gender_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//android.widget.TextView[contains(@text, '{gender_one.value}')]")))
        gender_select.click()
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        print("genero seleccionado")
        time.sleep(2)
        birthday = None
        birth = 0
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.ID, "android:id/button1").click()
        print("Fecha de nacimiento cargada")
        time.sleep(1)
        terms = None
        ter = 0
        terms = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for ter in terms:
            if terms[terms.index(ter)].text == "Acepto los términos y condiciones":
                terms[(terms.index(ter) - 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        # registro satisfactorio y comprobación de correo cargado en input correspondiente
        registre_succes = None
        registre_succes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//android.widget.EditText[contains(@text, '{email_two.value}')]")))
        assert registre_succes.is_displayed(), "Error el usuario registrado no tiene precargado el correo en el input correspondiente"
        print("Registro Exitoso")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE06 segundo Registro exitoso desde la seccion inicio.png")
        allure.attach(driver.get_screenshot_as_png(),
                      name="RE06 segundo Registro exitoso desde la seccion inicio",
                      attachment_type=AttachmentType.PNG)

        # Enviar Datos de registro al archivo par hacer proceso de login

        wb = openpyxl.load_workbook(filesheet)
        ws2 = wb.worksheets[0]
        receives = f"{PathAndVariablesUtil.db_path()}/recursos/Login.xlsx"
        wb3 = openpyxl.load_workbook(receives)
        ws3 = wb3.active
        for cicle in range(1, 3):
            ws3.insert_rows(1)
            for o in range(1, 2):
                for j in range(1, 11):
                    c = ws2.cell(row=o, column=j)

                    ws3.cell(row=o, column=j).value = c.value

            wb3.save(str(receives))

            # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
            wd = wb.active
            wd.delete_rows(1)  # para la fila 1
            wb.save(filesheet)
        wb.close()
        print("Registro concretado")

    def test_REG07(self):
        # Registro desde la pantalla de inicio al introducir un correo inexistente
        # Debera mostrarse una alerta indicando que el correo no existe por lo que si se pulsa registrar llevara
        # a la pantalla de registro con el correo precargado
        driver = self.driver
        SetUtilDevMode.SetDevelopMode(driver)
        # Ingresar a la seccion de registrate
        input_email_init = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/RegistroNuevo.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        gender = datos['D1']
        area = datos['H1']
        range_years = datos['I1']

        input_email_init.send_keys(email.value)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        alert_email_nonexistent = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            f"//android.widget.TextView[contains(@text, 'No te encuentras registrado')]"))).is_displayed()
        assert alert_email_nonexistent is True, "Error no se esta mostrando la alerta de correo inexistente"
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text,'Regístrate')]").click()
        check_email = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, f"//android.widget.EditText[contains(@text, '{email.value}')]")))
        assert check_email.is_displayed(), "Error el correo no se esta cargando en el campo email para registrar usaurio"
        print("Se mostro el correo cargado correctamente luego de mostrar alerta de usaurio inexistente")
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        time.sleep(1)
        # Ingresar codigo empresa y licencia para registrar usaurio nuevo
        driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'COD-123')]").send_keys(
            PathAndVariablesUtil.cod_habits())
        # Ingresar datos de Area y rango de edad
        are = None
        are = "/hierarchy/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "LinearLayout/android.widget.FrameLayout/android.widget." \
              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.widget" \
              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.widget.FrameLayout" \
              "/android.view.ViewGroup/android.view.ViewGroup/android.view." \
              "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
              ".view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView" \
              "/android.view.ViewGroup/android.view.ViewGroup[4]"
        camp_area = None
        camp_area = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, str(are))))
        camp_area.click()
        time.sleep(2)
        # Escoger una opcion de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{area.value}')]").click()
        time.sleep(1)
        # Pulsar boton Done (Hecho)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Seleccionar un rango de edad
        camp_range_year = None
        camp_range_year = driver.find_element(By.XPATH,
                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[5]")
        camp_range_year.click()
        time.sleep(2)

        # Escoger un rango de la lista
        driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{range_years.value}')]").click()
        time.sleep(1)
        # Pulsar boton de listo
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        time.sleep(1)
        # Ir al formulario de registro
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text, 'Completa tu información personal')]")))
        # Llenar formulario de registro
        campos = None
        campos = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")

        campos[0].send_keys(name.value)
        print("Nombre cargado")

        campos[1].send_keys(passw.value)
        driver.hide_keyboard()
        print("contraseña cargada")
        time.sleep(2)

        campos[2].send_keys(passw.value)
        driver.hide_keyboard()
        print("Confimacion de contraseña cargada")

        ScrollUtil.swipe_up_large(1, driver)

        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Masculino')]").click()
        gender_select = None
        gender_select = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//android.widget.TextView[contains(@text, '{gender.value}')]")))
        gender_select.click()
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Done')]").click()
        print("genero seleccionado")
        time.sleep(2)
        birthday = None
        birth = 0
        birthday = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for birth in birthday:
            if birthday[birthday.index(birth)].text == "Fecha de nacimiento":
                birthday[(birthday.index(birth) + 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.ID, "android:id/button1").click()
        print("Fecha de nacimiento cargada")
        time.sleep(1)
        terms = None
        ter = 0
        terms = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for ter in terms:
            if terms[terms.index(ter)].text == "Acepto los términos y condiciones":
                terms[(terms.index(ter) - 1)].click()
                time.sleep(1)
                break
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()
        # registro satisfactorio y comprobación de correo cargado en input correspondiente
        registre_succes = None
        registre_succes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//android.widget.EditText[contains(@text, '{email.value}')]")))
        assert registre_succes.is_displayed(), "Error el usuario registrado no tiene precargado el correo en el input correspondiente"
        print("Registro Exitoso")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Registro\\RE06 segundo Registro exitoso desde la seccion inicio.png")
        allure.attach(driver.get_screenshot_as_png(),
                      name="RE06 segundo Registro exitoso desde la seccion inicio",
                      attachment_type=AttachmentType.PNG)
        # Enviar Datos de registro al archivo par hacer proceso de login

        wb = openpyxl.load_workbook(filesheet)
        ws2 = wb.worksheets[0]
        receives = f"{PathAndVariablesUtil.db_path()}/recursos/Login.xlsx"
        wb3 = openpyxl.load_workbook(receives)
        ws3 = wb3.active

        for o in range(1, 2):
            ws3.insert_rows(1)
            for j in range(1, 11):
                c = ws2.cell(row=o, column=j)

                ws3.cell(row=o, column=j).value = c.value

        wb3.save(str(receives))

        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd = wb.active
        wd.delete_rows(1)  # para la fila 1
        wb.save(filesheet)
        wb.close()
        print("Usuario cambiado de BD")

    def teardown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(RegistroNew)
    unittest.TextTestRunner(verbosity=2).run(suite)
