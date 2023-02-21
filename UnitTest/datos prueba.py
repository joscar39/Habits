# Llenado de formulario de registro
import openpyxl

from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil

filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/Envia.xlsx"
wb = openpyxl.load_workbook(filesheet)
datos = wb["Hoja1"]
name = datos['A1']
email = datos['B1']
passw = datos['C1']
gender = datos['D1']
area = datos['H1']
range_years = datos['I1']

wb = openpyxl.load_workbook(filesheet)
ws2 = wb.worksheets[0]
receives = f"{PathAndVariablesUtil.db_path()}/recursos/Recibe.xlsx"
wb3 = openpyxl.load_workbook(receives)
ws3 = wb3.active
for cicle in range(1, 3):
    ws3.insert_rows(1)
    for o in range(1, 2):
        for j in range(1, 11):
            c = ws2.cell(row=o, column=j)

            ws3.cell(row=o, column=j).value = c.value

    wb3.save(str(receives))

    # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
    wd = wb.active
    wd.delete_rows(1)  # para la fila 1
    wb.save(filesheet)
wb.close()
print("Registro concretado")
