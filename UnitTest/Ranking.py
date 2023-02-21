import unittest
import time
import warnings

import openpyxl

from selenium.webdriver.common.by import By
from appium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from PageObjectModel.FlowsCostant.LoginAndRegisterFlows import LoginAndRegisterFlowsUtils
from PageObjectModel.FlowsCostant.ModalsNpsQualificate import NpsAndRatingStore
from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.action_app.RequestUtils import RequestMethod
from PageObjectModel.action_app.Scroll_util import ScrollUtil
from PageObjectModel.action_app.Tap_Drag_Util import TapAndDragUtil

ngrok_ruta = "https://4dd2-200-26-185-247.ngrok.io"


class MyTestChats(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_rank01(self):
        # Verificar posicion sin obtener puntos, luego realizar synchronization de datos para obtener puntos y verificar en ranking la posicion
        driver = self.driver
        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        secund_email = datos['B2']
        thirth_email = datos['B3']
        fourth_email = datos['B4']
        fifth_email = datos['B5']
        # obtains header request
        headers = RequestMethod.header_auth()
        # obtains data user
        data_user = RequestMethod.user_id(str(email.value), headers)
        id_user = data_user[0]
        name_user = data_user[1]
        company_user = data_user[2]
        # clean all ranking
        RequestMethod.delete_point_by_company(company_user, headers)
        # Add usar at ranking
        users = [secund_email.value, thirth_email.value, fourth_email.value, fifth_email.value]
        id_users = []
        name_users = []
        for i in range(0, 4):
            # obtains data user two
            data_users = None
            data_users = RequestMethod.user_id(str(users[i]), headers)
            id_users.append(data_users[0])
            name_users.append(data_users[1])
        # Add point for the secund user
        RequestMethod.masive_point_users(name_users[0], 10, id_users[0],
                                         name_users[1], 15, id_users[1],
                                         name_users[2], 20, id_users[2],
                                         name_users[3], 25, id_users[3],
                                         headers)

        # init sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element_find = None
        element_find = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Inicio']")

        assert element_find.is_displayed(), f"Error el usuario {name.value} no inicio sesion"
        print(f"Se inicio sesion correctamente con el usuario {name.value}")
        # Into Ranking and review if the user is positioned in the list
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
        time.sleep(2)
        # infography
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        time.sleep(1)
        # check if the user isn't in the list of ranking
        user_in_ranking = ScrollUtil.scroll_in_ranking(str(name.value), "4", driver)
        time.sleep(1)

        assert user_in_ranking is False, "Error Se esta mostrando puntos para este usuario en el ranking"
        print("El usuario no se encuentra en el ranking")
        time.sleep(1)
        # go back to home
        driver.back()
        time.sleep(1)
        # Add point for the user
        RequestMethod.point_user(1, name_user, id_user, headers)
        time.sleep(2)
        # Hide notification in app of the points
        ScrollUtil.close_notification(driver)
        # go back to the ranking, and check the user in the ranking with the posicion correctly
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))
        # Accept modal of infography
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        # Wait for the element is visible in the list of ranking
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='1']")))

        # find the user after obtains point and displayed in the ranking
        user_in_ranking = ScrollUtil.scroll_in_ranking(str(name.value), "Ver mas", driver)
        time.sleep(1)

        assert user_in_ranking is True, f"Error No Se esta mostrando en el ranking el usuario {name_user} luego de obtener puntos"
        print(f"Usuario {name_user} se muestra correctamente en el ranking")
        time.sleep(1)

    def test_rank02(self):
        # Comprobar que el icono para avance de posicion funciona
        # ingresando a ranking y visualizar si hay mas de 1 usuario en lista
        driver = self.driver

        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        secund_email = datos['B2']
        thirth_email = datos['B3']
        fourth_email = datos['B4']
        fifth_email = datos['B5']
        # obtains header request
        headers = RequestMethod.header_auth()
        # obtains data user
        data_user = RequestMethod.user_id(str(email.value), headers)
        id_user = data_user[0]
        name_user = data_user[1]
        # Iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element_find = None
        element_find = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Inicio']")

        assert element_find.is_displayed(), f"Error el usuario {name.value} no inicio sesion"
        print(f"Se inicio sesion correctamente con el usuario {name.value}")
        # Add point for the user
        RequestMethod.point_user(1, name_user, id_user, headers)
        time.sleep(1)
        # Hide notification in app of the points
        ScrollUtil.close_notification(driver)
        time.sleep(1)
        # go to the ranking, and check the user in the ranking with the posicion correctly in icon
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
        # infography
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='2']")))
        except:  # Add point a secund user for see him in the ranking

            print("No hay usuario en el ranking, se procedera a darle puntos a un grupo de usaurio")
            users = [secund_email.value, thirth_email.value, fourth_email.value, fifth_email.value]
            id_users = []
            name_users = []
            for i in range(0, 4):
                # obtains data user two
                data_users = None
                data_users = RequestMethod.user_id(str(users[i]), headers)
                id_users.append(data_users[0])
                name_users.append(data_users[1])
            # Add point for the secund user
            RequestMethod.masive_point_users(name_users[0], 10, id_users[0],
                                             name_users[1], 15, id_users[1],
                                             name_users[2], 20, id_users[2],
                                             name_users[3], 25, id_users[3],
                                             headers)
            time.sleep(2)
            driver.back()
            time.sleep(2)
            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
            # infography
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()

        # find  tab and swipe up for to see details of user and him point
        init_ele = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ver mas']")
        final_ele = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']")
        TapAndDragUtil.drag_drop_to_element(init_ele, final_ele, driver)
        time.sleep(2)

        elements = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        opponent = elements[45].text  # Obtain name of opponent
        position = int(elements[33].text)  # obtain My last position in ranking
        missing_point = int(elements[48].text) - int(elements[40].text)  # obtain difference of point between my opponent and me
        # go back to home
        driver.back()
        # Add point for the user
        RequestMethod.point_user((missing_point + 1), name_user, id_user, headers)
        time.sleep(2)
        # Hide notification in app of the points
        ScrollUtil.close_notification(driver)
        # go back to the ranking, and check the user is over of the another user
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
        # infography
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        # Wait the list is visible
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='1']")))

        # find  tab and swipe up for to see details of user and him point
        first = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ver mas']")
        secund = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']")
        TapAndDragUtil.drag_drop_to_element(first, secund, driver)
        time.sleep(2)
        see_position = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        result = False
        if (see_position[45].text != opponent) and (int(see_position[33].text) < position):
            result = True
        else:
            result = False
        assert result is True, f"Error El usuario {name_user} no logro avanzar de posicion"
        print(f"El usuario {name_user} Si logro avanzar de posicion, paso de la posicion {position} a la posicion {see_position[35].text}")

    def test_rank03(self):
        # Verificar que la obtención de puntos por avanzar de medalla funcione
        driver = self.driver

        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        secund_email = datos['B2']
        thirth_email = datos['B3']
        fourth_email = datos['B4']
        fifth_email = datos['B5']
        # obtains header request
        headers = RequestMethod.header_auth()
        # obtains data user
        data_user = RequestMethod.user_id(str(email.value), headers)
        id_user = data_user[0]
        name_user = data_user[1]
        # add point to user
        RequestMethod.point_user(1, name_user, id_user, headers)
        # Add user in ranking
        users = [secund_email.value, thirth_email.value, fourth_email.value, fifth_email.value]
        id_users = []
        name_users = []
        for i in range(0, 4):
            # obtains data user two
            data_users = None
            data_users = RequestMethod.user_id(str(users[i]), headers)
            id_users.append(data_users[0])
            name_users.append(data_users[1])
        # Add point for the secund user
        RequestMethod.masive_point_users(name_users[0], 10, id_users[0],
                                         name_users[1], 15, id_users[1],
                                         name_users[2], 20, id_users[2],
                                         name_users[3], 25, id_users[3],
                                         headers)
        # Point by medal
        point_medal = [20, 40, 80]
        level_medal = ["Habits Pro", "Habits Master", "Habits Hero"]

        # Iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element_find = None
        element_find = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Inicio']")

        assert element_find.is_displayed(), f"Error el usuario {name.value} no inicio sesion"
        print(f"Se inicio sesion correctamente con el usuario {name.value}")
        result_point = []
        result_level = []
        for i in range(0, 3):
            # Save point initial of user
            points = None
            valuepoint = None
            points = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            valuepoint = int(points.text)
            # Go into the ranking for see point to the next medal
            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
            # infography
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()

            # Search missing point to advance to the next
            time.sleep(2)
            search_missing_point = None
            text_with_missing_point = None
            separate_text = None
            point_for_the_next_level = None
            search_missing_point = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            text_with_missing_point = str(search_missing_point[4].text)
            separate_text = text_with_missing_point.split()
            point_for_the_next_level = int(separate_text[0])
            driver.back()
            # Add point for the user
            RequestMethod.point_user((point_for_the_next_level + 1), name_user, id_user, headers)
            time.sleep(1)
            # Hide notification in app of the points
            ScrollUtil.close_notification(driver)
            time.sleep(1)
            # Save new point after obtain point
            another_points = None
            new_value = None
            another_points = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            new_value = int(another_points.text)

            # Comparate point

            if int(valuepoint + (point_for_the_next_level + 1) + point_medal[i]) == new_value:
                result_point.append(True)
            else:
                result_point.append(False)

            assert result_point[-1] is True, f"Error los puntos esperados era {valuepoint + (point_for_the_next_level + 1) + point_medal[i]}, y se obtuvieron una cantidad de {new_value}"
            print(f"Se anexaron correctamente los {point_medal[i]} puntos al usuario {name_user} para cuando se avanza a la medalla {level_medal[i]}")
            time.sleep(1)
            # Go to the ranking and check the new medal
            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
            # infography
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
            time.sleep(1)
            find_level = None
            find_level = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            print(find_level[5].text)
            if i < 2:  # AL llegar a la ultima medalla el anunciado cambia y se pierde una posicion de ubicación del elemento de tipo de medalla
                if find_level[5].text == level_medal[i]:
                    result_level.append(True)
                else:
                    result_level.append(False)
            else:
                if find_level[4].text == level_medal[i]:
                    result_level.append(True)
                else:
                    result_level.append(False)

            assert result_level[-1] is True, f"Error el usuario {name_user}  No avanzó a la medalla estipulada: {level_medal[i]}"
            print(f"Se mostro la medalla correcta {level_medal[i]}, para el usuario {name_user}")
            driver.back()
            time.sleep(1)


    def test_rank04(self):
        # Comprobar que el icono para avance de posicion funciona cuando el usaurio sube de nivel o baja de nivel
        driver = self.driver

        # Datos para ingresar en formulario
        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        name = datos['A1']
        email = datos['B1']
        passw = datos['C1']
        secund_email = datos['B2']
        thirth_email = datos['B3']
        fourth_email = datos['B4']
        fifth_email = datos['B5']

        # obtains header request
        headers = RequestMethod.header_auth()
        # obtains data user
        data_user = RequestMethod.user_id(str(email.value), headers)
        id_user = data_user[0]
        name_user = data_user[1]
        # Iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element_find = None
        element_find = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Inicio']")

        assert element_find.is_displayed(), f"Error el usuario {name.value} no inicio sesion"
        print(f"Se inicio sesion correctamente con el usuario {name.value}")
        # Add point for the user
        RequestMethod.point_user(1, name_user, id_user, headers)
        time.sleep(1)
        # Hide notification in app of the points
        ScrollUtil.close_notification(driver)
        time.sleep(1)
        # go to the ranking, and check the user in the ranking with the posicion correctly in icon
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
        # infography
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='2']")))
        except:  # Add point a secund user for see him in the ranking

            print("No hay usuario en el ranking, se procedera a darle puntos a un grupo de usaurio")
            users = [secund_email.value, thirth_email.value, fourth_email.value, fifth_email.value]
            id_users = []
            name_users = []
            for i in range(0, 4):
                # obtains data user two
                data_users = None
                data_users = RequestMethod.user_id(str(users[i]), headers)
                id_users.append(data_users[0])
                name_users.append(data_users[1])
            # Add point for the secund user
            RequestMethod.masive_point_users(name_users[0], 10, id_users[0],
                                             name_users[1], 15, id_users[1],
                                             name_users[2], 20, id_users[2],
                                             name_users[3], 25, id_users[3],
                                             headers)
            time.sleep(2)
            driver.back()
            time.sleep(2)
            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
            # infography
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))

            driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()

        # find  tab and swipe up for to see details of user and him point
        init_ele = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ver mas']")
        final_ele = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']")
        TapAndDragUtil.drag_drop_to_element(init_ele, final_ele, driver)
        time.sleep(1)

        elements = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for e in elements:
            print(elements[elements.index(e)].text, "--position--", elements.index(e))

        # opponent = elements[45].text  # Obtain name of opponent
        # position = int(elements[33].text)  # obtain My last position in ranking
        # missing_point = int(elements[48].text) - int(elements[40].text)  # obtain difference of point between my opponent and me
        # # go back to home
        # driver.back()
        # # Add point for the user
        # RequestMethod.point_user((missing_point + 1), name_user, id_user, headers)
        # time.sleep(2)
        # # Hide notification in app of the points
        # ScrollUtil.close_notification(driver)
        # # go back to the ranking, and check the user is over of the another user
        # driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']").click()
        # # infography
        # WebDriverWait(driver, 10).until(
        #     EC.presence_of_element_located((By.XPATH,
        #                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[3]")))
        #
        # driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        # # Wait the list is visible
        # WebDriverWait(driver, 10).until(
        #     EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text='1']")))
        #
        # # find  tab and swipe up for to see details of user and him point
        # first = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ver mas']")
        # secund = driver.find_element(By.XPATH, "//android.widget.TextView[@text='Ranking']")
        # TapAndDragUtil.drag_drop_to_element(first, secund, driver)
        # time.sleep(2)
        # see_position = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        # result = False
        # if (see_position[45].text != opponent) and (int(see_position[33].text) < position):
        #     result = True
        # else:
        #     result = False
        # assert result is True, f"Error El usuario {name_user} no logro avanzar de posicion"
        # print(
        #     f"El usuario {name_user} Si logro avanzar de posicion, paso de la posicion {position} a la posicion {see_position[35].text}")


    def test_rank05(self):
        pass

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(MyTestChats)
    unittest.TextTestRunner(verbosity=2).run(suite)
