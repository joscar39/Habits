import time
import unittest
import warnings

import allure
import openpyxl
from allure_commons.types import AttachmentType
from appium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PageObjectModel.FlowsCostant.LoginAndRegisterFlows import LoginAndRegisterFlowsUtils
from PageObjectModel.FlowsCostant.ModalsNpsQualificate import NpsAndRatingStore
from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.Global_variables.Const_General import ConstGeneral
from PageObjectModel.action_app.Check_Elements import CheckElements
from PageObjectModel.action_app.Scroll_util import ScrollUtil


class MyTestBeneficios(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_BEN01(self, module="Gym QA", position_element=2, time_course=76):
        # Acceder a Beneficio Gym virtual (Cursos) y verificar que se muestre alertas de clase
        # inhabilitada por prerequisito, donde la alerta debe indicar que se necesita
        # ver el primer video para desbloquear el segundo y visualizarlo

        # position_element: Se refiere a la posicion en la cual se mantiene el beneficio en pantalla
        # module: Es el nombre del modulo que se desea acceder a traves de un

        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UsuariosBeneficios.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        name = datos['A1']

        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)
        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente, con el usuario {name.value}")
        # Obtain point of the score of the user
        score_user = None
        score_user = ConstGeneral.ScoreDashboard(driver)
        # Go to benefits module navbar
        driver.find_element(By.XPATH, PathAndVariablesUtil.benefits_navbar()).click()
        # call constants benefits of courses
        course = None
        course = ConstGeneral.Courses()
        time.sleep(1)
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Cursos']").click()
        time.sleep(1)
        p_element = f"/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout" \
                    f"/android.widget.FrameLayout/android.widget.LinearLayout/android.widget" \
                    f".FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android" \
                    f".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
                    f".widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/" \
                    f"android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup" \
                    f"/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup" \
                    f"/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup" \
                    f"/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout" \
                    f"/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/" \
                    f"android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup" \
                    f"/android.view.ViewGroup/android.view.ViewGroup[{position_element}]"
        driver.find_element(By.XPATH, p_element).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                        ((By.XPATH, f"//android.widget.TextView[contains(@text, '{course[1]}')]")))
        ScrollUtil.Scroll_without_limit(module, driver)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                        ((By.XPATH, f"//android.widget.TextView[contains(@text, '{module}')]")))
        class_courses = []
        find_class = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        # Search all class of the course
        for fc in find_class:
            if "Course" in find_class[find_class.index(fc)].text:
                class_courses.append(find_class[find_class.index(fc)].text)
        time.sleep(10)
        # Check alert when I click on the secund class
        driver.find_element(By.XPATH, f"//android.widget.TextView[@text='{class_courses[1]}']").click()
        time.sleep(1)
        alert_benefits = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        result = False
        for alert in alert_benefits:
            if alert_benefits[alert_benefits.index(
                    alert)].text == f"Este curso no estará disponible hasta que complete \"{class_courses[0]}\"":
                result = True
                break
            else:
                result = False
        assert result is True, "Error La alerta de curso inhabilitado por pre-requisito no se mostro"
        print("Se mostro correctamente la alerta de pre-requisito")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                      f"screenshots\\Beneficios\\"
                                      f"BEN01 Alerta de Pre-requisito se mostró correctamente.png")
        allure.attach(driver.get_screenshot_as_png(), name="BEN01 Alerta de Pre-requisito se mostró correctamente",
                      attachment_type=AttachmentType.PNG)
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Aceptar']").click()
        time.sleep(1)
        # See the first class and unlock the secund class
        driver.find_element(By.XPATH, f"//android.widget.TextView[@text='{class_courses[0]}']").click()
        time.sleep(1)
        play_course = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        play_course[6].click()
        time.sleep(1)
        timer = 0
        counter_video = None
        counter_video = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        while timer < time_course:
            element_counter = None
            element_counter = counter_video[3].text
            total_percentage = element_counter.split(sep="%")
            timer = int(total_percentage[0])
        driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Terminar']").click()
        print("Tiempo de visualizacion alcanzado para finalizar el curso")
        time.sleep(1)
        point = None
        point = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        text_point = str(point[4].text)
        save_point = text_point.split()
        the_point = int(save_point[0])
        range_point = driver.find_element(By.CLASS_NAME, "android.widget.SeekBar")
        range_point.send_keys(5)
        range_point.click()
        driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Enviar actividad']").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                        ((By.XPATH, f"//android.widget.TextView[contains(@text, 'Actividad completada')]")))
        ele = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        result_alert = False
        for e in ele:
            if ele[ele.index(e)].text == f"¡Bien hecho! Tu actividad ha sido aprobada y se te han otorgado {the_point} puntos. ":
                result_alert = True
                break
            else:
                result_alert = False

        assert result_alert is True, "No se mostro la alerta de actividad completada"
        print(f"Se mostro alerta de actividad completada con la obtención de {the_point} puntos")
        driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Aceptar']").click()
        # Check if the score is the correct then sum point of the course
        time.sleep(1)
        for i in range(0, 3):
            driver.back()
        time.sleep(1)
        new_score = None
        new_score = ConstGeneral.ScoreDashboard(driver)
        result_score = False
        if int(score_user) + int(the_point) == int(new_score):
            result_score = True
        else:
            result_score = False
        assert result_score is True, f"Error No se sumo la puntuacion del curso {class_courses[0]} al score del usuario"
        print(f"Se sumo correctamente la puntuacion del curso {class_courses[0]} al score")

        # # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        # wb = openpyxl.load_workbook(filesheet)
        # wd = wb.active
        # ws = wb.worksheets[0]
        #
        # receives = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        #
        # wb3 = openpyxl.load_workbook(receives)
        # ws3 = wb3.active
        #
        # ws3.insert_rows(1)
        # for i in range(1, 2):
        #     for j in range(1, 11):
        #         c = ws.cell(row=i, column=j)
        #
        #         ws3.cell(row=i, column=j).value = c.value
        #
        # wb3.save(str(receives))
        #
        # wd.delete_rows(1)  # para la fila 1
        # wb.save(filesheet)
        # wb.close()

    def test_BEN02(self, module="Clase para QA", position_element=3, time_class=80):
        # Acceder a Beneficios Meditation y yoga, entrar a clase grabad y omitir una clase 2 veces, para
        #  confirmar que la clase no se marque como completada, posteriormente visualizar la clase una tercera
        #  vez y asi corroborar que se marque como completada y se obtengan puntos

        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UsuariosBeneficios.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        name = datos['A1']

        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)
        # Verificar si mostrara modal de calificar aplicacion en tienda
        NpsAndRatingStore.SkipModalRating(email.value, driver)

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente, con el usuario {name.value}")
        score_user = None
        score_user = ConstGeneral.ScoreDashboard(driver)
        # Go to benefits module navbar
        driver.find_element(By.XPATH, PathAndVariablesUtil.benefits_navbar()).click()
        # call constants benefits of courses
        course = None
        course = ConstGeneral.LiveClassRecording()
        liveclassrecording = course[0]
        time.sleep(1)
        driver.find_element(By.XPATH, "//android.widget.TextView[@text='Clases grabadas']").click()
        time.sleep(1)
        p_element = None
        p_element = f"/hierarchy/android.widget.FrameLayout/android.widget." \
                    f"LinearLayout/android.widget.FrameLayout/android.widget." \
                    f"LinearLayout/android.widget.FrameLayout/android.widget." \
                    f"FrameLayout/android.view.ViewGroup/android.view.ViewGroup/" \
                    f"android.view.ViewGroup/android.view.ViewGroup/android.widget" \
                    f".FrameLayout/android.view.ViewGroup/android.view.ViewGroup" \
                    f"/android.view.ViewGroup/android.view.ViewGroup/android.view" \
                    f".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]" \
                    f"/android.view.ViewGroup/android.view.ViewGroup/android.widget" \
                    f".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android" \
                    f".view.ViewGroup[1]/android.widget.FrameLayout/android.view." \
                    f"ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android" \
                    f".view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup" \
                    f"/android.view.ViewGroup/android.view.ViewGroup[{position_element}]"
        driver.find_element(By.XPATH, p_element).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                        ((By.XPATH, f"//android.widget.TextView[contains(@text, '{liveclassrecording}')]")))
        ScrollUtil.Scroll_without_limit(module, driver)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                        ((By.XPATH, f"//android.widget.TextView[contains(@text, '{module}')]")))
        # See the first live class
        see_class = ConstGeneral.TitleLiveClassRecording()
        first_class = see_class[2]
        point_save = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        point_of_class = None
        value_text = None
        arr_text = None
        for p in point_save:
            if point_save[point_save.index(p)].text == first_class:
                value_text = point_save[point_save.index(p) + 3].text
                arr_text = value_text.split()
                point_of_class = arr_text[0]
                break
        status_first = None
        status_secund = None
        for enter in range(1, 3):
            driver.find_element(By.XPATH, f"//android.widget.TextView[@text= '{first_class}']").click()
            time.sleep(2)
            driver.back()
            if enter == 1:
                time.sleep(1)
                status_first = CheckElements.CheckIfTextElementIsVisible(f"¿Qué te pareció la clase de {first_class}", driver)
                assert status_first is True, "Error No se ha mostrado la primera alerta para comentar una clase grabada"
                print("Se mostro la alerta de comentario exitosamente")
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                              f"screenshots\\Beneficios\\"
                                              f"BEN02 Alerta para agregar comentario.png")
                allure.attach(driver.get_screenshot_as_png(),
                              name="BEN02 Alerta para agregar comentario", attachment_type=AttachmentType.PNG)
                driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Omitir']").click()
                time.sleep(1)
            elif enter > 1:
                time.sleep(1)
                status_secund = CheckElements.CheckIfTextElementIsVisible("Si sales ahora puede que no obtengas puntos.", driver)
                assert status_secund is True, "Erro no se mostro la segunda alerta paara cuando el usuairo sale de la clase grabada"
                print("Se mostro correctamente la segunda alerta al salir de la clase grabada")
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\"
                                              f"screenshots\\Beneficios\\"
                                              f"BEN02 Segunda alerta al salir de una clase grabada.png")
                allure.attach(driver.get_screenshot_as_png(),
                              name="BEN02 Segunda alerta al salir de una clase grabada", attachment_type=AttachmentType.PNG)
                driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Salir']").click()
                time.sleep(1)
        driver.find_element(By.XPATH, f"//android.widget.TextView[@text= '{first_class}']").click()
        time.sleep(1)
        timer = 0
        counter_video = None
        counter_video = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        while timer < time_class:
            element_counter = None
            element_counter = counter_video[2].text
            total_percentage = element_counter.split(sep="%")
            timer = int(total_percentage[0])
        driver.back()
        time.sleep(1)
        final_status = None
        final_status = CheckElements.CheckIfTextElementIsVisible(f"¿Qué te pareció la clase de {first_class}", driver)
        qualification = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        qualification[5].click()
        driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Enviar']").click()
        time.sleep(2)
        points_notification = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        result_point = False
        for po in points_notification:
            if points_notification[points_notification.index(po)].text == f"Hemos recibido tu valoración. Por responderla has ganado {point_of_class} puntos.":
                result_point = True
                break
            else:
                result_point = False
        assert result_point is True, "Error no se esta mostrando la alerta de obtención de puntos"
        print("Se mostró alerta de obtención de puntos correctamente")
        driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Aceptar']").click()
        time.sleep(2)
        for a in range(1, 4):
            driver.back()
        time.sleep(1)
        new_score = None
        new_score = ConstGeneral.ScoreDashboard(driver)
        result_score = False
        if int(score_user) + int(point_of_class) == int(new_score):
            result_score = True
        else:
            result_score = False
        assert result_score is True, f"Error No se sumo la puntuacion de la clase {first_class} al score del usuario"
        print(f"Se sumo correctamente la puntuacion de la clase {first_class} al score")

    def test_BEN03(self):
        # Entrar a Beneficios Premios, agregar puntos y ver si se reflejan en care coin
        # luego eliminar dichos puntos y ver que realmente ya no esten visibles en care coin
        pass

    def test_BEN04(self):
        # Entrar a beneficios premios, agregar puntos de care coin y comprar
        pass

        # # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        # wb = openpyxl.load_workbook(filesheet)
        # wd = wb.active
        # ws = wb.worksheets[0]
        #
        # receives = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        #
        # wb3 = openpyxl.load_workbook(receives)
        # ws3 = wb3.active
        #
        # ws3.insert_rows(1)
        # for i in range(1, 2):
        #     for j in range(1, 11):
        #         c = ws.cell(row=i, column=j)
        #
        #         ws3.cell(row=i, column=j).value = c.value
        #
        # wb3.save(str(receives))
        #
        # wd.delete_rows(1)  # para la fila 1
        # wb.save(filesheet)
        # wb.close()

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(MyTestBeneficios)
    unittest.TextTestRunner(verbosity=2).run(suite)
