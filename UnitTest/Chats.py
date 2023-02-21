import collections
import unittest
import time
import warnings

import numpy.random
import json
import requests
import openpyxl as xl
import openpyxl

from datetime import datetime
from datetime import timedelta

from appium.webdriver.appium_service import AppiumService
from selenium.webdriver.common.by import By
from appium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.action_app.Element_Constant import ElementConstantUtil
from PageObjectModel.action_app.Scroll_util import ScrollUtil
from PageObjectModel.action_app.Tap_Drag_Util import TapAndDragUtil

ngrok_ruta = "https://4dd2-200-26-185-247.ngrok.io"


class MyTestChats(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_chat(self):
        # Crear un grupo de chat público, y confirmar que se registren bien los miembros y los administrator seleccionados

        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\Chat01_ActivadoModoOculto.png")

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
        else:
            print("No se desplego la modal")
        time.sleep(2)
        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(2)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT01_ActivadoModoDesarrollo.png")

        time.sleep(0.5)

        # Iniciar Sesion

        login = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                              ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup[2]")
        login.click()

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        name = datos['A1']
        name1 = datos['A2']
        name2 = datos['A3']
        name3 = datos['A4']

        mail = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayou"
                                             "t/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                             "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGrou"
                                             "p/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                             "/android.widget.FrameLayout/android.view.ViewGroup/android.view.V"
                                             "iewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                             ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                             "view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup"
                                             "/android.view.ViewGroup[1]/android.widget.EditText")
        mail.send_keys(email.value)
        driver.hide_keyboard()
        time.sleep(1)

        password = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget.F"
                                                 "rameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "widget.FrameLayout/android.view.ViewGroup/android.view."
                                                 "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "view.ViewGroup/android.view.ViewGroup[2]/android.widget."
                                                 "ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                                 "[2]/android.widget.EditText")
        password.send_keys(passw.value)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\chats01_UsuarioCredenciales.png")

        access = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".FrameLayout/android.view.ViewGroup/android.view.ViewGrou"
                                               "p/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup[2]/android.widget"
                                               ".ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                               "[4]/android.widget.TextView")

        access.click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        enter_chats.click()
        time.sleep(2)
        new_chat = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")

        for new in new_chat:
            if new_chat[new_chat.index(new)].text == "Crear nuevo chat":
                new_chat[new_chat.index(new)].click()
                time.sleep(1)
                break
        # Formulario de nuevo chat
        time.sleep(1)
        input_img = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.ImageView")
        input_img.click()
        try:
            allow_bottom_permi = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,
                                                                                                "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")))
            allow_bottom_permi.click()
        except:
            print("Los permisos de acceso a multimedia ya fueron otorgados")

        select_type_file = driver.find_element(By.XPATH, "//android.widget.CompoundButton[contains(@text,'Imágenes')]")
        select_type_file.click()
        time.sleep(2)
        search_img = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for img in search_img:
            if ("IMG" or ".jpg") in search_img[search_img.index(img)].text:
                search_img[search_img.index(img)].click()
                time.sleep(1)
                break

        name_group = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[1]")))

        name_group.send_keys("Chat Grupal test 01")
        # Descripcion de grupo de chat
        description_group = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[2]")
        description_group.send_keys("Descripcion del grupo de chat01")
        # Tipo de grupo
        type_group_public = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Publico')]")
        type_group_public.click()
        # Enviar informacion para crear grupo de chat
        finish_new_group = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finish_new_group.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Agregar')]")))
        # Seleccionar un integrante de la lista filtrando por el buscador
        names = [
            name1.value,
            name2.value,
            name3.value
        ]
        for i in range(0, 3):
            search_input = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
            search_input.send_keys(names[i])
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text,'{names[i]}')]")))
            # ScrollUtil.scroll_to_text(name1, "Continuar", driver)
            find_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for find in find_user:
                if find_user[find_user.index(find)].text == names[i]:
                    find_user[(find_user.index(find) + 2)].click()
                    time.sleep(1)
                    break
        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)
        # Escoger Administradores
        admin_random = numpy.random.choice(names)
        select_admin = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for select in select_admin:
            if select_admin[select_admin.index(select)].text == admin_random:
                select_admin[(select_admin.index(select) + 2)].click()
                time.sleep(1)
                break
        TapAndDragUtil.search_buttom_contain_text("Crear mi equipo", driver)
        # Acceder al chat creado
        group_chat = WebDriverWait(driver, 60).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text,'Chat Grupal test 01')]")))
        group_chat.click()
        time.sleep(1)

        group_details = driver.find_element(By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView")
        group_details.click()
        time.sleep(1)
        # CConfirmation de integrantes y administrador

        ScrollUtil.swipe_up_large(1, driver)
        time.sleep(1)
        confirmation_members = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        list_members = []
        list_admin = [name.value]

        for confir in confirmation_members:
            if confirmation_members[confirmation_members.index(confir)].text in names:
                list_members.append(confirmation_members[confirmation_members.index(confir)].text)
                if (confirmation_members[(confirmation_members.index(confir) + 1)].text == "Admin") and (
                        confirmation_members[(confirmation_members.index(confir) + 1)].text != name.value):
                    list_admin.append(confirmation_members[confirmation_members.index(confir)].text)
                elif (confirmation_members[(confirmation_members.index(confir) + 1)].text == "Admin") and (
                        confirmation_members[(confirmation_members.index(confir) + 1)].text == name.value):
                    print(confirmation_members[confirmation_members.index(confir)].text, " Es el SuperdAdmin del grupo")
                else:
                    print(confirmation_members[confirmation_members.index(confir)].text, " No es un admin del grupo")

        assert collections.Counter(list_members) == collections.Counter(
            names), "Error los usuarios no son los seleccionados para el grupo de chat"
        print("Se asignaron los miembros correctos al grupo de chat")
        assert collections.Counter(list_admin) == collections.Counter(
            [name.value, admin_random]), " Error no son los administradores seleccionados"
        print("Se crearon los administradores correctamente")

    def test_chat02(self):
        # Crear un grupo de chat público, enviar mensaje, luego verificar que se muestre la notificacion badge
        # de mensaje nuevo con otro integrante del grupo de chat
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\Chats02_ActivadoModoOculto.png")

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
        else:
            print("No se desplego la modal")
        time.sleep(2)
        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(2)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT02_ActivadoModoDesarrollo.png")

        time.sleep(0.5)

        # Iniciar Sesion

        login = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                              ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup[2]")
        login.click()

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        secund_mail = datos['B3']
        secund_pass = datos['C3']
        name1 = datos['A5']
        name2 = datos['A3']
        name3 = datos['A6']

        mail = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayou"
                                             "t/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                             "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGrou"
                                             "p/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                             "/android.widget.FrameLayout/android.view.ViewGroup/android.view.V"
                                             "iewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                             ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                             "view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup"
                                             "/android.view.ViewGroup[1]/android.widget.EditText")
        mail.send_keys(email.value)
        driver.hide_keyboard()
        time.sleep(1)

        password = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget.F"
                                                 "rameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "widget.FrameLayout/android.view.ViewGroup/android.view."
                                                 "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "view.ViewGroup/android.view.ViewGroup[2]/android.widget."
                                                 "ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                                 "[2]/android.widget.EditText")
        password.send_keys(passw.value)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT02_UsuarioCredenciales.png")

        access = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".FrameLayout/android.view.ViewGroup/android.view.ViewGrou"
                                               "p/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup[2]/android.widget"
                                               ".ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                               "[4]/android.widget.TextView")

        access.click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        enter_chats.click()
        time.sleep(2)
        new_chat = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")

        for new in new_chat:
            if new_chat[new_chat.index(new)].text == "Crear nuevo chat":
                new_chat[new_chat.index(new)].click()
                time.sleep(1)
                break
        # Formulario de nuevo chat

        input_img = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.ImageView")
        input_img.click()
        try:
            allow_bottom_permi = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,
                                                                                                "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")))
            allow_bottom_permi.click()
        except:
            print("Los permisos de acceso a multimedia ya fueron otorgados")

        select_type_file = driver.find_element(By.XPATH, "//android.widget.CompoundButton[contains(@text,'Imágenes')]")
        select_type_file.click()
        time.sleep(1)
        search_img = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for img in search_img:
            if ("IMG" or ".jpg") in search_img[search_img.index(img)].text:
                search_img[search_img.index(img)].click()
                time.sleep(1)
                break

        name_group = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[1]")))

        name_group.send_keys("Test Chat 02")
        # Descripcion de grupo de chat
        description_group = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[2]")
        description_group.send_keys("Descripcion del grupo de chat01")
        # Tipo de grupo
        type_group_public = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Publico')]")
        type_group_public.click()
        # Enviar informacion para crear grupo de chat
        finish_new_group = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finish_new_group.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Agregar')]")))
        # Seleccionar un integrante de la lista filtrando por el buscador
        names = [
            name1.value,
            name2.value,
            name3.value
        ]
        for i in range(0, 3):
            search_input = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
            search_input.send_keys(names[i])
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text,'{names[i]}')]")))
            # ScrollUtil.scroll_to_text(name1, "Continuar", driver)
            find_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for find in find_user:
                if find_user[find_user.index(find)].text == names[i]:
                    find_user[(find_user.index(find) + 2)].click()
                    time.sleep(1)
                    break
        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)
        # Escoger Administradores
        admin_random = numpy.random.choice(names)
        select_admin = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for select in select_admin:
            if select_admin[select_admin.index(select)].text == admin_random:
                select_admin[(select_admin.index(select) + 2)].click()
                time.sleep(1)
                break
        TapAndDragUtil.search_buttom_contain_text("Crear mi equipo", driver)

        group_chat = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Test Chat 02')]")))
        group_chat.click()
        time.sleep(1)
        write_text = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Escribe un mensaje')]")
        write_text.send_keys("Bienvenidos al grupo test 02")

        send_text = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        send_text[-1].click()
        time.sleep(1)
        # Salir de la sala de chat
        send_text[0].click()
        time.sleep(2)
        # Salir de la app
        ScrollUtil.open_menu(driver)
        ScrollUtil.scroll_into_menu_to_text("Cerrar sesión", "Conectar", driver)

        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH, "//android.widget.TextView[contains(@text,"
                                                    "'¡Bienvenido! Inicia sesión o regístrate')]")))

        # Hacer Login con otro usuario perteneciente al equipo

        new_login = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]")
        new_login.click()
        time.sleep(2)
        try:
            toast_keychain = WebDriverWait(driver, 5). \
                until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.FrameLayout/android.widget"
                                                                ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup[2]/android.view.ViewGroup/"
                                                                "android.view.ViewGroup/android.view.ViewGroup")))
            verify_toast = toast_keychain.is_displayed()
            if verify_toast is True:
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\Chat02_Toast de keychain se muestra correctamente.png")
                time.sleep(1)
                click_to_toast = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                click_to_toast.click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\CHAT02_Se inicio automaticamente con el keychain.png")
                print("Se cumplio correctamente la funcion del keychain")
        except:
            print("Este dispositivo posee desbloqueo con sistema biometrico, se procedera a omitir el keychain")
            TapAndDragUtil.tap_screen(563, 1108, 50, driver)

        mail2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'ejemplo@mail.com')]")
        mail2.send_keys(secund_mail.value)

        passw2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        passw2.send_keys(secund_pass.value)

        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 10). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                   ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element2.is_displayed(), f"El usuario {name2.value} no inicio sesion se califica como error, ya que" \
                                        f" no se muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {name2.value}")
        module_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        module_chats.click()
        enter_group_chat = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Test Chat 02')]")))
        search_badge = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for badge in search_badge:
            # print(search_badge[search_badge.index(badge)].text, " posicion : ", search_badge.index(badge))
            if search_badge[search_badge.index(badge)].text == "Bienvenidos al grupo test 02":
                assert search_badge[
                           (search_badge.index(badge) + 1)].text == "1", "No Se mostro el badge de nuevo mensaje"
                print("Se visualiza el badge correctamente")
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\CHAT02_ Visualizacion de badge de mensaje nuevo.png")
                time.sleep(1)
                break

    def test_chat03(self):
        # Registrar un chat privado, invitar a usuarios y luego corroborar que el grupo de
        # chat no sea visualizados por integrantes no invitados
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\Chat03_ActivadoModoOculto.png")

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
        else:
            print("No se desplego la modal")
        time.sleep(2)
        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(2)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT03_ActivadoModoDesarrollo.png")

        time.sleep(0.5)

        # Iniciar Sesion

        login = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                              ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup[2]")
        login.click()

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        name1 = datos['A2']
        name2 = datos['A3']
        name3 = datos['A4']
        another_mail = datos['B5']
        another_pass = datos['C5']
        another_name = datos['A5']

        mail = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayou"
                                             "t/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                             "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGrou"
                                             "p/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                             "/android.widget.FrameLayout/android.view.ViewGroup/android.view.V"
                                             "iewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                             ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                             "view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup"
                                             "/android.view.ViewGroup[1]/android.widget.EditText")
        mail.send_keys(email.value)
        driver.hide_keyboard()
        time.sleep(1)

        password = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget.F"
                                                 "rameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "widget.FrameLayout/android.view.ViewGroup/android.view."
                                                 "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "view.ViewGroup/android.view.ViewGroup[2]/android.widget."
                                                 "ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                                 "[2]/android.widget.EditText")
        password.send_keys(passw.value)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT03_UsuarioCredenciales.png")

        access = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".FrameLayout/android.view.ViewGroup/android.view.ViewGrou"
                                               "p/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup[2]/android.widget"
                                               ".ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                               "[4]/android.widget.TextView")

        access.click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")

        enter_chats = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                      ((By.XPATH, PathAndVariablesUtil.chats_navbar())))
        enter_chats.click()
        time.sleep(2)
        new_chat = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")

        for new in new_chat:
            if new_chat[new_chat.index(new)].text == "Crear nuevo chat":
                new_chat[new_chat.index(new)].click()
                time.sleep(1)
                break
        # Formulario de nuevo chat

        input_img = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.ImageView")
        input_img.click()
        try:
            allow_bottom_permi = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,
                                                                                                "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")))
            allow_bottom_permi.click()
        except:
            print("Los permisos de acceso a multimedia ya fueron otorgados")

        select_type_file = driver.find_element(By.XPATH, "//android.widget.CompoundButton[contains(@text,'Imágenes')]")
        select_type_file.click()
        time.sleep(1)
        search_img = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for img in search_img:
            if ("IMG" or ".jpg") in search_img[search_img.index(img)].text:
                search_img[search_img.index(img)].click()
                time.sleep(1)
                break

        name_group = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[1]")))

        name_group.send_keys("Chat Privado Prueba 03")
        # Descripcion de grupo de chat
        description_group = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[2]")
        description_group.send_keys("Descripcion del grupo de chat01")
        # Tipo de grupo
        type_group_public = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Privado')]")
        type_group_public.click()
        # Enviar informacion para crear grupo de chat
        finish_new_group = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finish_new_group.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Agregar')]")))
        # Seleccionar un integrante de la lista filtrando por el buscador
        names = [
            name1.value,
            name2.value,
            name3.value
        ]
        for i in range(0, 3):
            search_input = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
            search_input.send_keys(names[i])
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text,'{names[i]}')]")))
            # ScrollUtil.scroll_to_text(name1, "Continuar", driver)
            find_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for find in find_user:
                if find_user[find_user.index(find)].text == names[i]:
                    find_user[(find_user.index(find) + 2)].click()
                    time.sleep(1)
                    break
        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)
        time.sleep(2)
        TapAndDragUtil.search_buttom_contain_text("Crear mi equipo", driver)
        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH,
                                          "//android.widget.TextView[contains(@text,'Chat Privado Prueba 03')]")))
        # Salir de la app
        ScrollUtil.open_menu(driver)
        ScrollUtil.scroll_into_menu_to_text("Cerrar sesión", "Conectar", driver)

        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH, "//android.widget.TextView[contains(@text,"
                                                    "'¡Bienvenido! Inicia sesión o regístrate')]")))

        # Hacer Login con otro usuario perteneciente al equipo

        new_login = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]")
        new_login.click()
        try:
            toast_keychain = WebDriverWait(driver, 5). \
                until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.FrameLayout/android.widget"
                                                                ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup[2]/android.view.ViewGroup/"
                                                                "android.view.ViewGroup/android.view.ViewGroup")))
            verify_toast = toast_keychain.is_displayed()
            if verify_toast is True:
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\Chat03_Toast de keychain se muestra correctamente.png")
                time.sleep(1)
                click_to_toast = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                click_to_toast.click()
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Login\\CHAT02_Se inicio automaticamente con el keychain.png")
                print("Se cumplio correctamente la funcion del keychain")
        except:
            print("Este dispositivo posee desbloqueo con sistema biometrico, se procedera a omitir el keychain")
            TapAndDragUtil.tap_screen(563, 1108, 50, driver)

        mail2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'ejemplo@mail.com')]")
        mail2.send_keys(another_mail.value)

        passw2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        passw2.send_keys(another_pass.value)

        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)

        element2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                   ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element2.is_displayed(), f"El usuario {another_name.value} no inicio sesion se califica como error, ya que" \
                                        f" no se muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {another_name.value}")
        module_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        module_chats.click()
        element_visible = False
        try:
            WebDriverWait(driver, 7).until(EC.presence_of_element_located(
                (By.XPATH, "//android.widget.TextView[contains(@text,'Chat Privado Prueba 03')]")))
            element_visible = True
        except:
            print("No esta visible el chat privado")

        assert element_visible is False, "Error se esta mostrando un chat privado para un usaurio no incluido en el grupo"
        print("Se cumplio el flujo correctamente, donde el chat privado no se mostro para el usuario loggeado")

    def test_chat04(self):

        # Crear un grupo privado, y editar integrantes se eliminando usuario, luego
        # se agregando otros y nombrando nuevos administradores y confirmar que los usuarios eliminados no puedan
        # visualizar el chat al que fueron eliminados

        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\Chat04_ActivadoModoOculto.png")

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
        else:
            print("No se desplego la modal")
        time.sleep(2)
        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(2)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT04_ActivadoModoDesarrollo.png")

        time.sleep(0.5)

        # Iniciar Sesion

        login = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                              ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup[2]")
        login.click()

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        name1 = datos['A2']
        name2 = datos['A3']
        name3 = datos['A4']
        email2 = datos['B2']
        pass2 = datos['C2']
        another_name = datos['A5']

        mail = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayou"
                                             "t/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                             "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGrou"
                                             "p/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                             "/android.widget.FrameLayout/android.view.ViewGroup/android.view.V"
                                             "iewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                             ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                             "view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup"
                                             "/android.view.ViewGroup[1]/android.widget.EditText")
        mail.send_keys(email.value)
        driver.hide_keyboard()
        time.sleep(1)

        password = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget.F"
                                                 "rameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "widget.FrameLayout/android.view.ViewGroup/android.view."
                                                 "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "view.ViewGroup/android.view.ViewGroup[2]/android.widget."
                                                 "ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                                 "[2]/android.widget.EditText")
        password.send_keys(passw.value)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT04_UsuarioCredenciales.png")

        access = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".FrameLayout/android.view.ViewGroup/android.view.ViewGrou"
                                               "p/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup[2]/android.widget"
                                               ".ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                               "[4]/android.widget.TextView")

        access.click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")

        enter_chats = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                      ((By.XPATH, PathAndVariablesUtil.chats_navbar())))
        enter_chats.click()
        time.sleep(2)
        new_chat = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")

        for new in new_chat:
            if new_chat[new_chat.index(new)].text == "Crear nuevo chat":
                new_chat[new_chat.index(new)].click()
                time.sleep(1)
                break
        # Formulario de nuevo chat

        input_img = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.ImageView")
        input_img.click()
        try:
            allow_bottom_permi = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,
                                                                                                "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")))
            allow_bottom_permi.click()
        except:
            print("Los permisos de acceso a multimedia ya fueron otorgados")

        select_type_file = driver.find_element(By.XPATH, "//android.widget.CompoundButton[contains(@text,'Imágenes')]")
        select_type_file.click()
        time.sleep(1)
        search_img = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for img in search_img:
            if ("IMG" or ".jpg") in search_img[search_img.index(img)].text:
                search_img[search_img.index(img)].click()
                time.sleep(1)
                break

        name_group = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[1]")))

        name_group.send_keys("Prueba de usuario eliminado")
        # Descripcion de grupo de chat
        description_group = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[2]")
        description_group.send_keys("Descripcion del grupo de chat01")
        # Tipo de grupo
        type_group_public = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Privado')]")
        type_group_public.click()
        # Enviar informacion para crear grupo de chat
        finish_new_group = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finish_new_group.click()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Agregar')]")))
        # Seleccionar un integrante de la lista filtrando por el buscador
        names = [
            name1.value,
            name2.value,
            name3.value
        ]
        for i in range(0, 3):
            search_input = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
            search_input.send_keys(names[i])
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text,'{names[i]}')]")))
            # ScrollUtil.scroll_to_text(name1, "Continuar", driver)
            find_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for find in find_user:
                if find_user[find_user.index(find)].text == names[i]:
                    find_user[(find_user.index(find) + 2)].click()
                    time.sleep(1)
                    break
        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)
        time.sleep(2)
        TapAndDragUtil.search_buttom_contain_text("Crear mi equipo", driver)
        # Acceder al chat creado
        group_chat = WebDriverWait(driver, 60).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text,'Prueba de usuario eliminado')]")))
        group_chat.click()
        time.sleep(1)

        group_details = driver.find_element(By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView")
        group_details.click()
        time.sleep(1)
        # Eliminar usuario del grupo
        delete_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for delete in delete_user:
            if delete_user[delete_user.index(delete)].text == f"{name2.value}":
                delete_user[(delete_user.index(delete) + 1)].click()
                time.sleep(1)
                break
        TapAndDragUtil.search_buttom_contain_text("Guardar Cambios", driver)
        # Acceder al chat creado
        group_chat_again = WebDriverWait(driver, 60).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text,'Prueba de usuario eliminado')]")))
        group_chat_again.click()
        time.sleep(1)

        group_details_again = driver.find_element(By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView")
        group_details_again.click()
        time.sleep(1)
        # Agregar nuevo usaurio al grupo
        add_new_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for add in add_new_user:
            if add_new_user[add_new_user.index(add)].text == "Añadir Participantes":
                add_new_user[(add_new_user.index(add) + 1)].click()
                variable = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText"
                search_new_user_input = WebDriverWait(driver, 60).until(EC.presence_of_element_located(
                    (By.XPATH, variable)))
                search_new_user_input.send_keys(another_name.value)
                break
        add_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for find in add_user:
            if add_user[add_user.index(find)].text == another_name.value:
                add_user[(add_user.index(find) + 2)].click()
                time.sleep(1)
                break
        TapAndDragUtil.search_buttom_contain_text("Actualizar", driver)
        time.sleep(2)
        # Salir de la app
        ScrollUtil.open_menu(driver)
        ScrollUtil.scroll_into_menu_to_text("Cerrar sesión", "Conectar", driver)

        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH, "//android.widget.TextView[contains(@text,"
                                                    "'¡Bienvenido! Inicia sesión o regístrate')]")))

        # Hacer Login con otro usuario perteneciente al equipo

        new_login = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]")
        new_login.click()
        try:
            toast_keychain = WebDriverWait(driver, 5). \
                until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.FrameLayout/android.widget"
                                                                ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup[2]/android.view.ViewGroup/"
                                                                "android.view.ViewGroup/android.view.ViewGroup")))
            verify_toast = toast_keychain.is_displayed()
            if verify_toast is True:
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\Chat04_Toast de keychain se muestra correctamente.png")
                time.sleep(1)
                click_to_toast = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                click_to_toast.click()
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Login\\CHAT04_Se inicio automaticamente con el keychain.png")
                print("Se cumplio correctamente la funcion del keychain")
        except:
            print("Este dispositivo posee desbloqueo con sistema biometrico, se procedera a omitir el keychain")
            TapAndDragUtil.tap_screen(563, 1108, 50, driver)

        mail2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'ejemplo@mail.com')]")
        mail2.send_keys(email2.value)

        passw2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        passw2.send_keys(pass2.value)

        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)

        element2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                   ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element2.is_displayed(), f"El usuario {name2.value} no inicio sesion se califica como error, ya que" \
                                        f" no se muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {name2.value}")
        module_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        module_chats.click()
        element_visible = False
        try:
            WebDriverWait(driver, 7).until(EC.presence_of_element_located(
                (By.XPATH, "//android.widget.TextView[contains(@text,'Prueba de usuario eliminad')]")))
            element_visible = True
        except:
            print("No esta visible el chat privado")

        assert element_visible is False, "Error se esta mostrando un chat privado para un usaurio no incluido en el grupo"
        print("Se cumplio el flujo correctamente, donde el chat privado no se mostro para el usuario loggeado")

    def test_chat05(self):

        # Registrar un chat publico, y una vez creado configurar para que el chat deje de ser publico a privado
        # y verificar que no se pueda visualizar para el resto de usuarios
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\Chats05_ActivadoModoOculto.png")

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
        else:
            print("No se desplego la modal")
        time.sleep(2)
        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(2)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT05_ActivadoModoDesarrollo.png")

        time.sleep(0.5)

        # Iniciar Sesion

        login = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                              ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup[2]")
        login.click()

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        admin_name = datos['A1']
        admin_email = datos['B1']
        admin_passw = datos['C1']
        secund_name = datos['A3']
        secund_mail = datos['B3']
        secund_pass = datos['C3']
        name1 = datos['A4']
        name2 = datos['A5']
        name3 = datos['A6']

        mail = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayou"
                                             "t/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                             "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGrou"
                                             "p/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                             "/android.widget.FrameLayout/android.view.ViewGroup/android.view.V"
                                             "iewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                             ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                             "view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup"
                                             "/android.view.ViewGroup[1]/android.widget.EditText")
        mail.send_keys(admin_email.value)
        driver.hide_keyboard()
        time.sleep(1)

        password = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget.F"
                                                 "rameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "widget.FrameLayout/android.view.ViewGroup/android.view."
                                                 "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "view.ViewGroup/android.view.ViewGroup[2]/android.widget."
                                                 "ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                                 "[2]/android.widget.EditText")
        password.send_keys(admin_passw.value)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT05_UsuarioCredenciales.png")

        access = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".FrameLayout/android.view.ViewGroup/android.view.ViewGrou"
                                               "p/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup[2]/android.widget"
                                               ".ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                               "[4]/android.widget.TextView")

        access.click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        enter_chats.click()
        time.sleep(2)
        new_chat = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")

        for new in new_chat:
            if new_chat[new_chat.index(new)].text == "Crear nuevo chat":
                new_chat[new_chat.index(new)].click()
                time.sleep(1)
                break
        # Formulario de nuevo chat

        input_img = driver.find_element(By.XPATH,
                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.ImageView")
        input_img.click()
        try:
            allow_bottom_permi = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,
                                                                                                "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")))
            allow_bottom_permi.click()
        except:
            print("Los permisos de acceso a multimedia ya fueron otorgados")

        select_type_file = driver.find_element(By.XPATH, "//android.widget.CompoundButton[contains(@text,'Imágenes')]")
        select_type_file.click()
        time.sleep(1)
        search_img = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
        for img in search_img:
            if ("IMG" or ".jpg") in search_img[search_img.index(img)].text:
                search_img[search_img.index(img)].click()
                time.sleep(1)
                break

        name_group = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[1]")))

        name_group.send_keys("private")
        # Descripcion de grupo de chat
        description_group = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.EditText[2]")
        description_group.send_keys("Descripcion del grupo de chat")
        # Tipo de grupo
        type_group_public = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Publico')]")
        type_group_public.click()
        # Enviar informacion para crear grupo de chat
        finish_new_group = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]")
        finish_new_group.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text,'Agregar')]")))
        # Seleccionar un integrante de la lista filtrando por el buscador
        names = [
            name1.value,
            name2.value,
            name3.value
        ]
        for i in range(0, 3):
            search_input = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
            search_input.send_keys(names[i])
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//android.widget.TextView[contains(@text,'{names[i]}')]")))
            # ScrollUtil.scroll_to_text(name1, "Continuar", driver)
            find_user = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for find in find_user:
                if find_user[find_user.index(find)].text == names[i]:
                    find_user[(find_user.index(find) + 2)].click()
                    time.sleep(1)
                    break
        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)
        time.sleep(1)
        TapAndDragUtil.search_buttom_contain_text("Crear mi equipo", driver)
        # Cerrar sesion e iniciar con usuario para corroborar que se muestra el chat grupal publico
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located(
                (By.XPATH, "//android.widget.TextView[contains(@text,'private')]")))

        # Salir de la app
        ScrollUtil.open_menu(driver)
        ScrollUtil.scroll_into_menu_to_text("Cerrar sesión", "Conectar", driver)

        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH, "//android.widget.TextView[contains(@text,"
                                                    "'¡Bienvenido! Inicia sesión o regístrate')]")))

        # Hacer Login con otro usuario perteneciente al equipo

        new_login = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]")
        new_login.click()
        time.sleep(2)
        try:
            toast_keychain = WebDriverWait(driver, 5). \
                until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.FrameLayout/android.widget"
                                                                ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup[2]/android.view.ViewGroup/"
                                                                "android.view.ViewGroup/android.view.ViewGroup")))
            verify_toast = toast_keychain.is_displayed()
            if verify_toast is True:
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\Chat05_Toast de keychain se muestra correctamente.png")
                time.sleep(1)
                click_to_toast = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                click_to_toast.click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\CHAT05_Se inicio automaticamente con el keychain.png")
                print("Se cumplio correctamente la funcion del keychain")
        except:
            print("Este dispositivo posee desbloqueo con sistema biometrico, se procedera a omitir el keychain")
            TapAndDragUtil.tap_screen(563, 1108, 50, driver)

        mail2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'ejemplo@mail.com')]")
        mail2.send_keys(secund_mail.value)

        passw2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        passw2.send_keys(secund_pass.value)

        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Continuar')]").click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 10). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                   ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element2.is_displayed(), f"El usuario {secund_name.value} no inicio sesion se califica como error, ya que" \
                                        f" no se muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {secund_name.value}")
        # Acceder al modulo chats
        driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar()).click()
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//android.widget.TextView[contains(@text,'private')]")))
            chat_is_visible = True
        except:
            chat_is_visible = False
        assert chat_is_visible is True, "Error No se esta mostrando el chat publico"
        print("Se ha visualizaco el chat publico correctamente")
        time.sleep(1)

        # Salir de la app para acceder como admin y cambiar el caht de publico a privado
        ScrollUtil.open_menu(driver)
        ScrollUtil.swipe_up_into_menu(2, driver)
        log_out = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Cerrar sesión')]")
        log_out.click()

        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH, "//android.widget.TextView[contains(@text,"
                                                    "'¡Bienvenido! Inicia sesión o regístrate')]")))

        # Hacer Login con otro usuario perteneciente al equipo

        new_login_again = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]")
        new_login_again.click()
        time.sleep(2)
        try:
            toast_keychain_again = WebDriverWait(driver, 5). \
                until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.FrameLayout/android.widget"
                                                                ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup[2]/android.view.ViewGroup/"
                                                                "android.view.ViewGroup/android.view.ViewGroup")))
            verify_toast_again = toast_keychain_again.is_displayed()
            if verify_toast_again is True:
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\Chat05_Toast de keychain se muestra correctamente.png")
                time.sleep(1)
                click_to_toast_again = driver.find_element(By.XPATH,
                                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                click_to_toast_again.click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\CHAT05_Se inicio automaticamente con el keychain.png")
                print("Se cumplio correctamente la funcion del keychain")
        except:
            print("Este dispositivo posee desbloqueo con sistema biometrico, se procedera a omitir el keychain")
            TapAndDragUtil.tap_screen(563, 1108, 50, driver)
        # Iniciar sesion como admin para acceder al cahat y cambiarlo de publico a privado
        mail_again = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'ejemplo@mail.com')]")
        mail_again.send_keys(admin_email.value)

        passw_again = driver.find_element(By.XPATH,
                                          "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        passw_again.send_keys(admin_passw.value)

        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 10). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        elemen = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                 ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert elemen.is_displayed(), f"El usuario {admin_name.value} no inicio sesion se califica como error, ya que" \
                                      f" no se muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {admin_name.value}")
        # Acceder al modulo chats
        driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar()).click()
        enter_group_chat_again = WebDriverWait(driver, 60).until(EC.presence_of_element_located(
            (By.XPATH, "//android.widget.TextView[contains(@text,'private')]")))
        enter_group_chat_again.click()
        time.sleep(1)
        group_details = driver.find_element(By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView")
        group_details.click()
        time.sleep(1)
        options = driver.find_elements(By.CLASS_NAME, "android.view.ViewGroup")
        options[1].click()
        time.sleep(1)
        do_it_private = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Hacer privado')]")
        do_it_private.click()
        time.sleep(2)
        # Cerrar sesion a iniciar con usuario para corroborar que no se muestre el chat privado
        ScrollUtil.open_menu(driver)
        ScrollUtil.swipe_up_into_menu(2, driver)
        driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Cerrar sesión')]").click()

        WebDriverWait(driver, 60).until(EC.presence_of_element_located
                                        ((By.XPATH, "//android.widget.TextView[contains(@text,"
                                                    "'¡Bienvenido! Inicia sesión o regístrate')]")))

        # Hacer Login con un usuario que no este en el equipo
        new_login2 = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]")
        new_login2.click()
        time.sleep(2)
        try:
            toast_keychain = WebDriverWait(driver, 5). \
                until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.FrameLayout/android.widget"
                                                                ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup[2]/android.view.ViewGroup/"
                                                                "android.view.ViewGroup/android.view.ViewGroup")))
            verify_toast = toast_keychain.is_displayed()
            if verify_toast is True:
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\Chat05_Toast de keychain se muestra correctamente.png")
                time.sleep(1)
                click_to_toast = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                click_to_toast.click()
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots"
                                              "\\Chats\\CHAT05_Se inicio automaticamente con el keychain.png")
                print("Se cumplio correctamente la funcion del keychain")
        except:
            print("Este dispositivo posee desbloqueo con sistema biometrico, se procedera a omitir el keychain")
            TapAndDragUtil.tap_screen(563, 1108, 50, driver)

        mail2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'ejemplo@mail.com')]")
        mail2.send_keys(secund_mail.value)

        passw2 = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        passw2.send_keys(secund_pass.value)

        TapAndDragUtil.search_buttom_contain_text("Continuar", driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 10). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                   ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element2.is_displayed(), f"El usuario {secund_name.value} no inicio sesion se califica como error, ya que" \
                                        f" no se muestran los elementos del navbar"
        print(f"Se inicio sesion correctamente con el usuario {secund_name.value}")
        # Acceder al modulo chats
        driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar()).click()
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//android.widget.TextView[contains(@text,'private')]")))
            chat_is_visible = True
        except:
            chat_is_visible = False
        assert chat_is_visible is False, "Error Se esta visualizando el chat privado"
        print("Se ha mantendio oculto el chat privado correctamente")
        time.sleep(1)

    def test_chat06(self):
        # buscar por nombre un grupo público, luego se debera anexar al equipo y enviar
        # una imagen y un mensaje de texto
        driver = self.driver

        driver.implicitly_wait(20)

        # Activar modo Desarrollo
        for i in range(1, 11):
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                          "LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout"
                                          "/android.widget.FrameLayout/android.widget.FrameLayout/android.view."
                                          "ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.widget.FrameLayout/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.v"
                                          "iew.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup/android.view.ViewGroup/android.view."
                                          "ViewGroup[1]/android.widget.ImageView").click()

        modal = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                              ".LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                              "android.view.ViewGroup/android.view.ViewGroup/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                              ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup").is_displayed()

        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\Chats06_ActivadoModoOculto.png")

        # Confirmar si la modal para modo desarrollo esta visible

        if modal is True:
            contra = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android.widget."
                                                   "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                                   ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android"
                                                   ".view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                                   "widget.EditText")
            contra.send_keys("12345678")
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/android"
                                          ".widget.FrameLayout/android.widget.FrameLayout/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.widget.FrameLayout/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android."
                                          "view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.widget.TextView").click()
        else:
            print("No se desplego la modal")
        time.sleep(2)
        modal2 = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                               "android.widget.LinearLayout/android.widget"
                                               ".FrameLayout/android.widget.LinearLayout/android."
                                               "widget.FrameLayout/android.widget.FrameLayout/android"
                                               ".view.ViewGroup[1]/android.view.ViewGroup/android.view.V"
                                               "iewGroup/android.view.ViewGroup/android.widget.FrameLayout"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android.vi"
                                               "ew.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGr"
                                               "oup/android.view.ViewGroup/android.widget.TextView").is_displayed()

        # Seleccionar opcion Desarrollo
        if modal2 is True:
            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout"
                                          "/android.widget.LinearLayout/android.widget"
                                          ".FrameLayout/android.widget.LinearLayout/"
                                          "android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android"
                                          ".view.ViewGroup/android.view.ViewGroup/android"
                                          ".view.ViewGroup/android.widget.FrameLayout"
                                          "/android.view.ViewGroup/android.view.ViewGroup"
                                          "/android.view.ViewGroup[2]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[3]/android.view.ViewGroup").click()

            driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android."
                                          "widget.LinearLayout/android.widget.FrameLayout/android.widget"
                                          ".LinearLayout/android.widget.FrameLayout/android.widget."
                                          "FrameLayout/android.view.ViewGroup[1]/android.view.ViewGroup"
                                          "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                          ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup/"
                                          "android.view.ViewGroup[2]/android.view.ViewGroup/android.view"
                                          ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup[5]").click()
            time.sleep(2)
        else:
            print("Modal par seleccionar opcion desarrollo no esta visible")
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT06_ActivadoModoDesarrollo.png")

        time.sleep(0.5)

        # Iniciar Sesion

        login = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "LinearLayout/android.widget.FrameLayout/android.widget."
                                              "FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.widget"
                                              ".FrameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                              ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                              "/android.view.ViewGroup[2]")
        login.click()

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/UserSinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']


        mail = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayou"
                                             "t/android.widget.FrameLayout/android.widget.LinearLayout/android."
                                             "widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGrou"
                                             "p/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                             "/android.widget.FrameLayout/android.view.ViewGroup/android.view.V"
                                             "iewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view"
                                             ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android."
                                             "view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup"
                                             "/android.view.ViewGroup[1]/android.widget.EditText")
        mail.send_keys(email.value)
        driver.hide_keyboard()
        time.sleep(1)

        password = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget."
                                                 "LinearLayout/android.widget.FrameLayout/android.widget.F"
                                                 "rameLayout/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "widget.FrameLayout/android.view.ViewGroup/android.view."
                                                 "ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                                 "/android.view.ViewGroup/android.view.ViewGroup/android."
                                                 "view.ViewGroup/android.view.ViewGroup[2]/android.widget."
                                                 "ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                                 "[2]/android.widget.EditText")
        password.send_keys(passw.value)
        driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Chats"
                                      "\\CHAT06_UsuarioCredenciales.png")

        access = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".LinearLayout/android.widget.FrameLayout/android.widget"
                                               ".FrameLayout/android.view.ViewGroup/android.view.ViewGrou"
                                               "p/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".widget.FrameLayout/android.view.ViewGroup/android.view"
                                               ".ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                                               "/android.view.ViewGroup/android.view.ViewGroup/android"
                                               ".view.ViewGroup/android.view.ViewGroup[2]/android.widget"
                                               ".ScrollView/android.view.ViewGroup/android.view.ViewGroup"
                                               "[4]/android.widget.TextView")

        access.click()

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, no se muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_chats = driver.find_element(By.XPATH, PathAndVariablesUtil.chats_navbar())
        enter_chats.click()
        # Encontrar icono de busqueda y filtrar un grupo de chat publico
        search_team = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]")))
        search_team.click()
        time.sleep(2)
        send_text = driver.find_elements(By.CLASS_NAME, "android.widget.EditText")
        send_text[0].send_keys("Chat Grupal test 01")
        time.sleep(2.5)
        result = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
        ele_result = driver.find_element(By.XPATH, str(result))
        ele_result.click()
        # Verificar la visualizacion de alerta de acceder al grupo cuando el usaurio no ha sido invitado a dicho grupo inicialmente
        alert_join_group = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, '¿Deseas ser participante de este grupo?')]")))
        assert alert_join_group.is_displayed(), "Error no esta mostrando la alerta para consultar si el usuario se desea anexar el grupo de chat"
        print("Alerta de participar en el grupo de chat mostrada correctamente")
        join_group = driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]")
        join_group.click()
        # Enviar mensaje al grupo
        text_box_in_chat = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//android.widget.EditText[contains(@text, 'Escribe un mensaje')]")))
        text_box_in_chat.send_keys("This is a message for the group, Im new here and I send my first words")
        send_text = driver.find_elements(By.CLASS_NAME, "android.view.ViewGroup")
        send_text[-1].click()
        # Enviar imagen al grupo
        search_img = driver.find_elements(By.XPATH, "android.widget.TextView")
        for img in search_img:
            print(search_img[search_img.index(img)].text, "posicion del texto es: ", search_img.index(img))




    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(MyTestChats)
    unittest.TextTestRunner(verbosity=2).run(suite)
