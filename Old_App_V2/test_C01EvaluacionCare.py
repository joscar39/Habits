import time
import unittest
import numpy.random
from selenium.webdriver.common.by import By
from appium import webdriver
import openpyxl
import numpy as np
import openpyxl as xl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from PageObjectModel.FlowsCostant.LoginAndRegisterFlows import LoginAndRegisterFlowsUtils
from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil
from PageObjectModel.action_app.Scroll_util import ScrollUtil
from PageObjectModel.action_app.Tap_Drag_Util import TapAndDragUtil


class EvaluacionCare(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": PathAndVariablesUtil.SetUpPlatform(),
            "appium:platformVersion": PathAndVariablesUtil.SetUpVersion(),
            "appium:deviceName": PathAndVariablesUtil.SetUpDeviceName(),
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_ENCU01(self):  # LLENAR ENCUESTA CARE Y SALIR DEL PROCESO DE SELECCIONAR HABITO
        driver = self.driver

        # Iniciar Sesion

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/Encuesta.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']

        # mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        # mail.send_keys(email.value)
        # driver.find_element(By.XPATH, "//android.widget.TextView[@text() = 'Continuar']").click()
        # time.sleep(1)
        #
        # password = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        # password.send_keys(passw.value)
        # # tap out the input to find the element correctly
        # TapAndDragUtil.tap_screen(887, 450, 50, driver)
        # # Click in button iniciar sesion
        # driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]").click()

        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        # Responder encuesta CARE
        # Verificar que el apartado de Care esta disponible
        care = WebDriverWait(driver, 10). \
            until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                                            ".LinearLayout/android.widget.FrameLayout/android.widget"
                                                            ".LinearLayout/android.widget.FrameLayout/android"
                                                            ".widget.FrameLayout/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.widget.FrameLayout/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup[1]/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.widget.FrameLayout/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android.view"
                                                            ".ViewGroup[1]/android.widget.FrameLayout/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android."
                                                            "view.ViewGroup/android.view.ViewGroup/android.view"
                                                            ".ViewGroup[2]/android.widget.ScrollView/android."
                                                            "view.ViewGroup/android.view.ViewGroup[2]/android"
                                                            ".view.ViewGroup/android.view.ViewGroup[1]/android"
                                                            ".view.ViewGroup/android.widget.TextView")))
        print(care.text)
        if care.text == "Evaluación CARE":  # Si hay coincidencia con el boton Evaluacion CARE
            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU02_BotonEvaluacionCare.png")
            care.click()
            print("Realizar evaluacion CARE")
            time.sleep(1)
            try:

                ScrollUtil.swipe_left_large(3, driver)

                # for i in range(1, 4):
                #     driver.swipe(591, 804, 131, 783)
                #     time.sleep(0.5)
                finish_intro = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget."
                                                   "LinearLayout/android.widget.FrameLayout/android."
                                                   "widget.LinearLayout/android.widget.FrameLayout/"
                                                   "android.widget.FrameLayout/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup/"
                                                   "android.view.ViewGroup/android.widget.FrameLayout"
                                                   "/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.widget."
                                                   "HorizontalScrollView/android.view.ViewGroup"
                                                   "/android.view.ViewGroup/android.widget.ScrollView"
                                                   "/android.view.ViewGroup/android.view."
                                                   "ViewGroup[5]").is_displayed()
                if finish_intro is True:
                    print("Continuar Encuesta")
            except:
                ScrollUtil.swipe_left_large(3, driver)
                # for i in range(1, 4):
                #     driver.swipe(591, 804, 131, 783)
                #     time.sleep(0.5)

            time.sleep(2)
            # Generar una lista de todos los campos visibles en la pantalla de encuesta
            initsurvey = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            initsurvey[10].click()
            time.sleep(1)
            onbording = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for i in range(1, 26):
                # se está generar un valor random entre 4 a 8 que representan
                # la posicion de los botones de las opciones a responder de la encuesta
                valor = np.random.randint(6, 11)
                onbording[valor].click()
                time.sleep(0.5)
            result = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text= 'Resultados']")))
            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU01_FinalizacionDeCareObtengo100ptos.png")
            result.click()
            time.sleep(2)

            print("Finalizado Proceso de Evaluacion Care")

        else:
            print("No esta disponible el boton de evaluacion")

        wb.close()
        # ENVIAR DATOS DEL USUARIO A LA BD DE PROCESO ONBORDING
        wb = xl.load_workbook(filesheet)
        wd = wb.active
        ws = wb.worksheets[0]
        receives = f"{PathAndVariablesUtil.db_path()}/recursos/EncuestaSeleccionarHab.xlsx"

        wb3 = xl.load_workbook(receives)
        ws3 = wb3.active

        ws3.insert_rows(1)
        for i in range(1, 2):
            for j in range(1, 11):
                c = ws.cell(row=i, column=j)

                ws3.cell(row=i, column=j).value = c.value
        wb3.save(str(receives))
        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
        wd.delete_rows(1)  # para la fila 1
        wb.save(filesheet)

    def test_ENCU02(self):
        # HACER PROCESO DESDE EL BOTON SELECCIONAR HABITO
        # Abrir Aplicacion y configurar modo Desarrollo
        global habits_selected
        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/EncuestaSeleccionarHab.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']

        # SetUtilDevMode.SetDevelopMode(driver)
        #
        # mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        # mail.send_keys(email.value)
        # driver.find_element(By.XPATH, "//android.widget.TextView[@text() = 'Continuar']").click()
        # time.sleep(1)
        #
        # password = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        # password.send_keys(passw.value)
        # # tap out the input to find the element correctly
        # TapAndDragUtil.tap_screen(887, 450, 50, driver)
        # # Click in button iniciar sesion
        # driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]").click()

        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        # Responder encuesta CARE
        # Verificar que el apartado de Care está disponible
        care = WebDriverWait(driver, 10). \
            until(EC.presence_of_element_located((By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")))

        if care.text == "Seleccionar habíto":  # Si hay coincidencia con el boton Evaluacion CARE
            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU02_BotonEvaluacionCare.png")
            care.click()

            time.sleep(3)

            connect = "Conectar con otros y consigo mismo, cultivar la gratitud, generosidad y bondad"
            activate = "Practicar de manera regular una actividad física"
            recharge = "Recargar de forma activa y pasiva la energía"
            eat_real = "Comer principalmente basado en plantas, sin procesados, local y de temporada"
            lt = [connect, activate, recharge, eat_real]
            pillar = numpy.random.choice(lt)
            time.sleep(3)
            percen = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for per in percen:
                if percen[percen.index(per)].text == pillar:
                    print(percen[percen.index(per)].text)
                    percen[percen.index(per)].click()
                    time.sleep(1)
                    break
            time.sleep(3)
            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU02_SeleccionarPilar.png")

            if pillar == connect:  # PILAR CONNECT
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Connect")
                # Posiciones de las opciones dentro de la lista de elementos
                ser_generoso = "Ser generoso"
                conectar_con_otro = "Conectar con otros"
                ser_agradecido = "Ser agradecido"
                objetivos = [ser_generoso, conectar_con_otro, ser_agradecido]

                # escoger una opcion random
                seleopc = numpy.random.choice(objetivos)
                time.sleep(1)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU02_SeleccionarObjetivo.png")
                time.sleep(1)
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)
                if seleopc == ser_generoso:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {ser_generoso}")
                    obj3 = "Comunico explícitamente admiración y afecto"
                    obj4 = "Practico un acto de bondad"
                    obj5 = "Comparto mis conocimientos e ideas con otros"
                    obj6 = "Dedico tiempo a otros para ayudar o acompañar"
                    objlist = [obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj3:
                            habits_selected = obj3
                            print(f"Habito seleccionado: {habits_selected}")
                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito seleccionado: {habits_selected}")
                        elif habi == obj5:
                            habits_selected = obj5
                            print(f"Habito seleccionado: {habits_selected}")
                        else:
                            habits_selected = obj6
                            print(f"Habito seleccionado: {habits_selected}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == conectar_con_otro:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {conectar_con_otro}")

                    obj3 = "Escucho al otro sin preparar mi respuesta"
                    obj4 = "Uso todos mis sentidos en cada interacción que tengo con otros"
                    obj5 = "Guardo los dispositivos electrónicos cuando hablo con otros"
                    obj6 = "Hablo con un ser querido al menos una vez al día"
                    objlist = [obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")

                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj5:
                            habits_selected = obj5
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj6
                            print(f"Habito selecionado: {habits_selected}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo Selecionado: {ser_agradecido}")

                    obj1 = "Expreso activamente agradecimiento a otros"
                    obj2 = "Escribo al despertarme 3 cosas que agradezco en mi vida"
                    obj3 = "Identifico al finalizar el día mi mejor momento"
                    obj4 = "Resalto lo bueno de las personas o situaciones"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado es: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado es: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado es: {habits_selected}")
                        else:
                            habits_selected = obj4
                            print(f"Habito selecionado es: {habits_selected}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            elif pillar == activate:  # PILAR ACTIVATE
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Activate")

                # Posiciones de las opciones dentro de la lista de elementos
                empezar_vida_activa = "Empezar una vida activa"
                mantener_ganar = "Mantener/ganar condición física"
                fortalecer_core = "Fortalecer core (abdominales, espalda, glúteos)"
                ganar_tonificar_masa = "Ganar/tonificar masa muscular"

                opc = [empezar_vida_activa, mantener_ganar, fortalecer_core, ganar_tonificar_masa]
                # escoger una opcion random
                seleopc = numpy.random.choice(opc)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                time.sleep(1)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU02_SeleccionarObjetivo.png")
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)

                if seleopc == empezar_vida_activa:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {empezar_vida_activa}")

                    obj1 = "Estiro durante 10 mins, alternando la parte superior e inferior"
                    obj2 = "Camino 15 min"
                    obj3 = "Hago una pausa cada hora y camino 5 min"
                    obj4 = "Camino 10 min después de mi comida de la tarde"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"El Habito seleccionado es: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"El Habito seleccionado es: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"El Habito seleccionado es: {habits_selected}")
                        else:
                            habits_selected = obj4
                            print(f"El Habito seleccionado es: {habits_selected}")


                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == mantener_ganar:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {mantener_ganar}")

                    obj1 = "Hago 30 min de mi cardio preferido"
                    obj2 = "Hago 20 rep de saltos largos y 30 segs de escaladores"
                    obj3 = "Hago uno de estos ejercicios cada día: Saltos con elevación de rodillas: 30segx10seg de descanso. Desplantes: 3 seriesx30seg"
                    obj4 = "Contraigo abdomen y glúteos 5vecesx5sg, 3 veces al día, en el lugar en donde esté"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.swipe_up_large(1, driver)
                    time.sleep(1)
                    ele_contains = driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{habi}')]")
                    ele_contains.click()
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"El Habito seleccionado es: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"El Habito seleccionado es: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"El Habito seleccionado es: {habits_selected}")
                        else:
                            habits_selected = obj4
                            print(f"El Habito seleccionado es: {habits_selected}")


                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == fortalecer_core:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo Selecionado: {fortalecer_core}")
                    obj1 = "Hago planchas laterales por cada lado entre 30seg a 1min. Hago abs en V entre 30seg a 1min"
                    obj2 = "Mantengo la posición de superman por 5seg, 10rep. Mantengo la posición de puente durante 30seg"
                    obj3 = "Hago plancha por 1 min. Hago giros en V o giros rusos por 2 min"
                    objlist = [obj1, obj2, obj3]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado es: {habits_selected}")

                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado es: {habits_selected}")
                        else:
                            habits_selected = obj3
                            print(f"Habito selecionado es: {habits_selected}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    finish_buttom = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in finish_buttom:
                        if finish_buttom[finish_buttom.index(sig)].text == " Finalizar":
                            finish_buttom[finish_buttom.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {ganar_tonificar_masa}")

                    obj1 = "Hago uno de estos ejercicios cada día de la semana: Bíceps: 40rep de bíceps. Extensiones de triceps:16rep. Sentadillas: 20repx3 series"
                    obj2 = "Hago uno de estos ejercicios cada día de la semana: Tríceps en una silla: 25repx3 series. Flexiones de pecho: 20repx3 series. Desplantes: 15repx2 series"
                    obj3 = "Hago uno de estos ejercicios cada día de la semana: Extensión de hombros: 15repx3 series. Flexiones de pecho: 20repx3 series. Desplantes: 15repx2 series"
                    obj4 = "Hago uno de estos ejercicios cada día de la semana: Bíceps 16repx4 series, mientras lo haces deja el brazo contrario doblado. Extensión de hombros: 15repx3 series. Sentadillas 20repx3 series"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"EL habitos seleccionado es: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"EL habitos seleccionado es: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"EL habitos seleccionado es: {habits_selected}")
                        else:
                            habits_selected = obj3
                            print(f"EL habitos seleccionado es: {habits_selected}")

                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                            "\\ENCU02_Seleccionado el Habito.png")

                        # Pulsar boton siguiente
                        finish_buttom = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                        for sig in finish_buttom:
                            if finish_buttom[finish_buttom.index(sig)].text == " Finalizar":
                                finish_buttom[finish_buttom.index(sig)].click()
                                time.sleep(1)
                                break
                        time.sleep(1)

                        starthabits = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/"
                                                          "android.widget.LinearLayout/android."
                                                          "widget.FrameLayout/android.widget."
                                                          "LinearLayout/android.widget.Frame"
                                                          "Layout/android.widget.FrameLayout"
                                                          "/android.view.ViewGroup/android."
                                                          "view.ViewGroup/android.view.ViewGroup"
                                                          "/android.view.ViewGroup/android."
                                                          "widget.FrameLayout/android.view."
                                                          "ViewGroup/android.view.ViewGroup"
                                                          "/android.view.ViewGroup/android."
                                                          "view.ViewGroup/android.view."
                                                          "ViewGroup/android.view.ViewGroup"
                                                          "/android.view.ViewGroup/android"
                                                          ".view.ViewGroup")
                        starthabits.click()
                        a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                        habits_is_visible = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, a)))
                        see_more_habits = driver.find_elements(By.XPATH,
                                                               "//android.widget.TextView[contains(@text, 'Ver más')]")
                        see_more_habits[0].click()
                        b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                        time.sleep(10)
                        compare_habits_app = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, b)))
                        assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                        print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                            "\\ENCU02_Habito seleccionado correctamente.png")
                        driver.find_element(By.XPATH,
                                            "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

            elif pillar == recharge:  # PILAR RECHARGE
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Recharge")

                # Posiciones de las opciones dentro de la lista de elementos
                dormir_profundamente = "Dormir profundamente"
                recargar_activamente = "Recargar activamente mi energía"
                opc = [dormir_profundamente, recargar_activamente]
                # escoger una opcion random
                seleopc = numpy.random.choice(opc)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                time.sleep(1)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU02_SeleccionarObjetivo.png")
                time.sleep(1)
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)

                if seleopc == dormir_profundamente:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {dormir_profundamente}")
                    time.sleep(1)

                    obj1 = "Hago mi rutina de descanso previo a dormirme"
                    obj2 = "Hago una lista de pendientes antes de dormir"
                    obj3 = "Hago una actividad que me relaje antes de dormir"
                    obj4 = "Como mín 3 hrs antes de dormir"
                    obj5 = "Me desconecto de la tecnología 2 hrs antes de ir a dormir"
                    obj6 = "Estiro antes de dormir"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.swipe_up_large(1, driver)
                    TapAndDragUtil.search_buttom_contain_text(habi, driver)
                    # ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj5:
                            habits_selected = obj5
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj6
                            print(f"Habito selecionado: {habits_selected}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {recargar_activamente}")
                    time.sleep(1)

                    obj1 = "Respiro de forma consciente entre 5 y 10 min"
                    obj2 = "Practico un arte oriental (Tai Chi, Aikido, Chi Kung)"
                    obj3 = "Conecto diariamente con la naturaleza"
                    obj4 = "Dedico tiempo a algo que me apasiona"
                    obj5 = "Hago mi rutina de recarga de energía"
                    obj6 = "Practico yoga al menos 10 min"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj5:
                            habits_selected = obj5
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj6
                            print(f"Habito selecionado: {habits_selected}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            else:  # PILAR EAT REAL
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Eat Real")

                # Posiciones de las opciones dentro de la lista de elementos
                equilibrar_alimentacion = "Equilibrar mi alimentación"
                comer_conscientemente = "Comer conscientemente"
                comer_sano = "Comer sano y delicioso"
                cuidar_alimentacion = "Cuidar mi alimentación"
                opc = [equilibrar_alimentacion, comer_conscientemente, comer_sano, cuidar_alimentacion]
                # escoger una opcion random
                seleopc = numpy.random.choice(opc)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                time.sleep(1)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU02_SeleccionarObjetivo.png")
                time.sleep(1)
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)

                if seleopc == equilibrar_alimentacion:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {equilibrar_alimentacion}")

                    obj1 = "Incluyo verduras de 3 colores diferentes en mis ensaladas"
                    obj2 = "Evito alimentos empacados"
                    obj3 = "Incluyo más fibra en mi alimentación"
                    obj4 = "Disminuyo las bebidas envasadas"
                    obj5 = "Dedico 20 minutos a cada comida"
                    obj6 = "Incluyo los 3 grupos de alimentos en mis tres comidas principales"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.swipe_up_large(1, driver)
                    time.sleep(1)
                    ele_contains = driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{habi}')]")
                    ele_contains.click()
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj5:
                            habits_selected = obj5
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj6
                            print(f"Habito selecionado: {habits_selected}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == comer_conscientemente:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {comer_conscientemente}")
                    time.sleep(1)

                    obj1 = "Tengo un horario fijo para cada comida y lo respeto"
                    obj2 = "Varío mis alimentos en cada grupo alimenticio"
                    obj3 = "Mi cena es vegetariana"
                    obj4 = "No como más, luego de cenar"
                    obj5 = "Como en 20 min en la mesa y sin distractores"
                    obj6 = "Evito los alimentos con azúcar"
                    obj7 = "Evito los alimentos fritos"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6, obj7]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj5:
                            habits_selected = obj5
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj6:
                            habits_selected = obj6
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj7
                            print(f"Habito selecionado: {habits_selected}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente

                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == comer_sano:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo Selecionado: {comer_sano}")
                    time.sleep(1)

                    obj1 = "Tengo un horario para comer y no pico entre comidas"
                    obj2 = "Enfoco mis 5 sentidos a la comida para disfrutarla"
                    obj3 = "Como hasta sentirme satisfecho en un 80%"
                    obj4 = "Evito la comida rápida"
                    obj5 = "Diseño un menú semanal y lo cumplo diariamente"
                    objlist = [obj1, obj2, obj3, obj4, obj5]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado: {habits_selected}")

                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj4:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj5
                            print(f"Habito selecionado: {habits_selected}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {cuidar_alimentacion}")
                    time.sleep(2)

                    obj1 = "Cuido mis porciones para evitar excesos"
                    obj2 = "Como mínimo 2 frutas de diferentes colores"
                    obj3 = "Me hidrato con 6 a 8 vasos de agua"
                    obj4 = "Consumo frutas y verduras de temporada"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_selected = obj1
                            print(f"Habito selecionado: {habits_selected}")

                        elif habi == obj2:
                            habits_selected = obj2
                            print(f"Habito selecionado: {habits_selected}")
                        elif habi == obj3:
                            habits_selected = obj3
                            print(f"Habito selecionado: {habits_selected}")
                        else:
                            habits_selected = obj4
                            print(f"Habito selecionado: {habits_selected}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_selected == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_selected}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_selected}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU02_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            wb.close()
            # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
            wb = xl.load_workbook(filesheet)
            wd = wb.active
            ws = wb.worksheets[0]

            receives = f"{PathAndVariablesUtil.db_path()}/recursos/RachaDiaria.xlsx"

            wb3 = xl.load_workbook(receives)
            ws3 = wb3.active

            ws3.insert_rows(1)
            for i in range(1, 2):
                for j in range(1, 11):
                    c = ws.cell(row=i, column=j)

                    ws3.cell(row=i, column=j).value = c.value

            wb3.save(str(receives))

            wd.delete_rows(1)  # para la fila 1
            wb.save(filesheet)

        else:
            print("El elemento no coincide con el boton buscado ")

    def test_ENCU03(self):  # HACER PROCESO COMPLETO DESDE BOTON ENCUESTA CARE
        # Abrir Aplicacion y configurar modo Desarrollo
        global habits_sele
        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/ProcesoOnbording.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']

        # SetUtilDevMode.SetDevelopMode(driver)
        #
        # mail = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Correo electrónico')]")
        # mail.send_keys(email.value)
        # driver.find_element(By.XPATH, "//android.widget.TextView[@text() = 'Continuar']").click()
        # time.sleep(1)
        #
        # password = driver.find_element(By.XPATH, "//android.widget.EditText[contains(@text, 'Ingresa tu contraseña')]")
        # password.send_keys(passw.value)
        # # tap out the input to find the element correctly
        # TapAndDragUtil.tap_screen(887, 450, 50, driver)
        # # Click in button iniciar sesion
        # driver.find_element(By.XPATH, "//android.widget.TextView[contains(@text, 'Iniciar sesión')]").click()
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        # Responder encuesta CARE
        # Esperar a que se muestre el apartado care
        care = WebDriverWait(driver, 10). \
            until(EC.presence_of_element_located((By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                                            ".LinearLayout/android.widget.FrameLayout/android.widget"
                                                            ".LinearLayout/android.widget.FrameLayout/android"
                                                            ".widget.FrameLayout/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.widget.FrameLayout/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android"
                                                            ".view.ViewGroup[1]/android.view.ViewGroup/android"
                                                            ".view.ViewGroup/android.widget.FrameLayout/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android.view"
                                                            ".ViewGroup[1]/android.widget.FrameLayout/android"
                                                            ".view.ViewGroup/android.view.ViewGroup/android."
                                                            "view.ViewGroup/android.view.ViewGroup/android.view"
                                                            ".ViewGroup[2]/android.widget.ScrollView/android."
                                                            "view.ViewGroup/android.view.ViewGroup[2]/android"
                                                            ".view.ViewGroup/android.view.ViewGroup[1]/android"
                                                            ".view.ViewGroup/android.widget.TextView")))

        if care.text == "Evaluación CARE":  # Si hay coincidencia con el boton Evaluacion CARE
            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU03_BotonEvaluacionCare.png")
            care.click()
            print("Realizar evaluacion CARE COMPLETO")
            try:

                ScrollUtil.swipe_left_large(3, driver)
                # for i in range(1, 4):
                #     driver.swipe(591, 804, 131, 783)
                #     time.sleep(0.25)
                finish_intro = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/android.widget"
                                                             ".LinearLayout/android.widget.FrameLayout/android."
                                                             "widget.LinearLayout/android.widget.FrameLayout/"
                                                             "android.widget.FrameLayout/android.view.ViewGroup"
                                                             "/android.view.ViewGroup/android.view.ViewGroup"
                                                             "/android.view.ViewGroup/android.widget.FrameLayout"
                                                             "/android.view.ViewGroup/android.view.ViewGroup/"
                                                             "android.view.ViewGroup/android.view.ViewGroup/android"
                                                             ".view.ViewGroup/android.widget.HorizontalScrollVie"
                                                             "/android.view.ViewGroup/android.view.ViewGroup"
                                                             "/android.widget.ScrollView/android.view.ViewGroup"
                                                             "/android.view.ViewGroup[5]").is_displayed()
                if finish_intro is True:
                    print("Continuar Encuesta")
            except:
                ScrollUtil.swipe_left_large(3, driver)
                # for i in range(1, 4):
                #     driver.swipe(591, 804, 131, 783)
                #     time.sleep(0.25)

            time.sleep(2)
            # Generar una lista de todos los campos visibles en la pantalla de encuesta
            initsurvey = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            initsurvey[10].click()
            time.sleep(1)
            onbording = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for i in range(1, 26):
                # se está generar un valor random entre 6 a 11 que representan
                # la posicion de los botones de las opciones a responder de la encuesta
                valor = np.random.randint(6, 11)
                onbording[valor].click()
                time.sleep(0.5)

            result = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[@text= 'Resultados']")))

            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU03_FinalizacionDeCareObtengo100ptos.png")
            result.click()

            time.sleep(3)

            # INCIALIZAR CON EL BOTON ELEGIR TU HABITO
            driver.find_element(By.XPATH, "//android.widget.TextView[@text= 'Elige tu hábito']").click()

            # Elegir un pilar

            time.sleep(3)
            connect = "Conectar con otros y consigo mismo, cultivar la gratitud, generosidad y bondad"
            activate = "Practicar de manera regular una actividad física"
            recharge = "Recargar de forma activa y pasiva la energía"
            eat_real = "Comer principalmente basado en plantas, sin procesados, local y de temporada"
            lt = [connect, activate, recharge, eat_real]
            pillar = numpy.random.choice(lt)
            percen = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
            for per in percen:
                if percen[percen.index(per)].text == pillar:
                    print(percen[percen.index(per)].text)
                    percen[percen.index(per)].click()
                    time.sleep(1)
                    break
            time.sleep(3)
            driver.get_screenshot_as_file(
                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                "\\ENCU03_SeleccionarPilar.png")

            if pillar == connect:  # PILAR CONNECT
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Connect")

                # Posiciones de las opciones dentro de la lista de elementos
                ser_generoso = "Ser generoso"
                conectar_con_otro = "Conectar con otros"
                ser_agradecido = "Ser agradecido"
                objetivos = [ser_generoso, conectar_con_otro, ser_agradecido]

                # escoger una opcion random
                seleopc = numpy.random.choice(objetivos)
                time.sleep(1)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU03_SeleccionarObjetivo.png")
                time.sleep(1)
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)

                if seleopc == ser_generoso:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {ser_generoso}")

                    obj3 = "Comunico explícitamente admiración y afecto"
                    obj4 = "Practico un acto de bondad"
                    obj5 = "Comparto mis conocimientos e ideas con otros"
                    obj6 = "Dedico tiempo a otros para ayudar o acompañar"
                    objlist = [obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj3:
                            habits_sele = obj3
                            print(f"Habito seleccionado: {habits_sele}")
                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito seleccionado: {habits_sele}")
                        elif habi == obj5:
                            habits_sele = obj5
                            print(f"Habito seleccionado: {habits_sele}")
                        else:
                            habits_sele = obj6
                            print(f"Habito seleccionado: {habits_sele}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == conectar_con_otro:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {conectar_con_otro}")
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj3 = "Escucho al otro sin preparar mi respuesta"
                    obj4 = "Uso todos mis sentidos en cada interacción que tengo con otros"
                    obj5 = "Guardo los dispositivos electrónicos cuando hablo con otros"
                    obj6 = "Hablo con un ser querido al menos una vez al día"
                    objlist = [obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")

                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj5:
                            habits_sele = obj5
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj6
                            print(f"Habito selecionado: {habits_sele}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo Selecionado: {ser_agradecido}")
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Expreso activamente agradecimiento a otros"
                    obj2 = "Escribo al despertarme 3 cosas que agradezco en mi vida"
                    obj3 = "Identifico al finalizar el día mi mejor momento"
                    obj4 = "Resalto lo bueno de las personas o situaciones"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado es: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado es: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado es: {habits_sele}")
                        else:
                            habits_sele = obj4
                            print(f"Habito selecionado es: {habits_sele}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            elif pillar == activate:
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Activate")
                # Posiciones de las opciones dentro de la lista de elementos
                empezar_vida_activa = "Empezar una vida activa"
                mantener_ganar = "Mantener/ganar condición física"
                fortalecer_core = "Fortalecer core (abdominales, espalda, glúteos)"
                ganar_tonificar_masa = "Ganar/tonificar masa muscular"

                opc = [empezar_vida_activa, mantener_ganar, fortalecer_core, ganar_tonificar_masa]
                # escoger una opcion random
                seleopc = numpy.random.choice(opc)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                time.sleep(1)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU02_SeleccionarObjetivo.png")
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        break
                time.sleep(2)

                if seleopc == empezar_vida_activa:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {empezar_vida_activa}")
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Estiro durante 10 mins, alternando la parte superior e inferior"
                    obj2 = "Camino 15 min"
                    obj3 = "Hago una pausa cada hora y camino 5 min"
                    obj4 = "Camino 10 min después de mi comida de la tarde"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"El Habito seleccionado es: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"El Habito seleccionado es: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"El Habito seleccionado es: {habits_sele}")
                        else:
                            habits_sele = obj4
                            print(f"El Habito seleccionado es: {habits_sele}")


                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()

                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                elif seleopc == mantener_ganar:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {mantener_ganar}")
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Hago 30 min de mi cardio preferido"
                    obj2 = "Hago 20 rep de saltos largos y 30 segs de escaladores"
                    obj3 = "Hago uno de estos ejercicios cada día: Saltos con elevación de rodillas: 30segx10seg de descanso. Desplantes: 3 seriesx30seg"
                    obj4 = "Contraigo abdomen y glúteos 5vecesx5sg, 3 veces al día, en el lugar en donde esté"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.swipe_up_large(1, driver)
                    time.sleep(1)
                    ele_contains = driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{habi}')]")
                    ele_contains.click()
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"El Habito seleccionado es: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"El Habito seleccionado es: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"El Habito seleccionado es: {habits_sele}")
                        else:
                            habits_sele = obj4
                            print(f"El Habito seleccionado es: {habits_sele}")


                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                elif seleopc == fortalecer_core:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo Selecionado: {fortalecer_core}")
                    obj1 = "Hago planchas laterales por cada lado entre 30seg a 1min. Hago abs en V entre 30seg a 1min"
                    obj2 = "Mantengo la posición de superman por 5seg, 10rep. Mantengo la posición de puente durante 30seg"
                    obj3 = "Hago plancha por 1 min. Hago giros en V o giros rusos por 2 min"
                    objlist = [obj1, obj2, obj3]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado es: {habits_sele}")

                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado es: {habits_sele}")
                        else:
                            habits_sele = obj3
                            print(f"Habito selecionado es: {habits_sele}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    finish_buttom = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in finish_buttom:
                        if finish_buttom[finish_buttom.index(sig)].text == " Finalizar":
                            finish_buttom[finish_buttom.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {ganar_tonificar_masa}")

                    obj1 = "Hago uno de estos ejercicios cada día de la semana: Bíceps: 40rep de bíceps. Extensiones de triceps:16rep. Sentadillas: 20repx3 series"
                    obj2 = "Hago uno de estos ejercicios cada día de la semana: Tríceps en una silla: 25repx3 series. Flexiones de pecho: 20repx3 series. Desplantes: 15repx2 series"
                    obj3 = "Hago uno de estos ejercicios cada día de la semana: Extensión de hombros: 15repx3 series. Flexiones de pecho: 20repx3 series. Desplantes: 15repx2 series"
                    obj4 = "Hago uno de estos ejercicios cada día de la semana: Bíceps 16repx4 series, mientras lo haces deja el brazo contrario doblado. Extensión de hombros: 15repx3 series. Sentadillas 20repx3 series"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)

                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"EL habitos seleccionado es: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"EL habitos seleccionado es: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"EL habitos seleccionado es: {habits_sele}")
                        else:
                            habits_sele = obj3
                            print(f"EL habitos seleccionado es: {habits_sele}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    finish_buttom = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    for sig in finish_buttom:
                        if finish_buttom[finish_buttom.index(sig)].text == " Finalizar":
                            finish_buttom[finish_buttom.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/"
                                                      "android.widget.LinearLayout/android."
                                                      "widget.FrameLayout/android.widget."
                                                      "LinearLayout/android.widget.Frame"
                                                      "Layout/android.widget.FrameLayout"
                                                      "/android.view.ViewGroup/android."
                                                      "view.ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android."
                                                      "widget.FrameLayout/android.view."
                                                      "ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android."
                                                      "view.ViewGroup/android.view."
                                                      "ViewGroup/android.view.ViewGroup"
                                                      "/android.view.ViewGroup/android"
                                                      ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele== compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            elif pillar == recharge:
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Recharge")
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Posiciones de las opciones dentro de la lista de elementos
                dormir_profundamente = "Dormir profundamente"
                recargar_activamente = "Recargar activamente mi energía"
                opc = [dormir_profundamente, recargar_activamente]
                # escoger una opcion random
                seleopc = numpy.random.choice(opc)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                time.sleep(1)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU03_SeleccionarObjetivo.png")
                time.sleep(1)
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)

                if seleopc == dormir_profundamente:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {dormir_profundamente}")
                    time.sleep(1)
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Hago mi rutina de descanso previo a dormirme"
                    obj2 = "Hago una lista de pendientes antes de dormir"
                    obj3 = "Hago una actividad que me relaje antes de dormir"
                    obj4 = "Como mín 3 hrs antes de dormir"
                    obj5 = "Me desconecto de la tecnología 2 hrs antes de ir a dormir"
                    obj6 = "Estiro antes de dormir"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.swipe_up_large(1, driver)
                    TapAndDragUtil.search_buttom_contain_text(habi, driver)
                    # ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj5:
                            habits_sele = obj5
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj6
                            print(f"Habito selecionado: {habits_sele}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {recargar_activamente}")
                    time.sleep(1)
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Respiro de forma consciente entre 5 y 10 min"
                    obj2 = "Practico un arte oriental (Tai Chi, Aikido, Chi Kung)"
                    obj3 = "Conecto diariamente con la naturaleza"
                    obj4 = "Dedico tiempo a algo que me apasiona"
                    obj5 = "Hago mi rutina de recarga de energía"
                    obj6 = "Practico yoga al menos 10 min"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj5:
                            habits_sele = obj5
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj6
                            print(f"Habito selecionado: {habits_sele}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele== compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            else:  # PILAR EAT REAL
                # INDICAR OBJETIVOS DEL PILAR SELECCIONADO
                print("Seleccionado Pilar Eat Real")
                # listado de elementos en pantalla
                option = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                # Posiciones de las opciones dentro de la lista de elementos
                equilibrar_alimentacion = "Equilibrar mi alimentación"
                comer_conscientemente = "Comer conscientemente"
                comer_sano = "Comer sano y delicioso"
                cuidar_alimentacion = "Cuidar mi alimentación"
                opc = [equilibrar_alimentacion, comer_conscientemente, comer_sano, cuidar_alimentacion]
                # escoger una opcion random
                seleopc = numpy.random.choice(opc)
                ScrollUtil.scroll_to_text(seleopc, " Siguiente", driver)
                driver.get_screenshot_as_file(
                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                    "\\ENCU03_SeleccionarObjetivo.png")
                time.sleep(1)
                # Pulsar boton siguiente
                for sig in option:
                    if option[option.index(sig)].text == " Siguiente":
                        option[option.index(sig)].click()
                        time.sleep(1)
                        break
                time.sleep(1)

                if seleopc == equilibrar_alimentacion:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {equilibrar_alimentacion}")
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Incluyo verduras de 3 colores diferentes en mis ensaladas"
                    obj2 = "Evito alimentos empacados"
                    obj3 = "Incluyo más fibra en mi alimentación"
                    obj4 = "Disminuyo las bebidas envasadas"
                    obj5 = "Dedico 20 minutos a cada comida"
                    obj6 = "Incluyo los 3 grupos de alimentos en mis tres comidas principales"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.swipe_up_large(1, driver)
                    time.sleep(1)
                    ele_contains = driver.find_element(By.XPATH, f"//android.widget.TextView[contains(@text, '{habi}')]")
                    ele_contains.click()
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado: {habits_sele}")

                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj5:
                            habits_sele = obj5
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj6
                            print(f"Habito selecionado: {habits_sele}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)

                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == comer_conscientemente:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo seleccionado: {comer_conscientemente}")
                    time.sleep(1)
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Tengo un horario fijo para cada comida y lo respeto"
                    obj2 = "Varío mis alimentos en cada grupo alimenticio"
                    obj3 = "Mi cena es vegetariana"
                    obj4 = "No como más, luego de cenar"
                    obj5 = "Como en 20 min en la mesa y sin distractores"
                    obj6 = "Evito los alimentos con azúcar"
                    obj7 = "Evito los alimentos fritos"
                    objlist = [obj1, obj2, obj3, obj4, obj5, obj6, obj7]
                    habi = numpy.random.choice(objlist)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj5:
                            habits_sele = obj5
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj6:
                            habits_sele = obj6
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj7
                            print(f"Habito selecionado: {habits_sele}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

                elif seleopc == comer_sano:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"Objetivo Selecionado: {comer_sano}")
                    time.sleep(1)
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Tengo un horario para comer y no pico entre comidas"
                    obj2 = "Enfoco mis 5 sentidos a la comida para disfrutarla"
                    obj3 = "Como hasta sentirme satisfecho en un 80%"
                    obj4 = "Evito la comida rápida"
                    obj5 = "Diseño un menú semanal y lo cumplo diariamente"
                    objlist = [obj1, obj2, obj3, obj4, obj5]
                    habi = numpy.random.choice(objlist)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado: {habits_sele}")

                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj4:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj5
                            print(f"Habito selecionado: {habits_sele}")
                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")

                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()
                else:
                    # SELECCIONAR HABITO DE OBJETIVO SELECCIONADO
                    print(f"objetivo seleccionado: {cuidar_alimentacion}")
                    time.sleep(2)
                    obj = driver.find_elements(By.CLASS_NAME, "android.widget.TextView")
                    obj1 = "Cuido mis porciones para evitar excesos"
                    obj2 = "Como mínimo 2 frutas de diferentes colores"
                    obj3 = "Me hidrato con 6 a 8 vasos de agua"
                    obj4 = "Consumo frutas y verduras de temporada"
                    objlist = [obj1, obj2, obj3, obj4]
                    habi = numpy.random.choice(objlist)
                    print(habi)
                    ScrollUtil.scroll_to_text(habi, " Finalizar", driver)
                    time.sleep(1)
                    try:
                        if habi == obj1:
                            habits_sele = obj1
                            print(f"Habito selecionado: {habits_sele}")

                        elif habi == obj2:
                            habits_sele = obj2
                            print(f"Habito selecionado: {habits_sele}")
                        elif habi == obj3:
                            habits_sele = obj3
                            print(f"Habito selecionado: {habits_sele}")
                        else:
                            habits_sele = obj4
                            print(f"Habito selecionado: {habits_sele}")

                    except:
                        print("ERROR no se consiguio la opcion de habito seleccionado revisar codigo")

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU03_Seleccionado el Habito.png")
                    # Pulsar boton siguiente
                    for sig in obj:
                        if obj[obj.index(sig)].text == " Finalizar":
                            obj[obj.index(sig)].click()
                            time.sleep(1)
                            break
                    time.sleep(1)
                    starthabits = driver.find_element(By.XPATH, "/hierarchy/android.widget.FrameLayout/"
                                                                "android.widget.LinearLayout/android."
                                                                "widget.FrameLayout/android.widget."
                                                                "LinearLayout/android.widget.Frame"
                                                                "Layout/android.widget.FrameLayout"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "widget.FrameLayout/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android."
                                                                "view.ViewGroup/android.view."
                                                                "ViewGroup/android.view.ViewGroup"
                                                                "/android.view.ViewGroup/android"
                                                                ".view.ViewGroup")
                    starthabits.click()
                    a = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[2]/android.widget.TextView[2]"
                    habits_is_visible = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, a)))
                    see_more_habits = driver.find_elements(By.XPATH,
                                                           "//android.widget.TextView[contains(@text, 'Ver más')]")
                    see_more_habits[0].click()
                    b = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[2]"
                    time.sleep(10)
                    compare_habits_app = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, b)))
                    assert habits_sele == compare_habits_app.text, f" Error no es el habito correcto. Se selecciono: {habits_sele}\n El habito mostrado es: {compare_habits_app.text}"
                    print(f"Finalizada la prueba exitosamente el habito correcto es: {habits_sele}")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\EncuestaHabito"
                        "\\ENCU0_Habito seleccionado correctamente.png")
                    driver.find_element(By.XPATH,
                                        "//android.widget.TextView[contains(@text, 'Aceptar')]").click()

            wb = xl.load_workbook(filesheet)
            wd = wb.active
            ws = wb.worksheets[0]

            receives = f"{PathAndVariablesUtil.db_path()}/recursos/RachaDiaria.xlsx"

            wb3 = xl.load_workbook(receives)
            ws3 = wb3.active

            ws3.insert_rows(1)
            for i in range(1, 2):
                for j in range(1, 11):
                    c = ws.cell(row=i, column=j)

                    ws3.cell(row=i, column=j).value = c.value
            wb3.save(str(receives))

            wd.delete_rows(1)  # para la fila 1
            wb.save(filesheet)


        else:
            print("El elemento no coincide con el boton buscado ")

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(EvaluacionCare)
    unittest.TextTestRunner(verbosity=2).run(suite)
