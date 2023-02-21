import time
import unittest
import json
import requests
import openpyxl as xl
import openpyxl

from datetime import datetime
from datetime import timedelta

from selenium.webdriver.common.by import By
from appium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from PageObjectModel.FlowsCostant.LoginAndRegisterFlows import LoginAndRegisterFlowsUtils
from PageObjectModel.Global_variables.Path_Data import PathAndVariablesUtil


class Retos(unittest.TestCase):
    def setUp(self):
        # Inicializacion de Manejador Appium
        caps = {
            "platformName": "Android",
            "appium:platformVersion": "10",
            "appium:deviceName": "dandelion_global ",
            "appium:automationName": "UiAutomator2",
            "appium:appPackage": "com.habitsv2",
            "appium:appActivity": ".MainActivity",
            "appium:ensureWebviewsHavePages": True,
            "appium:nativeWebScreenshot": True,
            "appium:newCommandTimeout": 3600,
            "appium:connectHardwareKeyboard": True
        }

        self.driver = webdriver.Remote("http://localhost:4723/wd/hub", caps)

    def test_RET01(self):
        # Inscribirse a un reto INDIVIDUAL y cumplir con la meta acordada para recibir los puntos indicados
        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/SinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']

        # Iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)
        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_navbar = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        # Guardar puntos del score del usuario
        points = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
        valuepoint = points.text

        # Acceder a MI DIA
        enter_navbar.click()
        time.sleep(5)
        driver.swipe(328, 1034, 328, 823)
        time.sleep(6)

        # VERIFICAR SI LOS RETOS ESTA DISPONIBLES

        try:
            tipo = driver.find_element(By.XPATH,
                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            print(tipo.text)
            if tipo.text == "Reto Individual":
                title_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]")
                text_title = title_challe.text
                separar = text_title.split()
                if separar[3] == "Individual" and separar[4] == "1":

                    # Guardar el valor del reto en ptos antes de aceptar el reto
                    ptosreto = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[4]")
                    point_challenge = ptosreto.text
                    suma_points = int(point_challenge) + int(valuepoint)
                    # ENTRAR Y ACEPTAR RETO
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    obj_challen = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[2]/android.widget.TextView[1]")
                    text_obj_challen = str(obj_challen.text)
                    pts_obj_challe = text_obj_challen.split()
                    point_obje = pts_obj_challe[0]

                    unirme = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    unirme.click()
                    welcome_challen = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element2 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, str(welcome_challen))))

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET01_Bienvenido al reto.png")

                    acep = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    acep.click()
                    time.sleep(2)
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET01_Inscrito en el reto.png")
                    register_challenge = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element3 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, str(register_challenge))))
                    print("Inscrito en el reto")
                    back = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    back.click()
                    time.sleep(2)
                    home_page = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    home_page.click()
                    time.sleep(1)

                    # modificar fecha de inicio del reto
                    # Obtener token
                    url_auth = "https://apiv2desarrollo.habits.ai/login"
                    payload_auth = {
                        "mail": "lenyn@habits.ai",
                        "pass": "123456789"
                    }
                    response_auth = requests.post(url_auth, data=payload_auth)
                    response_auth_json = response_auth.json()  # Obtener un json de la respuesta
                    token = response_auth_json[
                        'token']  # Hacer un arreglo que guarde el valor de la coincidencia con el texto token
                    bearer = f'Bearer {token}'
                    headers = {'Content-Type': 'application/json', 'Authorization': f'{bearer}'}

                    # Obtener ID de empresa
                    url_id_company = "https://apiv2desarrollo.habits.ai/api/company/?where[name]=E.C.V Licencias"

                    response_id_company = requests.get(url_id_company, headers=headers)
                    rp_id_comp_json = response_id_company.json()
                    data_company = rp_id_comp_json['data']
                    id_company = data_company[0]['_id']

                    time.sleep(1)

                    # Listar retos para la empresa
                    url_challenge = f"https://apiv2desarrollo.habits.ai/api/challenge/?whereObject[company]={id_company}"
                    response_challenge = requests.get(url_challenge, headers=headers)
                    response_auth_challe = response_challenge.json()
                    data = response_auth_challe['data']
                    idchallenge = data[0]['_id']
                    print(f"El id del reto es {idchallenge}")

                    # Modificar fecha de inicio del reto
                    url_date_init = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date = datetime.today()
                    new_date = date + timedelta(days=-1)
                    dateinit = new_date.strftime('%Y-%m-%d')
                    newtime = date.strftime('%H:%M:%S')
                    payload_date = json.dumps({
                        "start_date": f"{dateinit}T{newtime}Z"
                    })
                    response3 = requests.put(url_date_init, headers=headers, data=payload_date)
                    print(f"la fecha de inicio fue modificada exitosamente, estatus: {response3.status_code}")

                    # Sincronizar Data salud
                    time.sleep(2)
                    driver.swipe(453, 1055, 446, 755)

                    sync_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    sync_data.click()

                    time.sleep(1)
                    term_sync = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                    term_sync.click()
                    time.sleep(2)
                    sync_cellphone = driver.find_element(By.XPATH,
                                                         "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]")
                    sync_cellphone.click()
                    time.sleep(2)
                    try:
                        init_sync = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        init_sync.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    try:
                        next_step = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        next_step.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    element3 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.ID, "com.google.android.gms:id/account_picker_container")))
                    # SELECCIONAR USUARIO PARA SINCRONIZAR DATA

                    user_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout")
                    user_data.click()
                    data_salud = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup"
                    element4 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(data_salud))))
                    driver.swipe(342, 1079, 347, 811)

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET01_Data_salud_sincronizada.png")
                    print("data sincronizada")

                    # Agregar data de salud
                    url_user_id = f"https://apiv2desarrollo.habits.ai/api/user/?where[mail]={email.value}"
                    response_user_id = requests.get(url_user_id, headers=headers)
                    response_user_id_json = response_user_id.json()
                    data_user = response_user_id_json['data']
                    id_user = data_user[0]['_id']
                    print(id_user)
                    cardioid = "5f63a3676070b30de1549bf3"
                    pasos = "603ededd3522a74b4107ffb4"
                    sleep = "5f87468222b4c921c00b5286"
                    date2 = datetime.today()
                    dateinit2 = date2.strftime('%Y-%m-%d')
                    time_data = date2 + timedelta(minutes=-2)
                    newtime2 = time_data.strftime('%H:%M:%S')
                    # Consultar segun el reto inscrito que tipo de objetivo es; cardio, pasos o sueño
                    if separar[2] == "Cardio":

                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"

                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{cardioid}",
                            "value": f"{point_obje}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                        if response_data_salud.status_code == "200":
                            print("Data salud cargada exitosamente")
                    elif separar[2] == "Sueño":
                        point_obje = pts_obj_challe[0]
                        text_poin_obj = str(point_obje)
                        sep_poin_obj = text_poin_obj.split(sep=":")
                        value_point_obj = int(sep_poin_obj[0])
                        new_point_obj = value_point_obj * 3600
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{sleep}",
                            "value": f"{new_point_obj}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                        if response_data_salud.status_code == "200":
                            print("Data salud cargada exitosamente")
                    else:
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{pasos}",
                            "value": f"{point_obje}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                        if response_data_salud.status_code == "200":
                            print("Data salud cargada exitosamente para el usuario")

                    # Actualizar data salud nuevamente
                    driver.swipe(380, 244, 375, 1184)
                    time.sleep(1)
                    driver.swipe(410, 324, 408, 786)
                    time.sleep(20)
                    print("data Salud actualizada")

                    # Verificar si se sumoron los puntos de recompensa al score del usaurio

                    score = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    text_score = score.text
                    if int(text_score) == suma_points:
                        print(f"Suma de recompensa en score es correcta {suma_points}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET01_Score con recompensa sumada.png")
                        wb.close()
                        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
                        wb = xl.load_workbook(filesheet)
                        wd = wb.active
                        ws = wb.worksheets[0]

                        receives = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"

                        wb3 = xl.load_workbook(receives)
                        ws3 = wb3.active

                        ws3.insert_rows(1)
                        for i in range(1, 2):
                            for j in range(1, 11):
                                c = ws.cell(row=i, column=j)

                                ws3.cell(row=i, column=j).value = c.value

                        wb3.save(str(receives))

                        wd.delete_rows(1)  # para la fila 1
                        wb.save(filesheet)
                    else:
                        print("Error la suma no es la correcta")
                        print(
                            f"Suma de recompensa en score es INCORRECTA Valor Esperado = {suma_points} Valor obtenido = {text_score}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            f"\\RET01_Score Error en la suma{date2}.png")
                    # VERIFICAR SI EN EL RETO SE MUESTRA LA META CUMPLIDA
                    my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())

                    my_day.click()
                    time.sleep(3)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(5)
                    # VER DETALLE DL RETO
                    into_challen = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView[8]")
                    into_challen.click()
                    time.sleep(1)
                    # PULSAR BOTON DE VER RETO
                    again_into_challe = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    again_into_challe.click()
                    confi_challe = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element5 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(confi_challe))))
                    driver.swipe(385, 1316, 380, 1098)
                    time.sleep(1)
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET01_Reto cumplido.png")
                    time.sleep(1)
                    out_challenge = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    out_challenge.click()
                    time.sleep(1)

                    # FINALIZAR RETO Y COMPROBAR QUE YA NO SE MUESTRE EN PANTALLA
                    # Modificar fecha de inicio
                    url3 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date3 = datetime.today()

                    new_date3 = date3 + timedelta(days=-3)
                    dateinit3 = new_date3.strftime('%Y-%m-%d')
                    newtime3 = date3.strftime('%H:%M:%S')
                    payload_start_finish = json.dumps({
                        "start_date": f"{dateinit3}T{newtime3}Z"
                    })
                    response3 = requests.put(url3, headers=headers, data=payload_start_finish)

                    # Modificar fecha de finalizacion
                    url4 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date4 = datetime.today()
                    new_date4 = date4 + timedelta(days=-3)
                    dateinit4 = new_date4.strftime('%Y-%m-%d')
                    newtime4 = date2.strftime('%H:%M:%S')
                    payload_end_finish = json.dumps({
                        "end_date": f"{dateinit4}T{newtime4}Z"
                    })
                    response4 = requests.put(url4, headers=headers, data=payload_end_finish)
                    print(f"Se cambio la fecha del reto exitosamente, estatus: {response4.status_code}")
                    go_home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    go_home.click()
                    time.sleep(3)
                    go_challenge = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    go_challenge.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(5)
                    challenge_finish = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]"
                    find_challenge = driver.find_element(By.XPATH, str(challenge_finish))
                    text_challen_finish = find_challenge.text
                    if text_challen_finish != text_title:
                        print("Reto se finalizo correctamente")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET01_Reto Finalizado exitosamente.png")

                    else:
                        print('Error no se finalizo el reto')

            else:
                print("No es un reto individual")

        except AssertionError:
            print("No hay retos individuales")

    def test_RET02(self):
        # INSCRIBIRSE EN UN RETO INDIVIDUAL YA INICIALIZADO, CUMPLIR LA META ESTIPULADA PARA OBTENER PUNTOS
        driver = self.driver


        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/SinDataSincronizada.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']

       # iniciar sesion
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)
        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_element = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        # Guardar puntos del score del usuario
        points = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
        valuepoint = points.text

        # Acceder a MI DIA
        enter_element.click()
        time.sleep(5)
        driver.swipe(328, 1034, 328, 823)
        time.sleep(6)

        # VERIFICAR SI LOS RETOS ESTA DISPONIBLES

        try:
            tipo = driver.find_element(By.XPATH,
                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            print(tipo.text)
            if tipo.text == "Reto Individual":
                title_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]")
                text_title = title_challe.text
                separar = text_title.split()
                if separar[3] == "Individual" and separar[4] == "4":

                    # Guardar el valor del reto en ptos antes de aceptar el reto
                    ptosreto = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[4]")
                    point_challenge = ptosreto.text
                    suma_points = int(point_challenge) + int(valuepoint)
                    # ENTRAR Y ACEPTAR RETO
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    obj_challen = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[2]/android.widget.TextView[1]")
                    text_obj_challen = str(obj_challen.text)
                    pts_obj_challe = text_obj_challen.split()
                    point_obje = pts_obj_challe[0]

                    unirme = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    unirme.click()
                    welcome_challen = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element2 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, str(welcome_challen))))

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET01_Bienvenido al reto.png")

                    acep = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    acep.click()
                    time.sleep(2)
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET02_Inscrito en el reto.png")
                    register_challenge = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element3 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, str(register_challenge))))
                    print("Inscrito en el reto")
                    back = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    back.click()
                    time.sleep(2)
                    home_page = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    home_page.click()
                    time.sleep(1)

                    # modificar fecha de inicio del reto
                    # Obtener token
                    url_auth = "https://apiv2desarrollo.habits.ai/login"
                    payload_auth = {
                        "mail": "lenyn@habits.ai",
                        "pass": "123456789"
                    }
                    response_auth = requests.post(url_auth, data=payload_auth)
                    response_auth_json = response_auth.json()  # Obtener un json de la respuesta
                    token = response_auth_json[
                        'token']  # Hacer un arreglo que guarde el valor de la coincidencia con el texto token
                    bearer = f'Bearer {token}'
                    headers = {'Content-Type': 'application/json', 'Authorization': f'{bearer}'}

                    # Obtener ID de empresa
                    url_id_company = "https://apiv2desarrollo.habits.ai/api/company/?where[name]=E.C.V Licencias"

                    response_id_company = requests.get(url_id_company, headers=headers)
                    rp_id_comp_json = response_id_company.json()
                    data_company = rp_id_comp_json['data']
                    id_company = data_company[0]['_id']

                    time.sleep(1)

                    # Listar retos para la empresa
                    url_challenge = f"https://apiv2desarrollo.habits.ai/api/challenge/?whereObject[company]={id_company}"
                    response_challenge = requests.get(url_challenge, headers=headers)
                    response_auth_challe = response_challenge.json()
                    data = response_auth_challe['data']
                    idchallenge = data[3]['_id']
                    print(f"El id del reto es {idchallenge}")

                    # Modificar fecha de inicio del reto
                    url_date_init = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date = datetime.today()
                    new_date = date + timedelta(days=-1)
                    dateinit = new_date.strftime('%Y-%m-%d')
                    newtime = date.strftime('%H:%M:%S')
                    payload_date = json.dumps({
                        "start_date": f"{dateinit}T{newtime}Z"
                    })
                    response3 = requests.put(url_date_init, headers=headers, data=payload_date)

                    print(f"Modificacion de fecha de inicio del reto ejecutado estatus: {response3.status_code}")

                    # Sincronizar Data salud
                    time.sleep(2)
                    driver.swipe(453, 1055, 446, 755)

                    sync_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    sync_data.click()

                    time.sleep(1)
                    term_sync = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                    term_sync.click()
                    time.sleep(2)
                    sync_cellphone = driver.find_element(By.XPATH,
                                                         "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]")
                    sync_cellphone.click()
                    time.sleep(2)
                    try:
                        init_sync = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        init_sync.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    try:
                        next_step = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        next_step.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    element3 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.ID, "com.google.android.gms:id/account_picker_container")))
                    # SELECCIONAR USUARIO PARA SINCRONIZAR DATA

                    user_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout")
                    user_data.click()
                    data_salud = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup"
                    element4 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(data_salud))))
                    driver.swipe(342, 1079, 347, 811)

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET02_Data_salud_sincronizada.png")
                    print("data salud sincronizada")

                    # Agregar data de salud
                    url_user_id = f"https://apiv2desarrollo.habits.ai/api/user/?where[mail]={email.value}"
                    response_user_id = requests.get(url_user_id, headers=headers)
                    response_user_id_json = response_user_id.json()
                    data_user = response_user_id_json['data']
                    id_user = data_user[0]['_id']
                    print(f"el ID de usaurio es: {id_user}")
                    cardioid = "5f63a3676070b30de1549bf3"
                    pasos = "603ededd3522a74b4107ffb4"
                    sleep = "5f87468222b4c921c00b5286"
                    date2 = datetime.today()
                    dateinit2 = date2.strftime('%Y-%m-%d')
                    time_data = date2 + timedelta(minutes=-2)
                    newtime2 = time_data.strftime('%H:%M:%S')
                    # Consultar segun el reto inscrito que tipo de objetivo es; cardio, pasos o sueño
                    if separar[2] == "Cardio":

                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"

                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{cardioid}",
                            "value": f"{point_obje}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                        if response_data_salud.status_code == "200":
                            print("Data salud cargada exitosamente")
                    elif separar[2] == "Sueño":
                        point_obje = pts_obj_challe[0]
                        text_poin_obj = str(point_obje)
                        sep_poin_obj = text_poin_obj.split(sep=":")
                        value_point_obj = int(sep_poin_obj[0])
                        new_point_obj = value_point_obj * 3600
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{sleep}",
                            "value": f"{new_point_obj}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                        if response_data_salud.status_code == "200":
                            print("Data salud cargada exitosamente")
                    else:
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{pasos}",
                            "value": f"{point_obje}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                        if response_data_salud.status_code == "200":
                            print("Data salud cargada exitosamente")

                    # Actualizar data salud nuevamente
                    driver.swipe(380, 244, 375, 1184)
                    time.sleep(1)
                    driver.swipe(410, 324, 408, 786)
                    time.sleep(20)
                    print("data Salud actualizada")

                    # Verificar si se sumoron los puntos de recompensa al score del usaurio

                    score = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    text_score = score.text
                    if int(text_score) == suma_points:
                        print(f"Suma de recompensa en score es correcta {suma_points}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET02_Score con recompensa sumada.png")
                        wb.close()
                        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
                        wb = xl.load_workbook(filesheet)
                        wd = wb.active
                        ws = wb.worksheets[0]

                        receives = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"

                        wb3 = xl.load_workbook(receives)
                        ws3 = wb3.active

                        ws3.insert_rows(1)
                        for i in range(1, 2):
                            for j in range(1, 11):
                                c = ws.cell(row=i, column=j)

                                ws3.cell(row=i, column=j).value = c.value

                        wb3.save(str(receives))

                        wd.delete_rows(1)  # para la fila 1
                        wb.save(filesheet)
                    else:
                        print("Error la suma no es la correcta")
                        print(
                            f"Suma de recompensa en score es INCORRECTA Valor Esperado = {suma_points} Valor obtenido = {text_score}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            f"\\RET02_Score sin suma correcta {date2}.png")
                    # VERIFICAR SI EN EL RETO SE MUESTRA LA META CUMPLIDA
                    my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())

                    my_day.click()
                    time.sleep(3)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(5)
                    # VER DETALLE DL RETO
                    into_challen = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView[8]")
                    into_challen.click()
                    time.sleep(1)
                    # PULSAR BOTON DE VER RETO
                    again_into_challe = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    again_into_challe.click()
                    confi_challe = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element5 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(confi_challe))))
                    driver.swipe(385, 1316, 380, 1098)
                    time.sleep(1)
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET02_Reto cumplido.png")
                    time.sleep(1)
                    out_challenge = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    out_challenge.click()
                    time.sleep(1)
                    # FINALIZAR RETO Y COMPROBAR QUE YA NO SE MUESTRE EN PANTALLA
                    # Modificar fecha de inicio
                    url3 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date3 = datetime.today()

                    new_date3 = date3 + timedelta(days=-3)
                    dateinit3 = new_date3.strftime('%Y-%m-%d')
                    newtime3 = date3.strftime('%H:%M:%S')
                    payload_start_finish = json.dumps({
                        "start_date": f"{dateinit3}T{newtime3}Z"
                    })
                    response3 = requests.put(url3, headers=headers, data=payload_start_finish)
                    # Modificar fecha de finalizacion
                    url4 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date4 = datetime.today()
                    new_date4 = date4 + timedelta(days=-3)
                    dateinit4 = new_date4.strftime('%Y-%m-%d')
                    newtime4 = date2.strftime('%H:%M:%S')
                    payload_end_finish = json.dumps({
                        "end_date": f"{dateinit4}T{newtime4}Z"
                    })
                    response4 = requests.put(url4, headers=headers, data=payload_end_finish)
                    print(f"Se cambio la fecha del reto correctamente estatus: {response4.status_code}")

                    go_home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    go_home.click()
                    time.sleep(3)
                    go_challenge = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    go_challenge.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(5)
                    challenge_finish = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]"
                    find_challenge = driver.find_element(By.XPATH, str(challenge_finish))
                    text_challen_finish = find_challenge.text
                    if text_challen_finish != text_title:
                        print("Reto se finalizo correctamente")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET02_Reto Finalizado exitosamente.png")

                    else:
                        print('Error no se finalizo el reto')

            else:
                print("No es un reto individual")

        except AssertionError:
            print("No hay retos individuales")

    def test_RET03(self):
        # RETO GRUPAL  4 participantes 2 usuarios dentro de un equipo, 1 usuario como incripcion individual y uno aleatorio
        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        name = datos['A1']
        passw = datos['C1']
        email_secund = datos['B2']
        name_secund = datos['A2']
        passw_secund = datos['C2']
        name_3 = datos['A3']
        email_3 = datos['B3']
        name_4 = datos['A4']
        email_4 = datos['B4']

       # Iniciar sesion como usuario que se registra de forma individual
        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)
        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        element = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        # Guardar puntos del score del usuario
        points = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
        valuepoint = points.text

        # Acceder a MI DIA
        enter = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())

        enter.click()
        time.sleep(5)
        driver.swipe(328, 1034, 328, 823)
        time.sleep(6)

        # Verificar si hay retos disponibles

        try:
            tipo = driver.find_element(By.XPATH,
                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            print(tipo.text)
            if tipo.text == "Reto en Equipo":
                title_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]")
                text_title = title_challe.text
                separar = text_title.split()
                if separar[3] == "Grupal" and separar[4] == "1":

                    # Guardar el valor del reto en ptos antes de aceptar el reto
                    ptosreto = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[4]")
                    point_challenge = ptosreto.text
                    suma_points = int(point_challenge) + int(valuepoint)
                    suma_points_sec = 0
                    # ENTRAR AL DETALLE DEL RETO
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    try:
                        # Esperar infografia de retos
                        modal_info = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView"
                        element0 = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, str(modal_info))))
                        no_see_more = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[4]")
                        no_see_more.click()
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Cerrar infografia.png")
                        time.sleep(0.5)
                        close_modal = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView")
                        close_modal.click()
                    except:
                        print("No hay Modal de infografia")

                    # REGISTRARSE COMO USUARIO INDIVIDUAL
                    obj_challen = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[2]/android.widget.TextView[1]")
                    text_obj_challen = str(obj_challen.text)
                    pts_obj_challe = text_obj_challen.split()
                    point_obje = pts_obj_challe[0]

                    buttom_individual = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView")
                    buttom_individual.click()
                    time.sleep(1)

                    try:  # MODAL DE CONFIRMACION DE REGISTRO INDIVIDUAL
                        unirme_challen = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")))

                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Confirmacion de incripcion individual.png")
                        unirme_challen.click()

                    except:
                        print("ERROR NO SE MOSTRO LA MODAL DE CONFIRMACION DE INCRIPCION INDIVIDUAL")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_ERROR NO MOSTRO MODAL DE CONFIRMACION DE"
                            " incripcion individual.png")
                        time.sleep(2)
                    try:  # Modal de Registro Individual exitoso
                        aceptar_modal_bienvenida = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")))

                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Inscrito en el reto.png")
                        print("Inscrito en el reto")
                        aceptar_modal_bienvenida.click()
                        time.sleep(2)
                    except:
                        print("Error no se mostro la modal de registro exitoso")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_No se mostro modal de registro exitoso.png")
                        time.sleep(3)

                    # CERRAR SESION E INICIAR CON USUARIO QUE CREARA UN EQUIPO DE 2 INTEGRANTES
                    dashboard = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.home_navbar())))
                    dashboard.click()
                    time.sleep(2)
                    menu_burger = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]")
                    menu_burger.click()
                    time.sleep(2)
                    for i in range(1, 3):
                        driver.swipe(30, 1292, 38, 652)
                        time.sleep(0.5)
                    log_out = driver.find_element(By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    log_out.click()

                    # Iniciar Sesion Con usuario que creara grupo de 2 integrantes

                    time.sleep(6)
                    LoginAndRegisterFlowsUtils.LoginUser(email_secund.value, passw_secund.value, driver)

                    # Alerta de calificacion de aplicacion
                    try:
                        skip = WebDriverWait(driver, 8). \
                            until(
                            EC.presence_of_element_located(
                                (By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
                        # Omitir calificacion
                        skip.click()
                        print("Se Omitio la modal de calificar aplicacion en la tienda")
                    except:
                        print("No hay notificacion de calificacion de aplicacion")

                    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                              ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
                    assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
                    print("Se inicio sesion correctamente")
                    element = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
                    # Guardar puntos del score del usuario
                    points_secund = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    valuepoint_sec = points_secund.text

                    # Acceder a MI DIA
                    enter = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    enter.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(6)

                    tipo_sec = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    print(tipo_sec.text)
                    if tipo_sec.text == "Reto en Equipo":
                        title_challe_sec = driver.find_element(By.XPATH,
                                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]")
                        text_title_sec = title_challe_sec.text
                        separar = text_title_sec.split()
                        if separar[3] == "Grupal" and separar[4] == "1":

                            # Guardar la sumatoria de puntos que deberia quedar el usaurio al culminar el reto

                            suma_points_sec = int(point_challenge) + int(valuepoint_sec)
                            # ENTRAR AL DETALLE DL RETO
                            in_challenge = driver.find_element(By.XPATH,
                                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                            in_challenge.click()
                            try:
                                modal_info = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView"
                                element0 = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, str(modal_info))))
                                no_see_more = driver.find_element(By.XPATH,
                                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[4]")
                                no_see_more.click()
                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET03_Cerrar infografia.png")
                                time.sleep(0.5)
                                close_modal = driver.find_element(By.XPATH,
                                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView")
                                close_modal.click()
                            except:
                                print("No hay Modal de infografia")

                            # REGISTRARSE COMO EQUIPO

                            team_challeng = driver.find_element(By.XPATH,
                                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.TextView")
                            team_challeng.click()
                            time.sleep(1)
                            # MODAL DE CONFIRMACION DE REGISTRO EN EQUIPO
                            try:
                                modal_team = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup"
                                element2 = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, str(modal_team))))

                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET03_Confirmacion de registro de equipo.png")

                                team_registro = driver.find_element(By.XPATH,
                                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                                team_registro.click()
                                time.sleep(2)
                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET03_Listado de usaurios para enviar invitacion.png")

                            except:
                                print("ERROR NO SE MOSTRO LA MODAL DE CONFIRMACION DE INCRIPCION DE EQUIPO")
                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET03_ERROR NO MOSTRO MODAL DE CONFIRMACION DE"
                                    " incripcion individual.png")
                            # INVITAR USUARIOS

                            time.sleep(2)
                            search_box = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
                            search_box.send_keys(name_3.value)
                            time.sleep(2)
                            send_invi = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                            send_invi.click()
                            print("Usuario A invitado")
                            time.sleep(1)

                            search_box.clear()
                            time.sleep(2)
                            search_box.send_keys(name_4.value)
                            time.sleep(2)
                            send_invi_two = driver.find_element(By.XPATH,
                                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                            send_invi_two.click()
                            print("Usuario B Invitado")
                            time.sleep(1)

                            # CREAR EQUIPO

                            next_step = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                            next_step.click()
                            time.sleep(0.5)
                            upload_img = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ImageView")
                            upload_img.click()
                            time.sleep(1)
                            try:
                                accep_permi = driver.find_element(By.XPATH,
                                                                  "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")
                                accep_permi.click()
                            except:
                                print("Ya se otrogaron permisos de galeria")
                            pictures = driver.find_element(By.XPATH,
                                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.drawerlayout.widget.DrawerLayout/android.view.ViewGroup/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.HorizontalScrollView/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.CompoundButton[1]")
                            pictures.click()
                            time.sleep(0.5)
                            select_pic = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.drawerlayout.widget.DrawerLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.FrameLayout[2]/android.widget.LinearLayout/android.view.ViewGroup/androidx.recyclerview.widget.RecyclerView/androidx.cardview.widget.CardView[1]/androidx.cardview.widget.CardView/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.ImageView[1]")
                            select_pic.click()
                            name_team = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.EditText")
                            name_team.send_keys("Team RT-03")
                            create_team = driver.find_element(By.XPATH,
                                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.widget.TextView")
                            create_team.click()
                            challen_view = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[1]"
                            element2 = WebDriverWait(driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, str(challen_view))))
                            home_challe = driver.find_element(By.XPATH,
                                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView")
                            home_challe.click()
                            time.sleep(3)
                            dashboard = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                            dashboard.click()
                        else:
                            print("No esta disponible el reto para el usaurio 2")
                    else:
                        print("No hay retos en grupo disponibles")

                    # Obtener token
                    url_auth = "https://apiv2desarrollo.habits.ai/login"
                    payload_auth = {
                        "mail": "lenyn3@habits.ai",
                        "pass": "123456789"
                    }
                    response_auth = requests.post(url_auth, data=payload_auth)
                    response_auth_json = response_auth.json()  # Obtener un json de la respuesta
                    token = response_auth_json[
                        'token']  # Hacer un arreglo que guarde el valor de la coincidencia con el texto token
                    bearer = f'Bearer {token}'
                    headers = {'Content-Type': 'application/json', 'Authorization': f'{bearer}'}

                    # Obtener ID de empresa
                    url_id_company = "https://apiv2desarrollo.habits.ai/api/company/?where[name]=E.C.V Licencias"

                    response_id_company = requests.get(url_id_company, headers=headers)
                    rp_id_comp_json = response_id_company.json()
                    data_company = rp_id_comp_json['data']
                    id_company = data_company[0]['_id']

                    time.sleep(1)

                    # Listar retos para la empresa

                    url_challenge = f"https://apiv2desarrollo.habits.ai/api/challenge/?whereObject[company]={id_company}"
                    response_challenge = requests.get(url_challenge, headers=headers)
                    response_auth_challe = response_challenge.json()
                    data = response_auth_challe['data']
                    idchallenge = data[4]['_id']

                    time.sleep(1)

                    # PROCESAR CRON DE WAITING ROOM
                    url_waiting_room = f"https://apiv2desarrollo.habits.ai/api/challenge/processWaitingRoom/{idchallenge}"

                    response_waiting_room = requests.post(url_waiting_room, headers=headers)
                    print(response_waiting_room.status_code)
                    time.sleep(3)

                    # CONFIRMAR CREACION D EQUIPO CORRECTAMENTE
                    enter_my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    enter_my_day.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(6)
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    time.sleep(2)
                    see_teams = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.widget.TextView")
                    see_teams.click()
                    time.sleep(2)
                    admin = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[1]").text
                    user_invi_3 = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.TextView[1]").text
                    user_invi_4 = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView[1]").text
                    user_waitingroom = driver.find_element(By.XPATH,
                                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup/android.widget.TextView[1]").text
                    list_users = [admin, user_invi_3, user_invi_4, user_waitingroom]
                    try:
                        if str(name_secund.value) in list_users:
                            print("El admin es el correcto en el equipo")
                            if str(name.value) in list_users:
                                print("El usaurio en waiting room se anexo correctamente al equipo")
                                if str(name_3.value) in list_users:
                                    print(f"Usuario {name_3.value} Se agrego correctamente al equipo")
                                    if str(name_4.value) in list_users:
                                        print(f"El usaurio {name_4.value} se anexo correctamente al equipo")
                                        print("El equipo se creo correctamente")
                                        driver.get_screenshot_as_file(
                                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                            "\\RET03_ Se creo el equipo correctamente")
                    except:
                        print("Error no es el equipo correcto")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_No es el equipo correcto ERROR.png")
                    out_chall = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView")
                    out_chall.click()
                    time.sleep(2)
                    dash_home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    dash_home.click()
                    time.sleep(2)
                    # Modificar fecha de inicio del reto

                    url_date_init = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date = datetime.today()
                    new_date = date + timedelta(days=-1)
                    dateinit = new_date.strftime('%Y-%m-%d')
                    newtime = date.strftime('%H:%M:%S')
                    payload_date = json.dumps({
                        "start_date": f"{dateinit}T{newtime}Z"
                    })
                    response3 = requests.put(url_date_init, headers=headers, data=payload_date)

                    time.sleep(0.5)

                    # Sincronizar Data salud

                    time.sleep(2)
                    driver.swipe(453, 1055, 446, 755)

                    sync_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    sync_data.click()

                    time.sleep(1)
                    term_sync = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                    term_sync.click()
                    time.sleep(2)
                    sync_cellphone = driver.find_element(By.XPATH,
                                                         "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]")
                    sync_cellphone.click()
                    time.sleep(2)
                    try:
                        init_sync = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        init_sync.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    try:
                        next_step2 = driver.find_element(By.XPATH,
                                                         "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        next_step2.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    element3 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.ID, "com.google.android.gms:id/account_picker_container")))
                    # SELECCIONAR USUARIO PARA SINCRONIZAR DATA

                    user_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout")
                    user_data.click()
                    data_salud = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup"
                    element4 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(data_salud))))
                    driver.swipe(342, 1079, 347, 811)

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET03_Data_salud_sincronizada.png")
                    print("data sincronizada")

                    # Agregar data de salud Usuario Que creo el equipo
                    url_user_id = f"https://apiv2desarrollo.habits.ai/api/user/?where[mail]={email_secund.value}"
                    response_user_id = requests.get(url_user_id, headers=headers)
                    response_user_id_json = response_user_id.json()
                    data_user = response_user_id_json['data']
                    id_user = data_user[0]['_id']
                    print(id_user)
                    cardioid = "5f63a3676070b30de1549bf3"
                    pasos = "603ededd3522a74b4107ffb4"
                    sleep = "5f87468222b4c921c00b5286"
                    date2 = datetime.today()
                    dateinit2 = date2.strftime('%Y-%m-%d')
                    time_data = date2 + timedelta(minutes=-2)
                    newtime2 = time_data.strftime('%H:%M:%S')
                    if separar[2] == "Cardio":

                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        cardio_point = int(point_obje) * 4
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{cardioid}",
                            "value": f"{cardio_point}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                        status = response_data_salud.status_code
                        if status == "200":
                            print("Data salud cargada exitosamente")
                    elif separar[2] == "Sueño":
                        point_obje = pts_obj_challe[0]
                        text_poin_obj = str(point_obje)
                        sep_poin_obj = text_poin_obj.split(sep=":")
                        value_point_obj = int(sep_poin_obj[0])
                        new_point_obj = value_point_obj * 3600
                        sleep_point = int(new_point_obj) * 4
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{sleep}",
                            "value": f"{sleep_point}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                        status = response_data_salud.status_code
                        if status == "200":
                            print("Data salud cargada exitosamente")
                    else:
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        step_point = int(point_obje) * 4
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{pasos}",
                            "value": f"{step_point}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                        status = response_data_salud.status_code
                        if status == "200":
                            print("Data salud cargada exitosamente")
                        time.sleep(0.5)

                    # Actualizar data salud nuevamente
                    driver.swipe(380, 244, 375, 1184)
                    time.sleep(1)
                    driver.swipe(410, 324, 408, 786)
                    time.sleep(20)
                    print("data Salud actualizada")

                    # Verificar si se sumoron los puntos de recompensa al score del usaurio 2 que creo el equipo
                    time.sleep(2)
                    score = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    text_score = score.text
                    if int(text_score) == int(suma_points_sec):
                        print(
                            f"Suma de recompensa en score es correcta para el usuario 2 total de puntos = {suma_points_sec}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Score con recompensa sumada correctamente para usaurio 2.png")
                    else:
                        print("Error la suma no es la correcta")
                        print(
                            f"Suma de recompensa en score es INCORRECTA Valor Esperado = {suma_points_sec} Valor obtenido = {text_score}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            f"\\RET03_Score con recompensa sumada correctamente para usaurio 2 {date2}.png")
                    # Verificar si se sumaron lso puntos correctos para el usaurio 1 que se incribio de forma individual
                    # CERRAR SESION E INICIAR CON USUARIO QUE SE INSCRIBIO D FORMA INDIVIDUAL
                    menu_burger = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    menu_burger.click()
                    time.sleep(1)
                    for i in range(1, 3):
                        driver.swipe(286, 1360, 267, 636)
                    log_out = driver.find_element(By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    log_out.click()
                    time.sleep(2)
                    # Iniciar Sesion Con Usuario que se registro individualmente

                    LoginAndRegisterFlowsUtils.LoginUser(email.value, passw.value, driver)
                    # Verificar si se sumoron los puntos de recompensa al score del usaurio 2 que creo el equipo

                    score_sec = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    text_score_sec = score_sec.text
                    if int(text_score_sec) == suma_points:
                        print(
                            f"Suma de recompensa en score es correcta para el usuario en waiting room = {suma_points}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Score con recompensa sumada correctamente para usaurio en waiting room.png")
                    else:
                        print("Error la suma no es la correcta")
                        print(
                            f"Suma de recompensa en score es INCORRECTA Valor Esperado = {suma_points} Valor obtenido = {text_score_sec}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            f"\\RET03_Score con recompensa sumada correctamente para usaurio en waiting room {date2}.png")

                    # VERIFICAR SI EN EL RETO SE MUESTRA LA META CUMPLIDA
                    my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())

                    my_day.click()
                    time.sleep(3)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(5)
                    # VER DETALLE DL RETO
                    into_challen = driver.find_element(By.XPATH,
                                                       "hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    into_challen.click()
                    time.sleep(1)
                    # PULSAR BOTON DE VER RETO
                    again_into_challe = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    again_into_challe.click()
                    time.sleep(2)
                    ver_progreso = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                    ver_progreso.click()
                    time.sleep(1)
                    promedio = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.TextView[3]").text
                    objetivo = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.TextView[4]").text
                    promedio_value = int(promedio)
                    objetivo_value = int(objetivo)
                    if promedio_value >= objetivo_value:
                        print("Reto cumplido")

                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03 Progreso de reto grupal.png")
                        wb.close()
                        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
                        wb = xl.load_workbook(filesheet)
                        wd = wb.active
                        ws = wb.worksheets[0]

                        receives = f"{PathAndVariablesUtil.db_path()}/recursos/UsuariosBeneficios.xlsx"

                        wb3 = xl.load_workbook(receives)
                        ws3 = wb3.active

                        ws3.insert_rows(1)
                        for i in range(1, 2):
                            for j in range(1, 11):
                                c = ws.cell(row=i, column=j)

                                ws3.cell(row=i, column=j).value = c.value

                        wb3.save(str(receives))

                        wd.delete_rows(1)  # para la fila 1
                        wb.save(filesheet)
                    else:
                        print("ERROR no se cumplio el objetivo del reto grupal")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03 No se cumplio el objetivo de reto grupal ERROR.png")
                    time.sleep(1)
                    out_challenge = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    out_challenge.click()
                    time.sleep(1)
                    go_home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    go_home.click()
                    # FINALIZAR RETO Y COMPROBAR QUE YA NO SE MUESTRE EN PANTALLA
                    # Modificar fecha de inicio para finalizar el reto
                    url3 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date3 = datetime.today()

                    new_date3 = date3 + timedelta(days=-3)
                    dateinit3 = new_date3.strftime('%Y-%m-%d')
                    newtime3 = date3.strftime('%H:%M:%S')
                    payload_start_finish = json.dumps({
                        "start_date": f"{dateinit3}T{newtime3}Z"
                    })
                    response3 = requests.put(url3, headers=headers, data=payload_start_finish)
                    # Modificar fecha de finalizacion
                    url4 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date4 = datetime.today()
                    new_date4 = date4 + timedelta(days=-3)
                    dateinit4 = new_date4.strftime('%Y-%m-%d')
                    newtime4 = date2.strftime('%H:%M:%S')
                    payload_end_finish = json.dumps({
                        "end_date": f"{dateinit4}T{newtime4}Z"
                    })
                    response4 = requests.put(url4, headers=headers, data=payload_end_finish)
                    print("Se cambio la fecha del reto")
                    print(response4.status_code)

                    time.sleep(1)
                    go_challenge = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    go_challenge.click()
                    time.sleep(3)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(3)
                    challenge_finish = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]"
                    find_challenge = driver.find_element(By.XPATH, str(challenge_finish))
                    text_challen_finish = find_challenge.text
                    if text_challen_finish != text_title:
                        print("Reto se finalizo correctamente")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Reto Finalizado exitosamente.png")

                    else:
                        print('Error no se finalizo el reto')
            else:
                print("No es un reto Grupal")

        except AssertionError:
            print("No hay retos Grupales")

    def test_RET04(self):
        # RETO GRUPAL 4 participantes 1 usuario Se inscribe creando un equipo completo,
        # 1 usuario como incripcion individual y donde se registraran 2 equipos uno aleatorio
        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        name_waitingroom = datos['A1']
        passw = datos['C1']
        email_secund = datos['B2']
        name_secund = datos['A2']
        passw_secund = datos['C2']
        name_3 = datos['A3']
        name_4 = datos['A4']
        name_5 = datos['A5']

        # Iniciar sesion como usaurio individual

        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_element = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        # Guardar puntos del score del usuario que estara en waiting room
        points = driver.find_element(By.XPATH,
                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
        valuepoint = points.text

        # Acceder a MI DIA
        enter_element.click()
        time.sleep(5)
        driver.swipe(328, 1034, 328, 823)
        time.sleep(6)

        # Verificar si hay retos disponibles

        try:
            tipo = driver.find_element(By.XPATH,
                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            print(tipo.text)
            if tipo.text == "Reto en Equipo":
                title_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]")
                text_title = title_challe.text
                separar = text_title.split()
                if separar[3] == "Grupal" and separar[4] == "1":

                    # Guardar el valor del reto en ptos antes de aceptar el reto
                    ptosreto = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[4]")
                    point_challenge = ptosreto.text
                    suma_points = int(point_challenge) + int(valuepoint)
                    suma_points_sec = 0
                    # ENTRAR AL DETALLE DEL RETO
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    try:
                        modal_info = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView"
                        element0 = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, str(modal_info))))
                        no_see_more = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[4]")
                        no_see_more.click()
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET04_Cerrar infografia.png")
                        time.sleep(0.5)
                        close_modal = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView")
                        close_modal.click()
                    except:
                        print("No hay Modal de infografia")

                    # REGISTRARSE COMO USUARIO INDIVIDUAL
                    obj_challen = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[2]/android.widget.TextView[1]")
                    text_obj_challen = str(obj_challen.text)
                    pts_obj_challe = text_obj_challen.split()
                    point_obje = pts_obj_challe[0]

                    buttom_individual = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView")
                    buttom_individual.click()
                    time.sleep(1)

                    try:  # MODAL DE CONFIRMACION DE REGISTRO INDIVIDUAL
                        unirme_challen = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")))

                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Confirmacion de incripcion individual.png")
                        unirme_challen.click()

                    except:
                        print("ERROR NO SE MOSTRO LA MODAL DE CONFIRMACION DE INCRIPCION INDIVIDUAL")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_ERROR NO MOSTRO MODAL DE CONFIRMACION DE"
                            " incripcion individual.png")
                        time.sleep(2)
                    try:  # Modal de Registro Individual exitoso
                        aceptar_modal_bienvenida = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")))

                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_Inscrito en el reto.png")
                        print("Inscrito en el reto")
                        aceptar_modal_bienvenida.click()
                        time.sleep(2)
                    except:
                        print("Error no se mostro la modal de registro exitoso")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET03_No se mostro modal de registro exitoso.png")

                    # CERRAR SESION E INICIAR CON USUARIO QUE CREARA UN EQUIPO DE 2 INTEGRANTES
                    dashboard = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.home_navbar())))
                    dashboard.click()
                    time.sleep(2)
                    menu_burger = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]")
                    menu_burger.click()
                    time.sleep(2)
                    for i in range(1, 3):
                        driver.swipe(19, 800, 19, 244)
                        time.sleep(1)
                    log_out = driver.find_element(By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    log_out.click()

                    # Iniciar Sesion Con usuario que creara grupo de 2 integrantes

                    time.sleep(6)
                    LoginAndRegisterFlowsUtils.LoginUser(email_secund.value, passw_secund.value, driver)

                    # Alerta de calificacion de aplicacion
                    try:
                        skip = WebDriverWait(driver, 8). \
                            until(
                            EC.presence_of_element_located(
                                (By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
                        # Omitir calificacion
                        skip.click()
                        print("Se Omitio la modal de calificar aplicacion en la tienda")
                    except:
                        print("No hay notificacion de calificacion de aplicacion")

                    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                              ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
                    assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
                    print("Se inicio sesion correctamente")
                    enter_element = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
                    # Guardar puntos del score del usuario
                    points_secund = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    valuepoint_sec = points_secund.text

                    # Acceder a MI DIA
                    enter_element.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(6)

                    tipo_sec = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    print(tipo_sec.text)
                    if tipo_sec.text == "Reto en Equipo":
                        title_challe_sec = driver.find_element(By.XPATH,
                                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]")
                        text_title_sec = title_challe_sec.text
                        separar = text_title_sec.split()
                        if separar[3] == "Grupal" and separar[4] == "1":

                            # Guardar la sumatoria de puntos que deberia quedar el usaurio al culminar el reto

                            suma_points_sec = int(point_challenge) + int(valuepoint_sec)
                            # ENTRAR AL DETALLE DL RETO
                            in_challenge = driver.find_element(By.XPATH,
                                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                            in_challenge.click()
                            try:
                                modal_info = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView"
                                element0 = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, str(modal_info))))
                                no_see_more = driver.find_element(By.XPATH,
                                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[4]")
                                no_see_more.click()
                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET04_Cerrar infografia.png")
                                time.sleep(0.5)
                                close_modal = driver.find_element(By.XPATH,
                                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[5]/android.widget.TextView")
                                close_modal.click()
                            except:
                                print("No hay Modal de infografia")

                            # REGISTRARSE COMO EQUIPO

                            team_challeng = driver.find_element(By.XPATH,
                                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.TextView")
                            team_challeng.click()
                            time.sleep(1)
                            try:
                                modal_team = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup"
                                element2 = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, str(modal_team))))

                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET04_Confirmacion de registro de equipo.png")

                                team_registro = driver.find_element(By.XPATH,
                                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                                team_registro.click()
                                time.sleep(2)
                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET04_Listado de usaurios para enviar invitacion.png")

                            except:
                                print("ERROR NO SE MOSTRO LA MODAL DE CONFIRMACION DE INCRIPCION DE EQUIPO")
                                driver.get_screenshot_as_file(
                                    f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                    "\\RET04_ERROR NO MOSTRO MODAL DE CONFIRMACION DE"
                                    " incripcion individual.png")
                            # INVITAR USUARIOS

                            time.sleep(2)
                            search_box = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.EditText")
                            search_box.send_keys(name_3.value)
                            time.sleep(2)
                            send_invi = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                            send_invi.click()
                            print("Usuario A invitado")
                            time.sleep(1)

                            search_box.clear()
                            time.sleep(2)
                            search_box.send_keys(name_4.value)
                            time.sleep(2)
                            send_invi_two = driver.find_element(By.XPATH,
                                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                            send_invi_two.click()
                            print("Usuario B Invitado")
                            time.sleep(1)

                            search_box.clear()
                            time.sleep(2)
                            search_box.send_keys(name_5.value)
                            time.sleep(2)
                            send_invi_two = driver.find_element(By.XPATH,
                                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.TextView")
                            send_invi_two.click()
                            print("Usuario C Invitado")

                            # CREAR EQUIPO

                            next_step = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                            next_step.click()
                            time.sleep(0.5)
                            upload_img = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ImageView")
                            upload_img.click()
                            time.sleep(1)
                            try:
                                accep_permi = driver.find_element(By.XPATH,
                                                                  "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.Button[1]")
                                accep_permi.click()
                            except:
                                print("Ya se otrogaron permisos de galeria")
                            pictures = driver.find_element(By.XPATH,
                                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.drawerlayout.widget.DrawerLayout/android.view.ViewGroup/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.HorizontalScrollView/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.CompoundButton[1]")
                            pictures.click()
                            time.sleep(0.5)
                            select_pic = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.drawerlayout.widget.DrawerLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.FrameLayout[2]/android.widget.LinearLayout/android.view.ViewGroup/androidx.recyclerview.widget.RecyclerView/androidx.cardview.widget.CardView[1]/androidx.cardview.widget.CardView/android.widget.RelativeLayout/android.widget.FrameLayout[1]/android.widget.ImageView[1]")
                            select_pic.click()
                            name_team = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.EditText")
                            name_team.send_keys("Team RT-03")
                            create_team = driver.find_element(By.XPATH,
                                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.widget.TextView")
                            create_team.click()
                            challen_view = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[1]"
                            element2 = WebDriverWait(driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, str(challen_view))))
                            home_challe = driver.find_element(By.XPATH,
                                                              "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView")
                            home_challe.click()
                            time.sleep(3)
                            dashboard = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                            dashboard.click()
                        else:
                            print("No esta disponible el reto para el usaurio 2")
                    else:
                        print("No hay retos en grupo disponibles")

                    # Obtener token
                    url_auth = "https://apiv2desarrollo.habits.ai/login"
                    payload_auth = {
                        "mail": "lenyn3@habits.ai",
                        "pass": "123456789"
                    }
                    response_auth = requests.post(url_auth, data=payload_auth)
                    response_auth_json = response_auth.json()  # Obtener un json de la respuesta
                    token = response_auth_json[
                        'token']  # Hacer un arreglo que guarde el valor de la coincidencia con el texto token
                    bearer = f'Bearer {token}'
                    headers = {'Content-Type': 'application/json', 'Authorization': f'{bearer}'}

                    # Obtener ID de empresa
                    url_id_company = "https://apiv2desarrollo.habits.ai/api/company/?where[name]=E.C.V Licencias"

                    response_id_company = requests.get(url_id_company, headers=headers)
                    rp_id_comp_json = response_id_company.json()
                    data_company = rp_id_comp_json['data']
                    id_company = data_company[0]['_id']

                    time.sleep(1)

                    # Listar retos para la empresa

                    url_challenge = f"https://apiv2desarrollo.habits.ai/api/challenge/?whereObject[company]={id_company}"
                    response_challenge = requests.get(url_challenge, headers=headers)
                    response_auth_challe = response_challenge.json()
                    data = response_auth_challe['data']
                    idchallenge = data[4]['_id']

                    time.sleep(1)

                    # PROCESAR CRON DE WAITING ROOM
                    url_waiting_room = f"https://apiv2desarrollo.habits.ai/api/challenge/processWaitingRoom/{idchallenge}"

                    response_waiting_room = requests.post(url_waiting_room, headers=headers)
                    print(response_waiting_room.status_code)
                    time.sleep(3)

                    # CONFIRMAR CREACION D EQUIPO CORRECTAMENTE
                    enter_my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    enter_my_day.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(6)
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    time.sleep(2)
                    see_teams = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.widget.TextView")
                    see_teams.click()
                    time.sleep(2)
                    admin = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[1]").text
                    user_invi_3 = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.TextView[1]").text
                    user_invi_4 = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView[1]").text
                    user_invi_5 = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup/android.widget.TextView[1]").text
                    list_users = [admin, user_invi_3, user_invi_4, user_invi_5]
                    try:
                        if str(name_secund.value) in list_users:
                            print("El admin es el correcto en el equipo")
                            if str(name_3.value) in list_users:
                                print(f"Usuario {name_3.value} Se agrego correctamente al equipo")
                                if str(name_4.value) in list_users:
                                    print(f"El usaurio {name_4.value} se anexo correctamente al equipo")
                                    print("El equipo se creo correctamente")
                                    if str(name_5.value) in list_users:
                                        print(f"El usaurio {name_5.value} se anexo correctamente al equipo")
                                        print("El equipo se creo correctamente")
                                        driver.get_screenshot_as_file(
                                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                            "\\RET04_ Se creo el equipo correctamente")
                    except:
                        print("Error no es el equipo correcto")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET04_No es el equipo correcto ERROR.png")
                    out_chall = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView")
                    out_chall.click()
                    time.sleep(2)
                    dash_home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    dash_home.click()
                    time.sleep(2)
                    # Modificar fecha de inicio del reto

                    url_date_init = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date = datetime.today()
                    new_date = date + timedelta(days=-1)
                    dateinit = new_date.strftime('%Y-%m-%d')
                    newtime = date.strftime('%H:%M:%S')
                    payload_date = json.dumps({
                        "start_date": f"{dateinit}T{newtime}Z"
                    })
                    response3 = requests.put(url_date_init, headers=headers, data=payload_date)

                    time.sleep(0.5)

                    # Sincronizar Data salud

                    time.sleep(2)
                    driver.swipe(453, 1055, 446, 755)

                    sync_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    sync_data.click()

                    time.sleep(1)
                    term_sync = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                    term_sync.click()
                    time.sleep(2)
                    sync_cellphone = driver.find_element(By.XPATH,
                                                         "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]")
                    sync_cellphone.click()
                    time.sleep(2)
                    try:
                        init_sync = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        init_sync.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    try:
                        next_step2 = driver.find_element(By.XPATH,
                                                         "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                        next_step2.click()
                        time.sleep(2)
                    except:
                        print("terminos aceptados para sincronizar")
                    element3 = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.ID, "com.google.android.gms:id/account_picker_container")))
                    # SELECCIONAR USUARIO PARA SINCRONIZAR DATA

                    user_data = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout")
                    user_data.click()
                    data_salud = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup"
                    element4 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(data_salud))))
                    driver.swipe(342, 1079, 347, 811)

                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET04_Data_salud_sincronizada.png")
                    print("data sincronizada")

                    # Agregar data de salud Usuario Que creo el equipo
                    url_user_id = f"https://apiv2desarrollo.habits.ai/api/user/?where[mail]={email_secund.value}"
                    response_user_id = requests.get(url_user_id, headers=headers)
                    response_user_id_json = response_user_id.json()
                    data_user = response_user_id_json['data']
                    id_user = data_user[0]['_id']
                    print(id_user)
                    cardioid = "5f63a3676070b30de1549bf3"
                    pasos = "603ededd3522a74b4107ffb4"
                    sleep = "5f87468222b4c921c00b5286"
                    date2 = datetime.today()
                    dateinit2 = date2.strftime('%Y-%m-%d')
                    time_data = date2 + timedelta(minutes=-2)
                    newtime2 = time_data.strftime('%H:%M:%S')
                    if separar[2] == "Cardio":

                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        cardio_point = int(point_obje) * 4
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{cardioid}",
                            "value": f"{cardio_point}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                        status = response_data_salud.status_code
                        if status == "200":
                            print("Data salud cargada exitosamente")
                    elif separar[2] == "Sueño":
                        point_obje = pts_obj_challe[0]
                        text_poin_obj = str(point_obje)
                        sep_poin_obj = text_poin_obj.split(sep=":")
                        value_point_obj = int(sep_poin_obj[0])
                        new_point_obj = value_point_obj * 3600
                        sleep_point = int(new_point_obj) * 4
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{sleep}",
                            "value": f"{sleep_point}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                        status = response_data_salud.status_code
                        if status == "200":
                            print("Data salud cargada exitosamente")
                    else:
                        url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                        step_point = int(point_obje) * 4
                        payload_heal = json.dumps({
                            "date": f"{dateinit2} {newtime2}",
                            "kind": f"{pasos}",
                            "value": f"{step_point}",
                            "user": f"{id_user}",
                            "timeZone": "America/Caracas"
                        })
                        response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                        status = response_data_salud.status_code
                        if status == "200":
                            print("Data salud cargada exitosamente")
                        time.sleep(0.5)

                    # Actualizar data salud nuevamente
                    driver.swipe(380, 244, 375, 1184)
                    time.sleep(1)
                    driver.swipe(410, 324, 408, 786)
                    time.sleep(20)
                    print("data Salud actualizada")

                    # Verificar si se sumoron los puntos de recompensa al score del usaurio 2 que creo el equipo
                    time.sleep(2)
                    score = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                    text_score = score.text

                    if int(text_score) == int(suma_points_sec):
                        print(
                            f"Suma de recompensa en score es correcta para el usuario 2 total de puntos = {suma_points_sec}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET04_Score con recompensa sumada correctamente para usaurio 2.png")
                    else:
                        print("Error la suma no es la correcta")
                        print(
                            f"Suma de recompensa en score es INCORRECTA Valor Esperado = {suma_points_sec} Valor obtenido = {text_score}")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            f"\\RET04_Score con recompensa sumada correctamente para usaurio 2 {date2}.png")

                    # CERRAR SESION E INICIAR CON USUARIO QUE SE INSCRIBIO D FORMA INDIVIDUAL y confirmar si se creo un nuevo equpo
                    menu_burger = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    menu_burger.click()
                    time.sleep(1)
                    for i in range(1, 3):
                        driver.swipe(286, 1360, 267, 636)
                    log_out = driver.find_element(By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                    log_out.click()
                    time.sleep(2)

                    LoginAndRegisterFlowsUtils.LoginUser(email.value, passw.value, driver)
                    # Verificar si se creo el equipo con usuarios random donde el admin seria el usaurio en waiting room

                    # ACCEDR A MI DIA
                    my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    my_day.click()
                    time.sleep(3)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(5)
                    # VER DETALLE DL RETO
                    in_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[8]")
                    in_challenge.click()
                    time.sleep(2)
                    see_teams = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.widget.TextView")
                    see_teams.click()
                    time.sleep(2)
                    # COMPROBAR INTEGRANTES DEL SEGUNDO EQUIPO
                    admin_team_random = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[1]").text
                    user2_team_random = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.TextView[1]").text
                    user3_team_random = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView[1]").text
                    user4_team_random = driver.find_element(By.XPATH,
                                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup/android.widget.TextView[1]").text
                    list_random = [user2_team_random, user3_team_random, user4_team_random]

                    if str(name_waitingroom.value) == str(admin_team_random):
                        print("Se asigno al usaurio del waiting como admin correctamente")
                    else:
                        print("No se agrego al usaurio del waiting room como admin de equipo")
                        try:
                            print(f"El usaurio {name_waitingroom.value} en waiting room se agrego "
                                  f"correctamente como admin de equipo")
                            if str(name_secund) not in list_random:
                                if str(name_3) not in list_random:
                                    if str(name_4) not in list_random:
                                        if str(name_5) not in list_random:
                                            print("El equipo secundario se creo correctamente donde no se "
                                                  "repitieron otros usaurio de otros retos")
                                            driver.get_screenshot_as_file(
                                                f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                                "\\RET04_ Se creo el segundo equipo correctamente sin repetir usuarios")
                                            wb.close()
                                            # Eliminar el dato del registro de usuario del excel
                                            # para descartar dicho usuario ya creados
                                            wb = xl.load_workbook(filesheet)
                                            wd = wb.active
                                            ws = wb.worksheets[0]

                                            receives = f"{PathAndVariablesUtil.db_path()}/recursos/UsuariosBeneficios.xlsx"

                                            wb3 = xl.load_workbook(receives)
                                            ws3 = wb3.active
                                            for e in range(1, 3):
                                                ws3.insert_rows(1)
                                                for i in range(1, 2):
                                                    for j in range(1, 11):
                                                        c = ws.cell(row=i, column=j)

                                                        ws3.cell(row=i, column=j).value = c.value

                                                wb3.save(str(receives))

                                                wd.delete_rows(1)  # para la fila 1
                                                wb.save(filesheet)
                        except:
                            print("Error el equipo no es el correcto se repitieron usuarios")
                    # Regresar al home

                    out_challenge = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                    out_challenge.click()
                    time.sleep(1)
                    go_home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                    go_home.click()
                    # FINALIZAR RETO Y COMPROBAR QUE YA NO SE MUESTRE EN PANTALLA
                    # Modificar fecha de inicio para finalizar el reto
                    url3 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date3 = datetime.today()

                    new_date3 = date3 + timedelta(days=-3)
                    dateinit3 = new_date3.strftime('%Y-%m-%d')
                    newtime3 = date3.strftime('%H:%M:%S')
                    payload_start_finish = json.dumps({
                        "start_date": f"{dateinit3}T{newtime3}Z"
                    })
                    response3 = requests.put(url3, headers=headers, data=payload_start_finish)
                    # Modificar fecha de finalizacion
                    url4 = f"https://apiv2desarrollo.habits.ai/api/challenge/{idchallenge}"
                    date4 = datetime.today()
                    new_date4 = date4 + timedelta(days=-3)
                    dateinit4 = new_date4.strftime('%Y-%m-%d')
                    newtime4 = date2.strftime('%H:%M:%S')
                    payload_end_finish = json.dumps({
                        "end_date": f"{dateinit4}T{newtime4}Z"
                    })
                    response4 = requests.put(url4, headers=headers, data=payload_end_finish)
                    print("Se cambio la fecha del reto")
                    print(response4.status_code)

                    time.sleep(1)
                    go_challenge = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    go_challenge.click()
                    time.sleep(3)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(3)
                    challenge_finish = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.widget.TextView[1]"
                    find_challenge = driver.find_element(By.XPATH, str(challenge_finish))
                    text_challen_finish = find_challenge.text
                    if text_challen_finish != text_title:
                        print("Reto se finalizo correctamente")
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET04_Reto Finalizado exitosamente.png")
                    else:
                        print('Error no se finalizo el reto')
            else:
                print("No es un reto Grupal")

        except AssertionError:
            print("No hay retos Grupales")

    def test_RET05(self):  # Reto paralelo
        driver = self.driver

        filesheet = f"{PathAndVariablesUtil.db_path()}/recursos/usuariosListos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        email = datos['B1']
        passw = datos['C1']
        email_two = datos['B2']
        passw_two = datos['C2']

        LoginAndRegisterFlowsUtils.ConfigEnvAndLoginUser(email.value, passw.value, driver)

        # Alerta de calificacion de aplicacion
        try:
            skip = WebDriverWait(driver, 8). \
                until(
                EC.presence_of_element_located((By.XPATH, "//android.widget.TextView[contains(@text, 'Omitir')]")))
            # Omitir calificacion
            skip.click()
            print("Se Omitio la modal de calificar aplicacion en la tienda")
        except:
            print("No hay notificacion de calificacion de aplicacion")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located
                                                  ((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        assert element.is_displayed(), "No se inicio sesion de forma correcta, nos e muestran los elementos del navbar"
        print("Se inicio sesion correctamente")
        enter_element = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
        # Guardar puntos del score del usuario A
        points_user_A = driver.find_element(By.XPATH,
                                            "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
        valuepoint_userA = points_user_A.text

        # Acceder a MI DIA
        enter_element.click()
        time.sleep(5)
        driver.swipe(328, 1034, 328, 823)
        time.sleep(6)

        # VERIFICAR SI LOS RETOS ESTA DISPONIBLES

        try:
            tipo = driver.find_element(By.XPATH,
                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
            print(tipo.text)
            if tipo.text == "Reto paralelo":
                # Guardar puntos de recompensa del reto paralelo
                point_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView[4]").text
                # Entrar a reto paralelo
                enter_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                enter_challe.click()
                time.sleep(5)
                # Seleccionar un tipo de reto para el usaurio A
                first_challe = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]")
                first_challe.click()
                time.sleep(5)
                # Guardar el tipo de reto a inscribir, sueño, pasos o cardio
                type_chall = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.widget.TextView[1]").text
                text_type_chall = type_chall.split()
                # Guardar puntos ojetivos para cumplir meta
                obj_challe = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[2]/android.widget.TextView[1]").text
                sep_obj_chall = obj_challe.split()
                point_obje_chall = sep_obj_chall[0]

                # Inscribirse en el reto
                unirme_al_reto = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                unirme_al_reto.click()
                aceptar = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                element0 = WebDriverWait(driver, 10). \
                    until(EC.presence_of_element_located((By.XPATH, str(aceptar))))
                acep_inscription = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                acep_inscription.click()
                time.sleep(2)
                see_user_in_chall = driver.find_element(By.XPATH,
                                                        "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                see_user_in_chall.click()
                time.sleep(2)
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                              "\\RET05_Usuario registrado en reto.png")
                back_to_home = driver.find_element(By.XPATH,
                                                   "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                back_to_home.click()
                # Obtener token
                url_auth = "https://apiv2desarrollo.habits.ai/login"
                payload_auth = {
                    "mail": "lenyn@habits.ai",
                    "pass": "123456789"
                }
                response_auth = requests.post(url_auth, data=payload_auth)
                response_auth_json = response_auth.json()  # Obtener un json de la respuesta
                token = response_auth_json[
                    'token']  # Hacer un arreglo que guarde el valor de la coincidencia con el texto token
                bearer = f'Bearer {token}'
                headers = {'Content-Type': 'application/json', 'Authorization': f'{bearer}'}
                # Agregar data de salud para usuario A
                url_user_id = f"https://apiv2desarrollo.habits.ai/api/user/?where[mail]={email.value}"
                response_user_id = requests.get(url_user_id, headers=headers)
                response_user_id_json = response_user_id.json()
                data_user = response_user_id_json['data']
                id_user = data_user[0]['_id']
                print(id_user)
                cardioid = "5f63a3676070b30de1549bf3"
                pasos = "603ededd3522a74b4107ffb4"
                sleep = "5f87468222b4c921c00b5286"
                date2 = datetime.today()
                dateinit2 = date2.strftime('%Y-%m-%d')
                time_data = date2 + timedelta(minutes=-2)
                newtime2 = time_data.strftime('%H:%M:%S')
                print(text_type_chall[2])
                if text_type_chall[2] == "Cardio":

                    url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                    cardio_point = int(point_obje_chall)
                    print(cardio_point)
                    payload_heal = json.dumps({
                        "date": f"{dateinit2} {newtime2}",
                        "kind": f"{cardioid}",
                        "value": f"{cardio_point}",
                        "user": f"{id_user}",
                        "timeZone": "America/Caracas"
                    })
                    response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                    if response_data_salud.status_code == "200":
                        print("Data salud cargada exitosamente")
                    else:
                        print(response_data_salud.text)
                        print(response_data_salud.status_code)
                elif text_type_chall[2] == "Sueño":

                    text_poin_obj = str(point_obje_chall)
                    sep_poin_obj = text_poin_obj.split(sep=":")
                    value_point_obj = int(sep_poin_obj[0])
                    new_point_obj = value_point_obj * 3600
                    sleep_point = int(new_point_obj)
                    print(sleep_point)
                    url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                    payload_heal = json.dumps({
                        "date": f"{dateinit2} {newtime2}",
                        "kind": f"{sleep}",
                        "value": f"{sleep_point}",
                        "user": f"{id_user}",
                        "timeZone": "America/Caracas"
                    })
                    response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                    if response_data_salud.status_code == "200":
                        print("Data salud cargada exitosamente")
                    else:
                        print(response_data_salud.text)
                        print(response_data_salud.status_code)
                else:
                    url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                    step_point = int(point_obje_chall)
                    print(step_point)
                    payload_heal = json.dumps({
                        "date": f"{dateinit2} {newtime2}",
                        "kind": f"{pasos}",
                        "value": f"{step_point}",
                        "user": f"{id_user}",
                        "timeZone": "America/Caracas"
                    })
                    response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)

                    if response_data_salud.status_code == "200":
                        print("Data salud cargada exitosamente")
                    else:
                        print(response_data_salud.text)
                        print(response_data_salud.status_code)
                    time.sleep(0.5)

                # Sincronizar data salud del usaurio A
                back_dashboard = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                back_dashboard.click()
                time.sleep(2)
                home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                home.click()
                time.sleep(4)
                driver.swipe(453, 1055, 446, 755)

                sync_data = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                sync_data.click()

                time.sleep(1)
                term_sync = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                term_sync.click()
                time.sleep(2)
                sync_cellphone = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]")
                sync_cellphone.click()
                time.sleep(2)
                try:
                    init_sync = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                    init_sync.click()
                    time.sleep(2)
                except:
                    print("terminos aceptados para sincronizar")
                try:
                    next_step = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                    next_step.click()
                    time.sleep(2)
                except:
                    print("terminos aceptados para sincronizar")
                element3 = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, "com.google.android.gms:id/account_picker_container")))
                # SELECCIONAR USUARIO PARA SINCRONIZAR DATA

                user_data = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout")
                user_data.click()
                data_salud = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup"
                element4 = WebDriverWait(driver, 50).until(
                    EC.presence_of_element_located((By.XPATH, str(data_salud))))
                driver.swipe(342, 1079, 347, 811)

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                              "\\RET05_Data_salud_sincronizada.png")
                print("data sincronizada")
                # Confirmar si se sumaron los puntos correctamente en el score del usuario A
                time.sleep(2)
                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                              "\\RET05_Notifiacion de reto cumplido.png")
                time.sleep(3)

                driver.swipe(436, 338, 415, 1213)
                into_score = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]")
                into_score.click()
                time.sleep(2)
                challen_finish = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[3]").text
                challen_point_finish = driver.find_element(By.XPATH,
                                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[2]").text
                if "¡Has terminado el desafío!" in challen_finish and f"Ganaste {point_challe} puntos" in challen_point_finish:
                    print("Se obtuvieron los puntos correctamente por cumplir el reto en paralelo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET05 Puntuacion recibida correctamente por completar el reto.png")
                    time.sleep(2)
                    # abrir menu lateral
                    el13 = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView")
                    el13.click()
                    time.sleep(2)
                    # Ir al home
                    el14 = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                    el14.click()
                    time.sleep(1)
                    # Confirmar si se sumo el score correctamente al usuario A

                    old_score_A = int(valuepoint_userA) + int(point_challe)
                    new_score_A = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]").text
                    if old_score_A == int(new_score_A):
                        print("Se sumo correctamente la puntuacion en el score al usuario A")
                    else:
                        print("No se sumaron los puntos en el score del usuario")

                    # Confirmar el progreso del reto para el usuario A
                    my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    my_day.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(6)
                    # Entrar a reto paralelo
                    enter_challe.click()
                    time.sleep(3)
                    # ingresar un tipo de reto para del usaurio A
                    first_challe.click()
                    time.sleep(3)
                    # Entrar al reto
                    see_chall_userA = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    see_chall_userA.click()
                    # VEr progreso del reto
                    up = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element4 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(up))))

                    see_progress = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                    see_progress.click()

                    progree_user = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.TextView[2]").text
                    objetive_challenge = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.TextView[3]").text
                    if progree_user == objetive_challenge:
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET05 Progreso de reto cumplido usuario A.png")
                        print("Progreso del reto cumplido para usaurio A")
                    else:
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET05 Error no se cumplio el reto del usuario A.png")
                        print("No se cumplio el progreso del reto del usaurio A")
                else:
                    print("Error en asignar puntuacion por completar el reto")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET05 Error no se obtuvo puntuacion por cumplir el reto.png")
                # Regresar a Reto paralelo
                el20 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                el20.click()
                time.sleep(1)
                # Regresar a mi dia
                el21 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                el21.click()
                time.sleep(1)
                # abrir menu lateral
                el22 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]")
                el22.click()
                # cerrar sesion
                for i in range(1, 3):
                    driver.swipe(260, 1000, 260, 514)

                el23 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                el23.click()
                time.sleep(6)
                # Iniciar sesion con usaurio B
                LoginAndRegisterFlowsUtils.LoginUser(email_two.value, passw_two.value, driver)
                # ir al reto en mi dia

                el29 = WebDriverWait(driver, 50).until(
                    EC.presence_of_element_located((By.XPATH, PathAndVariablesUtil.challenge_navbar())))
                # Guardar puntos del score del usuario B
                points_user_B = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]")
                valuepoint_userB = points_user_B.text
                time.sleep(1)

                el29.click()
                time.sleep(5)
                driver.swipe(328, 1034, 328, 823)
                time.sleep(6)
                # Entrar al reto
                el30 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.HorizontalScrollView/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                el30.click()
                time.sleep(5)
                # Seleccionar el reto numero 2 para usaurio B
                secund_challenge = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]")
                secund_challenge.click()
                time.sleep(5)
                challen_two = driver.find_element(By.XPATH,
                                                  "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.widget.TextView[1]").text
                sep_chall_two = challen_two.split()
                pts_obj_secund = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[2]/android.widget.TextView[1]").text
                sep_pts_obj = pts_obj_secund.split()
                point_secund_chall = sep_pts_obj[0]
                # Inscribirse en reto 2 para usaurio B
                el32 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                el32.click()
                time.sleep(1)
                acep_ins_two = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                element4 = WebDriverWait(driver, 50).until(
                    EC.presence_of_element_located((By.XPATH, str(acep_ins_two))))
                modal_acep = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                modal_acep.click()
                time.sleep(1)
                # Regresar a home
                el34 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                el34.click()
                time.sleep(3)
                el35 = driver.find_element(By.XPATH,
                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                el35.click()
                time.sleep(5)
                # Inyectar daTA SALUD para usaurio B
                # Obtener token
                url_auth = "https://apiv2desarrollo.habits.ai/login"
                payload_auth = {
                    "mail": "lenyn@habits.ai",
                    "pass": "123456789"
                }
                response_auth = requests.post(url_auth, data=payload_auth)
                response_auth_json = response_auth.json()  # Obtener un json de la respuesta
                token = response_auth_json[
                    'token']  # Hacer un arreglo que guarde el valor de la coincidencia con el texto token
                bearer = f'Bearer {token}'
                headers = {'Content-Type': 'application/json', 'Authorization': f'{bearer}'}
                # Agregar data de salud para usuario A
                url_user_id = f"https://apiv2desarrollo.habits.ai/api/user/?where[mail]={email_two.value}"
                response_user_id = requests.get(url_user_id, headers=headers)
                response_user_id_json = response_user_id.json()
                data_user = response_user_id_json['data']
                id_user = data_user[0]['_id']
                print(id_user)
                cardioid = "5f63a3676070b30de1549bf3"
                pasos = "603ededd3522a74b4107ffb4"
                sleep = "5f87468222b4c921c00b5286"
                date2 = datetime.today()
                dateinit2 = date2.strftime('%Y-%m-%d')
                time_data = date2 + timedelta(minutes=-2)
                newtime2 = time_data.strftime('%H:%M:%S')
                if sep_chall_two[2] == "Cardio":

                    url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                    cardio_point = int(point_obje_chall)
                    payload_heal = json.dumps({
                        "date": f"{dateinit2} {newtime2}",
                        "kind": f"{cardioid}",
                        "value": f"{cardio_point}",
                        "user": f"{id_user}",
                        "timeZone": "America/Caracas"
                    })
                    response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                    status = response_data_salud.status_code
                    if status == "200":
                        print("Data salud cargada exitosamente")
                elif sep_chall_two[2] == "Sueño":

                    text_poin_obj = str(point_obje_chall)
                    sep_poin_obj = text_poin_obj.split(sep=":")
                    value_point_obj = int(sep_poin_obj[0])
                    new_point_obj = value_point_obj * 3600
                    sleep_point = int(new_point_obj)
                    url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                    payload_heal = json.dumps({
                        "date": f"{dateinit2} {newtime2}",
                        "kind": f"{sleep}",
                        "value": f"{sleep_point}",
                        "user": f"{id_user}",
                        "timeZone": "America/Caracas"
                    })
                    response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                    status = response_data_salud.status_code
                    if status == "200":
                        print("Data salud cargada exitosamente")
                else:
                    url_data_heal = "https://apiv2desarrollo.habits.ai/api/health/addManualData/"
                    step_point = int(point_obje_chall)
                    payload_heal = json.dumps({
                        "date": f"{dateinit2} {newtime2}",
                        "kind": f"{pasos}",
                        "value": f"{step_point}",
                        "user": f"{id_user}",
                        "timeZone": "America/Caracas"
                    })
                    response_data_salud = requests.post(url_data_heal, headers=headers, data=payload_heal)
                    status = response_data_salud.status_code
                    if status == "200":
                        print("Data salud cargada exitosamente")
                    time.sleep(0.5)

                # Sincronizar data salud del usaurio B
                back_dashboard = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView")
                back_dashboard.click()
                time.sleep(2)
                home = driver.find_element(By.XPATH, PathAndVariablesUtil.home_navbar())
                home.click()
                time.sleep(2)
                driver.swipe(453, 1055, 446, 755)

                sync_data = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.TextView")
                sync_data.click()

                time.sleep(1)
                term_sync = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                term_sync.click()
                time.sleep(2)
                sync_cellphone = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]")
                sync_cellphone.click()
                time.sleep(2)
                try:
                    init_sync = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                    init_sync.click()
                    time.sleep(2)
                except:
                    print("terminos aceptados para sincronizar")
                try:
                    next_step = driver.find_element(By.XPATH,
                                                    "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]")
                    next_step.click()
                    time.sleep(2)
                except:
                    print("terminos aceptados para sincronizar")
                element3 = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, "com.google.android.gms:id/account_picker_container")))
                # SELECCIONAR USUARIO PARA SINCRONIZAR DATA

                user_data = driver.find_element(By.XPATH,
                                                "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout")
                user_data.click()
                data_salud = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup[3]/android.view.ViewGroup[1]/android.view.ViewGroup"
                element4 = WebDriverWait(driver, 50).until(
                    EC.presence_of_element_located((By.XPATH, str(data_salud))))
                driver.swipe(342, 1079, 347, 811)

                driver.get_screenshot_as_file(f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                                              "\\RET05_Data_salud_sincronizada usaurio B.png")
                print("data sincronizada")
                # Confirmar si se sumaron los puntos correctamente en el score del usuario A

                driver.swipe(436, 338, 415, 1213)
                into_score = driver.find_element(By.XPATH,
                                                 "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]")
                into_score.click()
                time.sleep(2)
                challen_finish = driver.find_element(By.XPATH,
                                                     "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[3]").text
                challen_point_finish = driver.find_element(By.XPATH,
                                                           "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[3]/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[2]/android.widget.TextView[2]").text
                if "¡Has terminado el desafío!" in challen_finish and f"Ganaste {point_challe} puntos" in challen_point_finish:
                    print("Se obtuvieron los puntos correctamente por cumplir el reto en paralelo")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET05 Puntuacion recibida correctamente por completar el reto para usaurio B.png")
                    # abrir menu lateral
                    el13 = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.widget.TextView")
                    el13.click()
                    # Ir al home
                    el14 = driver.find_element(By.XPATH,
                                               "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[2]/android.view.ViewGroup/android.view.ViewGroup")
                    el14.click()
                    # Confirmar si se sumo el score correctamente al usuario A

                    old_score_B = int(valuepoint_userB) + int(point_secund_chall)
                    new_score_B = driver.find_element(By.XPATH,
                                                      "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[2]/android.widget.ScrollView/android.view.ViewGroup/android.view.ViewGroup[1]/android.widget.TextView[2]").text
                    if old_score_B == int(new_score_B):
                        print("Se sumo correctamente la puntuacion en el score al usuario B")
                    else:
                        print("No se sumaron los puntos en el score del usuario B")

                    # Confirmar el progreso del reto para el usuario A
                    my_day = driver.find_element(By.XPATH, PathAndVariablesUtil.challenge_navbar())
                    my_day.click()
                    time.sleep(5)
                    driver.swipe(328, 1034, 328, 823)
                    time.sleep(6)
                    # Entrar a reto paralelo
                    enter_challe.click()
                    time.sleep(2)
                    # ingresar un tipo de reto para del usaurio A
                    first_challe.click()
                    # Entrar al reto
                    see_chall_userA = driver.find_element(By.XPATH,
                                                          "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup[4]/android.view.ViewGroup[3]/android.view.ViewGroup/android.widget.TextView")
                    see_chall_userA.click()
                    # VEr progreso del reto
                    up = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup"
                    element4 = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, str(up))))

                    see_progress = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup")
                    see_progress.click()

                    progree_user = driver.find_element(By.XPATH,
                                                       "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.TextView[2]").text
                    objetive_challenge = driver.find_element(By.XPATH,
                                                             "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.FrameLayout/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.widget.ScrollView/android.view.ViewGroup/android.widget.TextView[3]").text
                    if progree_user == objetive_challenge:
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET05 Progreso de reto cumplido usuario B.png")
                        print("Progreso del reto cumplido para usaurio B")
                        wb.close()
                        # Eliminar el dato del registro de usuario del excel para descartar dicho usuario ya creados
                        wb = xl.load_workbook(filesheet)
                        wd = wb.active
                        ws = wb.worksheets[0]

                        receives = f"{PathAndVariablesUtil.db_path()}/recursos/UsuariosBeneficios.xlsx"

                        wb3 = xl.load_workbook(receives)
                        ws3 = wb3.active

                        ws3.insert_rows(1)
                        for i in range(1, 2):
                            for j in range(1, 11):
                                c = ws.cell(row=i, column=j)

                                ws3.cell(row=i, column=j).value = c.value

                        wb3.save(str(receives))

                        wd.delete_rows(1)  # para la fila 1
                        wb.save(filesheet)
                    else:
                        driver.get_screenshot_as_file(
                            f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                            "\\RET05 Error no se cumplio el reto del usuario B.png")
                        print("No se cumplio el progreso del reto del usaurio B")
                else:
                    print("Error en asignar puntuacion por completar el reto")
                    driver.get_screenshot_as_file(
                        f"{PathAndVariablesUtil.evidence_path()}\\recursos\\screenshots\\Retos"
                        "\\RET05 Error no se obtuvo puntuacion por cumplir el reto del usuario B.png")

            else:
                print("El reto no es paralelo")
        except AssertionError:
            print("Error en el reto no existe")

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    suite = unittest.TestLoader().loadTestsFromTestCase(Retos)
    unittest.TextTestRunner(verbosity=2).run(suite)
